"""Project commercial API (QS): BOQ. Progress claims + P&L follow in later
slices. Commercial data is contract-sensitive, so access is gated to those who
may see the contract value (HO roles incl. QS, and the assigned PM)."""
from rest_framework import serializers
from rest_framework.decorators import (api_view, parser_classes,
                                       permission_classes)
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from . import commercial
from .models import (BoqItem, ClientReceipt, ProgressClaim, Project,
                     Variation, VariationItem)
from .views_projects import PROJECT_EDIT_ROLES, _can_view_value


def _get_project(request, pid):
    try:
        p = Project.objects.select_related("site").get(pk=pid)
    except Project.DoesNotExist:
        return None, Response({"detail": "Not found."}, status=404)
    if not _can_view_value(request.user, p):
        return None, Response({"detail": "Not permitted."}, status=403)
    return p, None


def _require_editor(request):
    if request.user.role not in PROJECT_EDIT_ROLES:
        return Response({"detail": "Only the QS / PM edits the BOQ."},
                        status=403)
    return None


class BoqItemSerializer(serializers.ModelSerializer):
    amount = serializers.DecimalField(max_digits=18, decimal_places=3,
                                      read_only=True)
    amount_supply = serializers.DecimalField(max_digits=18, decimal_places=3,
                                             read_only=True)
    amount_install = serializers.DecimalField(max_digits=18, decimal_places=3,
                                              read_only=True)
    rate_total = serializers.DecimalField(max_digits=16, decimal_places=3,
                                          read_only=True)

    class Meta:
        model = BoqItem
        fields = ["id", "sort_order", "section", "item_code", "description",
                  "unit", "qty", "rate_supply", "rate_install", "rate_total",
                  "is_heading", "is_discount", "amount", "amount_supply",
                  "amount_install"]


def _boq_payload(project):
    boq = getattr(project, "boq", None)
    if boq is None:
        return {"exists": False, "currency": "USD", "is_locked": False,
                "split_rates": False, "mode": "CONVENTIONAL", "total": 0,
                "total_supply": 0, "total_install": 0, "contract_value": 0,
                "items": [], "categories": []}
    data = {"exists": True, "currency": boq.currency,
            "is_locked": boq.is_locked, "split_rates": boq.split_rates,
            "mode": boq.mode, "total": boq.total,
            "total_supply": boq.total_supply, "total_install": boq.total_install,
            "contract_value": boq.contract_value,
            "items": BoqItemSerializer(boq.items.all(), many=True).data,
            "categories": []}
    if boq.mode == boq.Mode.UNIT:
        data["categories"] = [{
            "id": c.id, "ref": c.ref, "name": c.name, "unit": c.unit,
            "qty": c.qty, "is_lump": c.is_lump,
            "per_unit_total": c.per_unit_total, "line_total": c.line_total,
            # Split-certified categories (material vs workmanship claimed
            # independently) carry the two per-unit contract rates.
            "unit_amount_supply": c.unit_amount_supply,
            "unit_amount_install": c.unit_amount_install,
            "is_split": c.has_split_rates,
            # The captured works' sum — shown against the contract rate so the
            # QS can see where a bill's detail doesn't reconcile.
            "items_total": c.items_total,
            # The works that build up the per-unit rate (empty for lump bills);
            # material and labour rates kept separate, like a conventional BOQ.
            "items": [{"id": i.id, "description": i.description,
                       "unit": i.unit, "qty": i.qty,
                       "rate_material": i.rate_supply, "rate_labour": i.rate_install,
                       "rate": i.rate_total, "amount": i.amount}
                      for i in c.items.all()],
        } for c in boq.categories.all()]
    return data


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def boq_detail(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    return Response(_boq_payload(p))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def boq_save(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    boq, msg = commercial.set_boq_items(p, request.data.get("rows") or [],
                                        request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_boq_payload(p))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def boq_lock(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_boq_lock(p, request.data.get("locked", True),
                                     request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_boq_payload(p))


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def boq_delete(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.delete_boq(p, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    # Re-fetch: the deleted BOQ is still cached on the reverse relation.
    p, _ = _get_project(request, pid)
    return Response(_boq_payload(p))


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def boq_import(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    upload = request.FILES.get("file")
    if not upload:
        return Response({"detail": "Attach the filled BOQ Excel (.xlsx)."},
                        status=400)
    from openpyxl import load_workbook
    try:
        wb = load_workbook(upload, read_only=True, data_only=True)
    except Exception:
        return Response({"detail": "Could not read that file — save it as "
                         ".xlsx and try again."}, status=400)
    ws = wb["BOQ"] if "BOQ" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return Response({"detail": "The sheet is empty."}, status=400)
    keys = [commercial.normalise_header(h) for h in header]
    if "description" not in keys:
        return Response({"detail": "Need at least a Description column."},
                        status=400)
    rows = []
    for raw in rows_iter:
        if raw is None or all(c in (None, "") for c in raw):
            continue
        rows.append({k: v for k, v in zip(keys, raw) if k})
    if not rows:
        return Response({"detail": "No rows found below the header."},
                        status=400)
    boq, msg = commercial.import_boq_rows(p, rows, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_boq_payload(p))


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def boq_template(request, pid):
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font
    p, err = _get_project(request, pid)
    if err:
        return err
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    # Supply (Material) + Install (Labour) columns; leave Install blank for a
    # combined-rate contract.
    headers = ["Section", "Code", "Description", "Unit", "Qty",
               "Material", "Labour"]
    ws.append(headers)
    for i, w in enumerate([22, 10, 46, 8, 12, 12, 12], start=1):
        ws.cell(row=1, column=i).font = Font(bold=True)
        ws.column_dimensions[chr(64 + i)].width = w
    ws.append(["Bill 1 — Substructure", "", "", "", "", "", ""])
    ws.append(["", "1.1", "Excavate for foundations", "m3", "120", "5.00",
               "3.50"])
    ws.append(["", "1.2", "Mass concrete blinding", "m3", "35", "80.00",
               "15.00"])
    ws.freeze_panes = "A2"
    resp = HttpResponse(content_type="application/vnd.openxmlformats-"
                        "officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="boq-template.xlsx"'
    wb.save(resp)
    return resp


# ---- BOQ capture from PDF / Excel (extract → review → commit) ------------

def _get_import(request, pk):
    from .models import BoqImport
    try:
        imp = BoqImport.objects.select_related("project__site").get(pk=pk)
    except BoqImport.DoesNotExist:
        return None, Response({"detail": "Not found."}, status=404)
    if not _can_view_value(request.user, imp.project):
        return None, Response({"detail": "Not permitted."}, status=403)
    return imp, None


@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def boq_import_extract(request, pid):
    """Upload a client BOQ (PDF/Excel) and extract it into a review draft."""
    from . import boq_extract
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    upload = request.FILES.get("file")
    if not upload:
        return Response({"detail": "Attach the BOQ PDF or Excel."}, status=400)
    try:
        imp, msg = boq_extract.run_import(p, upload, request.user)
    except boq_extract.ExtractionError as e:
        return Response({"detail": str(e)}, status=400)
    except Exception as e:               # surface the reason, never a bare 500
        import logging
        logging.getLogger("boq").exception("BOQ capture failed")
        return Response({"detail": f"Capture failed: {e}"}, status=400)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(boq_extract.import_payload(imp), status=201)


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([IsAuthenticated])
def boq_capture_unit(request, pid):
    """Capture a unit-based BOQ (Excel or PDF) → review categories (no commit).
    Separate from the conventional import; the conventional path is untouched."""
    from . import boq_unit_extract as ue
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    upload = request.FILES.get("file")
    if not upload:
        return Response({"detail": "Attach the unit BOQ (Excel or PDF)."},
                        status=400)
    try:
        cats, gst, msg = ue.run_capture(upload)
    except ue.ExtractionError as e:
        return Response({"detail": str(e)}, status=400)
    except Exception as e:               # surface the reason, never a bare 500
        import logging
        logging.getLogger("boq").exception("BOQ unit capture failed")
        return Response({"detail": f"Capture failed: {e}"}, status=400)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response({"categories": cats, "gst_percent": gst,
                     "count": len(cats)})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def boq_commit_unit(request, pid):
    """Commit reviewed unit categories into the project's BOQ (UNIT mode)."""
    from . import boq_unit_extract as ue
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    boq, msg = ue.commit(p, request.data.get("categories") or [], request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response({"ok": True, "mode": boq.mode,
                     "categories": boq.categories.count(),
                     "contract_value": str(boq.contract_value)})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def boq_category_items(request, pid, cat_id):
    """Set a unit category's detail line items (the works that build its per-unit
    rate). The per-unit total then derives from them."""
    from . import boq_unit_extract as ue
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = ue.set_category_items(p, cat_id, request.data.get("rows") or [],
                                   request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_boq_payload(p))


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def boq_import_latest(request, pid):
    """The most recent open (un-committed) import draft for the project."""
    from . import boq_extract
    from .models import BoqImport
    p, err = _get_project(request, pid)
    if err:
        return err
    imp = (BoqImport.objects.filter(project=p, status="DRAFT")
           .order_by("-created_at").first())
    return Response(boq_extract.import_payload(imp) if imp else None)


@api_view(["PUT", "DELETE"])
@permission_classes([IsAuthenticated])
def boq_import_detail(request, pk):
    from . import boq_extract
    from .models import BoqImport
    imp, err = _get_import(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    if request.method == "DELETE":
        imp.delete()
        return Response(status=204)
    if imp.status == BoqImport.Status.COMMITTED:
        return Response({"detail": "This import is already loaded."}, status=400)
    rows = boq_extract.normalise_rows(request.data.get("rows") or [])
    imp.rows = rows
    rec = boq_extract.reconcile(
        rows, [imp.meta.get("printed_total")] if imp.meta.get("printed_total")
        else [])
    imp.meta = {**imp.meta, **rec}
    imp.save(update_fields=["rows", "meta", "updated_at"])
    return Response(boq_extract.import_payload(imp))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def boq_import_commit(request, pk):
    from . import boq_extract
    imp, err = _get_import(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    boq, msg = boq_extract.commit(imp, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_boq_payload(imp.project))


# ---- Variations (VOs) ---------------------------------------------------

class VariationItemSerializer(serializers.ModelSerializer):
    amount = serializers.DecimalField(max_digits=18, decimal_places=3,
                                      read_only=True)
    amount_supply = serializers.DecimalField(max_digits=18, decimal_places=3,
                                             read_only=True)
    amount_install = serializers.DecimalField(max_digits=18, decimal_places=3,
                                              read_only=True)
    rate_total = serializers.DecimalField(max_digits=16, decimal_places=3,
                                          read_only=True)

    class Meta:
        model = VariationItem
        fields = ["id", "sort_order", "section", "item_code", "description",
                  "unit", "qty", "rate_supply", "rate_install", "rate_total",
                  "is_heading", "amount", "amount_supply", "amount_install"]


class VariationSerializer(serializers.ModelSerializer):
    gross = serializers.DecimalField(max_digits=18, decimal_places=2,
                                     read_only=True)
    signed_total = serializers.DecimalField(max_digits=18, decimal_places=2,
                                            read_only=True)
    items = VariationItemSerializer(many=True, read_only=True)

    class Meta:
        model = Variation
        fields = ["id", "seq", "ref", "title", "kind", "status", "ref_date",
                  "gross", "signed_total", "items"]


def _variations_payload(project):
    vs = project.variations.prefetch_related("items").all()
    return {
        "currency": (getattr(project, "boq", None).currency
                     if getattr(project, "boq", None) else "USD"),
        "contract": {k: v for k, v in commercial.contract_summary(
            project).items()},
        "variations": VariationSerializer(vs, many=True).data,
    }


def _get_variation(request, pk):
    try:
        v = Variation.objects.select_related("project__site").get(pk=pk)
    except Variation.DoesNotExist:
        return None, Response({"detail": "Not found."}, status=404)
    if not _can_view_value(request.user, v.project):
        return None, Response({"detail": "Not permitted."}, status=403)
    return v, None


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def variation_list(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    return Response(_variations_payload(p))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def variation_create(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    commercial.create_variation(p, request.data, request.user)
    return Response(_variations_payload(p), status=201)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def variation_items(request, pk):
    v, err = _get_variation(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_variation_items(v, request.data.get("rows") or [],
                                            request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_variations_payload(v.project))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def variation_meta(request, pk):
    v, err = _get_variation(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_variation_meta(v, request.data, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_variations_payload(v.project))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def variation_status(request, pk):
    v, err = _get_variation(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_variation_status(
        v, request.data.get("status"), request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_variations_payload(v.project))


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def variation_delete(request, pk):
    v, err = _get_variation(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    if v.status != "DRAFT":
        return Response({"detail": "Only a draft variation can be deleted."},
                        status=400)
    project = v.project
    v.delete()
    return Response(_variations_payload(project))


# ---- Progress claims (interim payment applications) ---------------------

def _claim_meta(claim):
    return {
        "id": claim.id, "seq": claim.seq, "ref": claim.ref,
        "claim_type": claim.claim_type, "basis": claim.basis,
        "basis_locked": claim.previous_id is not None,
        "status": claim.status, "work_done_upto": claim.work_done_upto,
        "advance_pct": claim.advance_pct, "recovery_pct": claim.recovery_pct,
        "advance_recovered_override": claim.advance_recovered_override,
        "retention_pct": claim.retention_pct,
        "retention_held_override": claim.retention_held_override,
        "gst_pct": claim.gst_pct,
        "material_on_site": claim.material_on_site,
        "material_off_site": claim.material_off_site,
        "retention_released": claim.retention_released,
        "note": claim.note,
        "previous_ref": claim.previous.ref if claim.previous_id else None,
        "certified_at": claim.certified_at,
        "invoice_no": claim.invoice_no,
    }


def _receipt_json(r):
    return {
        "id": r.id, "amount": r.amount, "currency": r.currency,
        "received_on": r.received_on, "reference": r.reference,
        "note": r.note, "claim_ref": r.claim.ref if r.claim_id else None,
        "claim_id": r.claim_id,
        "recorded_by": (r.recorded_by.full_name if r.recorded_by_id else None),
    }


def _claims_payload(project):
    """The claims register: each claim's header plus its net-due / total from
    the waterfall, the contract summary, the money-in position and receipts."""
    claims = list(project.claims.all())
    rows = []
    for c in claims:
        w = commercial.claim_valuation(c)["waterfall"]
        rows.append({**_claim_meta(c),
                     "k_gross": w["k_gross"], "net_due": w["net_due"],
                     "gst": w["gst"], "total": w["total"]})
    receipts = project.receipts.all()
    return {
        "currency": (getattr(project, "boq", None).currency
                     if getattr(project, "boq", None) else "USD"),
        "contract": {k: v for k, v in commercial.contract_summary(
            project).items()},
        "can_raise": bool(getattr(project, "boq", None) and (
            project.boq.categories.exists() if project.boq.mode == "UNIT"
            else project.boq.items.exists())),
        "claims": rows,
        "revenue": commercial.project_revenue_summary(project),
        "receipts": [_receipt_json(r) for r in receipts],
    }


def _claim_detail(claim):
    val = commercial.claim_valuation(claim)
    return {"claim": _claim_meta(claim), **val}


def _get_claim(request, pk):
    try:
        c = ProgressClaim.objects.select_related(
            "project__site", "previous").get(pk=pk)
    except ProgressClaim.DoesNotExist:
        return None, Response({"detail": "Not found."}, status=404)
    if not _can_view_value(request.user, c.project):
        return None, Response({"detail": "Not permitted."}, status=403)
    return c, None


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def claim_list(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    return Response(_claims_payload(p))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def claim_create(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.create_claim(p, request.data, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_claims_payload(p), status=201)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def claim_detail(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    return Response(_claim_detail(c))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def claim_items(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_claim_items(c, request.data.get("rows") or [],
                                        request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_claim_detail(c))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def claim_meta(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_claim_meta(c, request.data, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_claim_detail(c))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def claim_deductions(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_claim_deductions(
        c, request.data.get("rows") or [], request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_claim_detail(c))


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def deduction_presets(request):
    """The non-inventory back-charge pick-list (QS/PM/Director/Admin maintain)."""
    from .models import DeductionPreset
    if request.method == "POST":
        if request.user.role not in PROJECT_EDIT_ROLES:
            return Response({"detail": "Not permitted."}, status=403)
        name = (request.data.get("name") or "").strip()
        if not name:
            return Response({"detail": "A name is required."}, status=400)
        p, _ = DeductionPreset.objects.get_or_create(name=name)
        return Response({"id": p.id, "name": p.name}, status=201)
    return Response([{"id": p.id, "name": p.name}
                     for p in DeductionPreset.objects.filter(is_active=True)])


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def claim_status(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.set_claim_status(c, request.data.get("status"),
                                         request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    detail = _claim_detail(c)
    # Warn (never block) if a REQUIRED bond/insurance cover isn't issued yet —
    # the client won't process the claim without it (owner 2026-08-03).
    from . import bonds
    gaps = bonds.required_gaps(c.project) if c.status != "DRAFT" else []
    detail["bond_warning"] = (
        "Required cover not yet issued — " + ", ".join(gaps)
        + ". The client may not process this claim until these are in place."
    ) if gaps else None
    return Response(detail)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def claim_delete(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    if c.status != "DRAFT":
        return Response({"detail": "Only a draft claim can be deleted."},
                        status=400)
    # Only the newest claim can be removed — earlier ones anchor the chain.
    if c.project.claims.filter(seq__gt=c.seq).exists():
        return Response({"detail": "Delete the later claim(s) first."},
                        status=400)
    project = c.project
    c.delete()
    return Response(_claims_payload(project))


# ---- Client receipts (money-in, P4) -------------------------------------

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def receipt_create(request, pid):
    p, err = _get_project(request, pid)
    if err:
        return err
    if (bad := _require_editor(request)):
        return bad
    _, msg = commercial.record_client_receipt(p, request.data, request.user)
    if msg:
        return Response({"detail": msg}, status=400)
    return Response(_claims_payload(p), status=201)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def receipt_delete(request, pk):
    try:
        r = ClientReceipt.objects.select_related("project__site").get(pk=pk)
    except ClientReceipt.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if not _can_view_value(request.user, r.project):
        return Response({"detail": "Not permitted."}, status=403)
    if (bad := _require_editor(request)):
        return bad
    project = r.project
    commercial.delete_client_receipt(r, request.user)
    return Response(_claims_payload(project))


# ---- Claim / invoice PDFs (P5) ------------------------------------------

def pdf_bytes(template, context):
    """Render a template to PDF bytes (raises if the engine is unavailable).
    Shared by the download responses and the meeting-minutes email."""
    from django.conf import settings
    from django.template.loader import render_to_string
    from weasyprint import HTML
    html = render_to_string(template, context)
    return HTML(string=html, base_url=str(settings.MEDIA_ROOT)).write_pdf()


def _render_pdf(template, context, filename):
    from django.http import HttpResponse
    try:
        pdf = pdf_bytes(template, context)
    except Exception as e:                       # pragma: no cover - env dep
        return Response({"detail": f"PDF engine unavailable: {e}"}, status=500)
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}.pdf"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def claim_ipa_pdf(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if c.status == "DRAFT":
        return Response({"detail": "Submit the claim before printing the "
                                   "application."}, status=400)
    # A certified application prints as the Interim Payment Certificate (IPC).
    tag = f"{c.ipc_ref}-IPC" if c.is_certified else f"{c.ref}-IPA"
    return _render_pdf("pdf/claim_ipa.html",
                       commercial.claim_pdf_context(c),
                       f"{c.project.code}-{tag}")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def claim_invoice_pdf(request, pk):
    c, err = _get_claim(request, pk)
    if err:
        return err
    if c.status not in ("CERTIFIED", "PAID"):
        return Response({"detail": "A tax invoice can be issued once the "
                                   "claim is certified."}, status=400)
    return _render_pdf("pdf/tax_invoice.html",
                       commercial.invoice_pdf_context(c),
                       f"{c.invoice_no or c.ref}")
