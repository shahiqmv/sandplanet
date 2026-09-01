"""Variation order — QS working copy (xlsx).

The VO PDF is the client's document: fixed, signed, filed. This is the other
thing a QS needs — the same variation as a sheet they can work in, before the
price is one we stand behind (owner 2026-09-01).

So the money here is FORMULAS, not answers. Change a qty or a rate and the
line amount, its section total, the variation total and the resulting contract
sum all move with it. A sheet of baked numbers would only be the PDF again,
with worse typography.

It carries the same figures as `variation_pdf_context`, from the same call, so
the working copy and the issued document can never quietly disagree.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# House palette, matching the PDF letterhead and the procurement export.
_NAVY = "10344F"
_BAND = "1685CC"
_SECTION = "E7F1F9"
_GREY = "F4F9FD"
_WHITE = "FFFFFF"
_AMBER = "8A6D00"

_THIN = Side(style="thin", color="C4D0DA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")

_MONEY = '#,##0.00;[Red]-#,##0.00'
_QTY = '#,##0.00###'


def _money_fmt(currency):
    """Money reads in the contract's currency — the contract decides, project
    by project, and a sheet that says MVR over a USD contract is worse than
    one that says nothing."""
    return f'"{currency} "' + _MONEY


def build_variation_xlsx(v):
    from .commercial import variation_pdf_context
    ctx = variation_pdf_context(v)
    split, ccy = ctx["split"], ctx["currency"]
    wb = Workbook()
    ws = wb.active
    ws.title = v.ref or "Variation"

    # Columns: Code | Description | Unit | Qty | (Material | Labour) | Amount
    cols = ([("Code", 12), ("Description", 52), ("Unit", 9), ("Qty", 11)]
            + ([("Material rate", 14), ("Labour rate", 14)] if split
               else [("Rate", 14)])
            + [("Amount", 16)])
    last = len(cols)
    last_letter = get_column_letter(last)
    amt_letter = last_letter
    qty_letter = get_column_letter(4)
    for i, (_, width) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    r = 1

    def band(text, size=13, fill=_NAVY, color=_WHITE, height=22):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=last)
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=size, color=color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = _LEFT
        ws.row_dimensions[r].height = height
        r += 1

    def field(label, value):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=9,
                                                          color="5A6B78")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last)
        c = ws.cell(row=r, column=2, value=value)
        c.alignment = _LEFT
        r += 1

    band(f"{v.ref} — {(v.title or '').strip() or 'Variation order'}")
    band(f"{ctx['project'].title}  ·  {ctx['project'].code}", size=10,
         fill=_BAND, height=18)
    r += 1

    field("Project", f"{ctx['project'].title} ({ctx['project'].code})")
    field("Employer", (ctx["employer"] or {}).get("name") or "—")
    field("Variation", f"{v.ref}  ·  {ctx['kind_label']}")
    field("Status", ctx["client_status"])
    field("Client instruction", v.ref_date.strftime("%d %b %Y")
          if v.ref_date else "—")
    field("Currency", ccy)
    field("Prepared by", ctx["prepared_by"] or "—")
    r += 1

    # The one thing a reader must not get wrong about this file.
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last)
    note = ws.cell(row=r, column=1, value=(
        "WORKING COPY — figures are live formulas. Edit a qty or a rate and "
        "the totals below follow. This is not the issued variation order; "
        "download the VO PDF for the document that goes to the Employer."))
    note.font = Font(bold=True, size=9, color=_AMBER)
    note.fill = PatternFill("solid", fgColor="FFF7E0")
    note.alignment = _LEFT
    ws.row_dimensions[r].height = 26
    r += 2

    # ---- priced items -------------------------------------------------
    head = r
    for i, (title, _) in enumerate(cols, start=1):
        c = ws.cell(row=head, column=i, value=title)
        c.font = Font(bold=True, size=10, color=_WHITE)
        c.fill = PatternFill("solid", fgColor=_NAVY)
        c.alignment = _CENTER
        c.border = _BORDER
    ws.row_dimensions[head].height = 22
    ws.freeze_panes = ws.cell(row=head + 1, column=1)
    r += 1

    # A single unnamed section has nothing to subtotal — its subtotal and the
    # gross would be the same figure twice, which reads as a mistake.
    show_subtotals = len(ctx["sections"]) > 1 or bool(
        ctx["sections"] and ctx["sections"][0]["section"])
    section_total_rows, line_rows = [], []
    for sec in ctx["sections"]:
        first_line = None
        if sec["section"]:
            ws.merge_cells(start_row=r, start_column=1, end_row=r,
                           end_column=last)
            c = ws.cell(row=r, column=1, value=sec["section"])
            c.font = Font(bold=True, size=10, color=_NAVY)
            c.fill = PatternFill("solid", fgColor=_SECTION)
            c.alignment = _LEFT
            for i in range(1, last + 1):
                ws.cell(row=r, column=i).border = _BORDER
            r += 1
        for ln in sec["lines"]:
            it = ln["item"]
            ws.cell(row=r, column=1, value=it.item_code or "")
            d = ws.cell(row=r, column=2, value=it.description or "")
            d.alignment = _LEFT_TOP
            if it.is_heading:
                # A heading carries no money — leave the cells genuinely
                # empty so a SUM over the block cannot pick anything up.
                ws.cell(row=r, column=1).font = Font(bold=True)
                d.font = Font(bold=True)
            else:
                ws.cell(row=r, column=3, value=it.unit or "")
                q = ws.cell(row=r, column=4, value=it.qty)
                q.number_format = _QTY
                q.alignment = _RIGHT
                if split:
                    rates = [(5, it.rate_supply), (6, it.rate_install)]
                    rate_expr = (f"({get_column_letter(5)}{r}"
                                 f"+{get_column_letter(6)}{r})")
                else:
                    rates = [(5, it.rate_total)]
                    rate_expr = f"{get_column_letter(5)}{r}"
                for col, val in rates:
                    c = ws.cell(row=r, column=col, value=val)
                    c.number_format = _MONEY
                    c.alignment = _RIGHT
                a = ws.cell(row=r, column=last,
                            value=f"={qty_letter}{r}*{rate_expr}")
                a.number_format = _MONEY
                a.alignment = _RIGHT
                if first_line is None:
                    first_line = r
                line_rows.append(r)
            for i in range(1, last + 1):
                ws.cell(row=r, column=i).border = _BORDER
            r += 1
        if not show_subtotals:
            continue
        # Section total — a SUM over its own block, so inserted rows count.
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=last - 1)
        c = ws.cell(row=r, column=1,
                    value=(f"Subtotal · {sec['section']}" if sec["section"]
                           else "Subtotal"))
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="right", vertical="center")
        t = ws.cell(row=r, column=last,
                    value=(f"=SUM({amt_letter}{first_line}:{amt_letter}{r-1})"
                           if first_line else 0))
        t.font = Font(bold=True)
        t.number_format = _money_fmt(ccy)
        for i in range(1, last + 1):
            ws.cell(row=r, column=i).fill = PatternFill("solid",
                                                        fgColor=_GREY)
            ws.cell(row=r, column=i).border = _BORDER
        section_total_rows.append(r)
        r += 1

    gross_row = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last - 1)
    c = ws.cell(row=r, column=1, value=f"Gross value of this variation ({ccy})")
    c.font = Font(bold=True, size=11, color=_WHITE)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.fill = PatternFill("solid", fgColor=_NAVY)
    if section_total_rows:
        gross_formula = "=" + "+".join(f"{amt_letter}{x}"
                                       for x in section_total_rows)
    elif line_rows:
        gross_formula = (f"=SUM({amt_letter}{line_rows[0]}"
                         f":{amt_letter}{line_rows[-1]})")
    else:
        gross_formula = 0
    g = ws.cell(row=r, column=last, value=gross_formula)
    g.font = Font(bold=True, size=11, color=_WHITE)
    g.fill = PatternFill("solid", fgColor=_NAVY)
    g.number_format = _money_fmt(ccy)
    r += 2

    # ---- effect on the contract sum -----------------------------------
    band("Effect on the contract sum", size=11, fill=_BAND, height=18)

    def row(label, value, bold=False, fill=None):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=last - 1)
        c = ws.cell(row=r, column=1, value=label)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.font = Font(bold=bold, size=10)
        m = ws.cell(row=r, column=last, value=value)
        m.number_format = _money_fmt(ccy)
        m.font = Font(bold=bold, size=10)
        if fill:
            for i in range(1, last + 1):
                ws.cell(row=r, column=i).fill = PatternFill("solid",
                                                            fgColor=fill)
        for i in range(1, last + 1):
            ws.cell(row=r, column=i).border = _BORDER
        r += 1
        return r - 1

    row("Original contract sum", ctx["original"])
    row("Variations already approved", ctx["prior_approved"])
    before_row = row("Contract sum before this variation",
                     f"={amt_letter}{r-2}+{amt_letter}{r-1}", bold=True)
    # An omission subtracts: the sign is the variation's kind, not the
    # QS's arithmetic, so the formula carries it.
    sign = "-" if ctx["is_omission"] else ""
    this_row = row(f"This variation ({ctx['kind_label'].lower()})",
                   f"={sign}{amt_letter}{gross_row}", bold=True)
    # Four VOs submitted together each read as if it were the first, so the
    # bottom line has to say what this one sits on top of — the same rule the
    # PDF follows (owner 2026-08-22).
    if ctx["prior_pending"]:
        pending_row = row(f"Other variations awaiting approval "
                          f"({ctx['prior_pending_refs']})",
                          ctx["prior_pending_net"])
        total = (f"={amt_letter}{before_row}+{amt_letter}{pending_row}"
                 f"+{amt_letter}{this_row}")
        label = "Anticipated contract sum if all are approved"
    else:
        total = f"={amt_letter}{before_row}+{amt_letter}{this_row}"
        label = "Resulting contract sum"
    row(label, total, bold=True, fill=_SECTION)
    return wb
