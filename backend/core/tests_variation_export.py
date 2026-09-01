"""The QS's working copy of a variation (xlsx).

The point of this sheet is that its money is live: change a qty and the
variation total and the contract sum follow. So these tests do not check that
formulas are present — they evaluate them, and compare the answer to what the
model says (owner 2026-09-01).
"""
import re
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from datetime import date, timedelta

from django.test import TestCase
from rest_framework.test import APIClient

from .models import Project, Site, User, Variation
from .tests import make_user


class _Sheet:
    """Just enough spreadsheet to resolve the formulas this export writes:
    cell references, SUM over a range, + - * and parentheses."""

    def __init__(self, ws):
        self.ws = ws

    def value(self, ref):
        v = self.ws[ref].value
        if isinstance(v, str) and v.startswith("="):
            return self._eval(v[1:])
        return Decimal(str(v)) if isinstance(v, (int, float)) else v

    def _eval(self, expr):
        def sum_range(m):
            col, a, b = m.group(1), int(m.group(2)), int(m.group(3))
            total = sum((self.value(f"{col}{i}") or Decimal("0")
                         for i in range(a, b + 1)), Decimal("0"))
            return f"Decimal('{total}')"

        expr = re.sub(r"SUM\(([A-Z]+)(\d+):[A-Z]+(\d+)\)", sum_range, expr)
        expr = re.sub(r"\b([A-Z]+)(\d+)\b",
                      lambda m: f"Decimal('{self.value(m.group(0)) or 0}')",
                      expr)
        return eval(expr, {"Decimal": Decimal, "__builtins__": {}})


def _load(v):
    from .variation_export import build_variation_xlsx
    buf = BytesIO()
    build_variation_xlsx(v).save(buf)
    buf.seek(0)
    return load_workbook(buf).active


def _find(ws, text):
    """Row number whose first cell starts with `text` (headers are merged)."""
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith(text):
                return c.row
    raise AssertionError(f"no row starting {text!r}")


class _VariationBase(TestCase):
    def setUp(self):
        self.site = Site.objects.create(
            code="VKR", name="Vakkaru", status=Site.Status.ACTIVE,
            start_date=date.today() - timedelta(days=90))
        self.project = Project.objects.create(
            site=self.site, code="POOLS17", title="17 Swimming Pools",
            contract_value="500000")
        self.qs = make_user("qs1", User.Role.QS)
        self.client = APIClient()
        self.client.force_authenticate(self.qs)

    def _create(self, kind="ADDITION"):
        r = self.client.post(
            f"/api/v1/projects/{self.project.id}/variations/create",
            {"title": "Extra pool coping", "kind": kind, "rows": [
                {"item_code": "V1", "description": "Coping stone", "unit": "m",
                 "qty": "40", "rate_supply": "25", "rate_install": "10"}]},
            format="json")
        self.assertEqual(r.status_code, 201, r.data)
        return r.data["variations"][-1]


class VariationWorkingCopyTests(_VariationBase):

    def test_the_sheet_totals_agree_with_the_variation(self):
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        ws = _load(v)
        sh = _Sheet(ws)
        col = get_column_letter(ws.max_column)
        gross_row = _find(ws, "Gross value of this variation")
        self.assertEqual(sh.value(f"{col}{gross_row}"), v.gross)
        self.assertEqual(v.gross, Decimal("1400"))      # 40 × (25 + 10)

    def test_editing_a_qty_carries_through_to_the_contract_sum(self):
        """The whole reason this is a spreadsheet and not a PDF."""
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        ws = _load(v)
        col = get_column_letter(ws.max_column)
        gross_row = _find(ws, "Gross value of this variation")
        result_row = _find(ws, "Resulting contract sum")
        self.assertEqual(_Sheet(ws).value(f"{col}{result_row}"),
                         Decimal("501400"))             # 500,000 + 1,400
        # the QS doubles the quantity in the sheet
        item_row = next(c.row for r in ws.iter_rows() for c in r
                        if c.value == "Coping stone")
        self.assertEqual(Decimal(str(ws[f"D{item_row}"].value)),
                         Decimal("40"))
        ws[f"D{item_row}"] = 80
        sh = _Sheet(ws)
        self.assertEqual(sh.value(f"{col}{gross_row}"), Decimal("2800"))
        self.assertEqual(sh.value(f"{col}{result_row}"), Decimal("502800"))

    def test_an_omission_subtracts_from_the_contract_sum(self):
        """The sign belongs to the variation's kind, not to the QS's
        arithmetic — so the formula carries it."""
        v = Variation.objects.get(pk=self._create("OMISSION")["id"])
        ws = _load(v)
        col = get_column_letter(ws.max_column)
        gross_row = _find(ws, "Gross value of this variation")
        result_row = _find(ws, "Resulting contract sum")
        sh = _Sheet(ws)
        self.assertEqual(sh.value(f"{col}{gross_row}"), Decimal("1400"))
        self.assertEqual(v.signed_total, Decimal("-1400"))
        self.assertEqual(sh.value(f"{col}{result_row}"), Decimal("498600"))

    def test_rates_and_qty_are_numbers_not_text(self):
        """A rate written as text cannot be worked on."""
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        ws = _load(v)
        row = next(c.row for r in ws.iter_rows() for c in r
                   if c.value == "Coping stone")
        for letter, expected in (("D", 40), ("E", 25), ("F", 10)):
            cell = ws[f"{letter}{row}"]
            self.assertIsInstance(cell.value, (int, float, Decimal),
                                  f"{letter} is {type(cell.value)}")
            self.assertEqual(Decimal(str(cell.value)), Decimal(str(expected)))

    def test_the_sheet_says_it_is_not_the_issued_document(self):
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        ws = _load(v)
        text = " ".join(str(c.value) for r in ws.iter_rows() for c in r
                        if c.value)
        self.assertIn("WORKING COPY", text)
        self.assertIn("not the issued variation order", text)

    def test_a_draft_can_be_exported_even_though_the_pdf_refuses(self):
        """Working on the price before we stand behind it is the point."""
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        self.assertEqual(v.status, "DRAFT")
        r = self.client.get(f"/api/v1/variations/{v.id}/vo.pdf")
        self.assertEqual(r.status_code, 400)            # the client document
        r = self.client.get(f"/api/v1/variations/{v.id}/vo.xlsx")
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])
        self.assertIn(f"{v.ref}-working.xlsx", r["Content-Disposition"])

    def test_an_empty_variation_is_refused(self):
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        v.items.all().delete()
        r = self.client.get(f"/api/v1/variations/{v.id}/vo.xlsx")
        self.assertEqual(r.status_code, 400)


class ApprovedWithoutADateTests(_VariationBase):
    """Three live variations are APPROVED with no employer approval date —
    approved before the rule that requires date + ref. Formatting the missing
    date threw, so the VO PDF button 500'd on them (owner 2026-09-01)."""

    def _approved_undated(self):
        v = Variation.objects.get(pk=self._create("ADDITION")["id"])
        Variation.objects.filter(pk=v.pk).update(
            status="APPROVED", employer_approved_on=None, employer_ref="")
        v.refresh_from_db()
        return v

    def test_the_pdf_context_prints_without_the_date(self):
        from . import commercial
        ctx = commercial.variation_pdf_context(self._approved_undated())
        self.assertEqual(ctx["client_status"], "Approved by the Employer")
        self.assertIn("forms part of the contract", ctx["status_note"])
        self.assertNotIn("None", ctx["status_note"])

    def test_both_documents_render(self):
        v = self._approved_undated()
        self.assertEqual(
            self.client.get(f"/api/v1/variations/{v.id}/vo.pdf").status_code,
            200)
        self.assertEqual(
            self.client.get(f"/api/v1/variations/{v.id}/vo.xlsx").status_code,
            200)
