"""Employees & site timesheets (spec §6A). Sensitive fields (basic_pay,
passport_no) are serialized only for HO HR / Admin, never logged, and
excluded from site-level responses."""

from datetime import date, datetime, timedelta
from decimal import Decimal

from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import serializers, viewsets
from rest_framework.decorators import action, api_view
from rest_framework.permissions import BasePermission
from rest_framework.response import Response

from .audit import audit
from .models import (
    Attendance,
    CompanyParameter,
    Employee,
    EmployeeSiteAllocation,
    ManpowerCategory,
    OnboardingCase,
    OvertimeRate,
    Site,
    TimesheetMonth,
    User,
)
from . import leave as _leave
from .permissions import PAY_ROLES, scoped_site_ids

# PA (Director's office) has full HR access alongside HR/Admin (owner 2026-08-03)
HR_ROLES = ("HO_HR", "ADMIN", "PA")
# R3 addendum, plus the signatory who signs every payroll PYR (2026-08-16).
# One definition, in permissions.py — the site-facing screens redact against
# the same list.
PAYROLL_ROLES = PAY_ROLES
# passport/permit/contact: HR+Admin only; basic_pay also visible to Finance
SENSITIVE_FIELDS = ("passport_no", "passport_expiry", "work_permit_no",
                    "work_visa_number", "medical_expiry", "insurance_expiry",
                    "emergency_contact")
PAY_FIELDS = ("basic_pay", "usd_basic_pay")


def _is_hr(user):
    return user.role in HR_ROLES


def _sees_pay(user):
    return user.role in PAYROLL_ROLES


class IsHrOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.method in ("GET", "HEAD", "OPTIONS"):
            return True
        return _is_hr(request.user)


class EmployeeSerializer(serializers.ModelSerializer):
    site_id = serializers.SerializerMethodField()
    site_code = serializers.SerializerMethodField()
    job_category_name = serializers.CharField(source="job_category.name",
                                              read_only=True, default=None)
    photo_url = serializers.SerializerMethodField()
    ot_rate = serializers.SerializerMethodField()
    ot_effective = serializers.SerializerMethodField()
    permit_state = serializers.SerializerMethodField()
    permit_days = serializers.SerializerMethodField()
    permit_pending = serializers.SerializerMethodField()

    def validate_passport_no(self, value):
        """One passport, one record.

        Rakib Hossain was on BVR's July payroll twice — as EMP-0020 with the
        23 days he had actually worked, and again as EMP-0603, a second record
        created for him in August. The site reported his July attendance as
        "gone missing"; it was on the first record all along. Thirty-eight
        passports are on more than one record and seventeen of those have two
        live records (owner 2026-08-15).

        Blocked rather than warned, because the duplicate is invisible
        afterwards — it looks like a new man with no history. HR can still
        proceed with `allow_duplicate_passport` when it is the OTHER record
        that carries the typo, which is about half of them.
        """
        pno = (value or "").strip()
        if not pno:
            return value
        clash = Employee.objects.filter(passport_no__iexact=pno)
        if self.instance is not None:
            clash = clash.exclude(pk=self.instance.pk)
        other = clash.first()
        req = self.context.get("request")
        override = req and str(
            req.data.get("allow_duplicate_passport", "")).lower() in (
            "1", "true", "yes")
        if override and self.instance is not None:
            self.instance._allow_duplicate_passport = True
        if other and not override:
            raise serializers.ValidationError(
                f"Passport {pno} is already on {other.emp_no} "
                f"{other.full_name}. If this is the same man, use that record "
                f"so his history stays with him. If the passport on "
                f"{other.emp_no} is wrong, correct it there first — or resend "
                f"with allow_duplicate_passport to proceed anyway.")
        return value

    def create(self, validated_data):
        """Carry the override onto the new record.

        On an edit the flag can be set on `self.instance`, but a create has no
        instance until here — without this the model guard would refuse the
        very case the override exists for (owner 2026-08-16).
        """
        obj = Employee(**validated_data)
        req = self.context.get("request")
        if req and str(req.data.get("allow_duplicate_passport", "")).lower() \
                in ("1", "true", "yes"):
            obj._allow_duplicate_passport = True
        obj.save()
        return obj

    class Meta:
        model = Employee
        fields = ["id", "emp_no", "full_name", "photo", "photo_url",
                  "date_of_birth", "gender", "marital_status",
                  "passport_no", "passport_expiry", "nationality",
                  "job_category", "job_category_name", "job_title",
                  "basic_pay",
                  "usd_basic_pay", "currency",
                  "ot_applies", "ot_rate", "ot_effective", "employment_type",
                  "work_permit_no", "work_permit_expiry", "work_visa_number",
                  "medical_expiry", "insurance_expiry", "permit_state",
                  "permit_days", "permit_pending", "emergency_contact",
                  "join_date", "is_active", "site_id", "site_code"]
        read_only_fields = ["emp_no", "photo_url", "ot_rate", "ot_effective",
                            "permit_state", "permit_days", "permit_pending"]
        extra_kwargs = {"photo": {"write_only": True, "required": False}}

    def get_photo_url(self, obj):
        return obj.photo.url if obj.photo else None

    def get_permit_state(self, obj):
        from . import permits
        return permits.permit_status(obj)[0]

    def get_permit_days(self, obj):
        from . import permits
        return permits.permit_status(obj)[1]

    def get_permit_pending(self, obj):
        return obj.permit_renewals.filter(applied=False).exists()

    def get_ot_rate(self, obj):
        return obj.ot_rate()

    def get_ot_effective(self, obj):
        return obj.ot_rate() > 0

    def get_site_id(self, obj):
        return obj.current_site_id()

    def get_site_code(self, obj):
        row = obj.site_allocations.filter(to_date__isnull=True) \
            .select_related("site").first()
        return row.site.code if row else None

    def validate(self, attrs):
        inst = self.instance
        usd = attrs.get("usd_basic_pay",
                        getattr(inst, "usd_basic_pay", None))
        emp_type = attrs.get("employment_type",
                             getattr(inst, "employment_type", "PERMANENT"))
        # Split pay (USD basic) is permanent-only, and a split worker is paid no
        # MVR basic — force it to 0 (owner 2026-08-06).
        if usd and usd > 0:
            if emp_type != "PERMANENT":
                raise serializers.ValidationError({"usd_basic_pay":
                    "Only permanent workers can be paid a USD basic."})
            attrs["basic_pay"] = Decimal("0")
        return attrs

    def to_representation(self, instance):
        data = super().to_representation(instance)
        request = self.context.get("request")
        # Site users see emp no, name, category only (spec §6A.1)
        if request and not _is_hr(request.user):
            for field in SENSITIVE_FIELDS:
                data.pop(field, None)
        if request and not _sees_pay(request.user):
            for field in PAY_FIELDS + ("ot_rate", "currency"):
                data.pop(field, None)
        return data


class EmployeeViewSet(viewsets.ModelViewSet):
    serializer_class = EmployeeSerializer
    permission_classes = [IsHrOrReadOnly]
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    # What makes an employee record REAL history — anything here and the record
    # is part of the company's books, so it may never be deleted (owner
    # 2026-08-12: admin delete exists to remove duplicate/erroneous records
    # found in the expat-portal reconciliation, not to erase people).
    # NB salary_revisions / permit_renewals CASCADE off Employee — listing them
    # here stops a delete silently taking real records with it.
    _HISTORY = (
        ("attendance", "attendance days"),
        ("payroll_lines", "payroll lines"),
        ("salary_advances", "salary advances"),
        ("change_items", "worker-change batches"),
        ("salary_revisions", "salary revisions"),
        ("permit_renewals", "work-permit renewals"),
    )

    def _delete_blockers(self, emp):
        out = []
        for attr, label in self._HISTORY:
            mgr = getattr(emp, attr, None)
            n = mgr.count() if mgr is not None else 0
            if n:
                out.append(f"{n} {label}")
        return out

    @action(detail=True, methods=["get"])
    def deletable(self, request, pk=None):
        """Why this record can (or cannot) be deleted — the confirm dialog
        reads this before offering the button."""
        emp = self.get_object()
        blockers = self._delete_blockers(emp)
        return Response({
            "can_delete": request.user.role == "ADMIN" and not blockers,
            "is_admin": request.user.role == "ADMIN",
            "blockers": blockers,
            "allocations": emp.site_allocations.count(),
            # an onboarding case merely detaches (SET_NULL, related_name="+"
            # so it has no reverse accessor) — never a blocker
            "onboarding_case": OnboardingCase.objects.filter(
                employee=emp).exists(),
        })

    def destroy(self, request, *args, **kwargs):
        """Delete an employee record — ADMIN only, and only when it carries no
        real history. Site allocations (the only soft link) are removed with
        it; an onboarding case is detached, not deleted."""
        if request.user.role != "ADMIN":
            return Response(
                {"detail": "Only an administrator can delete an employee "
                           "record."}, status=403)
        emp = self.get_object()
        blockers = self._delete_blockers(emp)
        if blockers:
            return Response(
                {"detail": f"{emp.emp_no} carries " + ", ".join(blockers)
                           + " — it can't be deleted. Deactivate it instead so "
                             "the history stays intact."}, status=400)
        emp_id = emp.id            # keep the id — the audit row outlives the record
        detail = {"emp_no": emp.emp_no, "name": emp.full_name,
                  "passport": emp.passport_no or "",
                  "engagement": emp.engagement_type,
                  "allocations_removed": emp.site_allocations.count()}
        with transaction.atomic():
            emp.site_allocations.all().delete()   # PROTECTed, but no history
            emp.delete()
        audit("employee", emp_id, "EMPLOYEE_DELETED", actor=request.user,
              detail=detail)
        return Response({"deleted": True, **detail})

    def get_queryset(self):
        qs = Employee.objects.select_related("job_category").order_by("emp_no")
        # HR owns direct employees only — subcontract workers are absent from the
        # register (D-b). Site workforce views opt in with include_subcontract=1.
        if self.request.GET.get("include_subcontract") != "1":
            qs = qs.hr_managed()
        site_ids = scoped_site_ids(self.request.user)
        if site_ids is not None:  # site roles: own roster only
            qs = qs.filter(site_allocations__site_id__in=site_ids,
                           site_allocations__to_date__isnull=True)
        if self.request.GET.get("site"):
            qs = qs.filter(site_allocations__site_id=self.request.GET["site"],
                           site_allocations__to_date__isnull=True)
        if self.request.GET.get("active") != "all":
            qs = qs.filter(is_active=True)
        return qs.distinct()

    @action(detail=False, methods=["get"])
    def export(self, request):
        """Download the employee register as xlsx, matching the on-screen
        filters (owner 2026-08-06). Pay columns only for pay-seeing roles."""
        if request.user.role not in PAYROLL_ROLES:
            return Response({"detail": "Not permitted."}, status=403)
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        qs = Employee.objects.select_related("job_category").order_by("emp_no")
        if request.GET.get("include_subcontract") != "1":
            qs = qs.hr_managed()
        site_ids = scoped_site_ids(request.user)
        if site_ids is not None:
            qs = qs.filter(site_allocations__site_id__in=site_ids,
                           site_allocations__to_date__isnull=True)
        data = EmployeeSerializer(qs.distinct(), many=True,
                                  context={"request": request}).data

        q = request.GET.get("q", "").lower()
        fsite = request.GET.get("site", "")
        fcat = request.GET.get("category", "")
        fcur = request.GET.get("currency", "")
        fstatus = request.GET.get("status", "active")
        femp = request.GET.get("employment", "")
        fnat = request.GET.get("nationality", "")

        def keep(e):
            if q and q not in f"{e['emp_no']} {e['full_name']}".lower():
                return False
            if fsite == "__none":
                if e.get("site_code"):
                    return False
            elif fsite and e.get("site_code") != fsite:
                return False
            if fcat and str(e.get("job_category")) != fcat:
                return False
            if fcur and e.get("currency") != fcur:
                return False
            if fstatus == "active" and not e["is_active"]:
                return False
            if fstatus == "inactive" and e["is_active"]:
                return False
            if femp and e["employment_type"] != femp:
                return False
            if fnat and e.get("nationality") != fnat:
                return False
            return True

        rows = [e for e in data if keep(e)]
        seespay = _sees_pay(request.user)
        # The master export adds sensitive identity fields (passport, DOB,
        # emergency contact) for reconciliation — HR only.
        full = request.GET.get("full") == "1"
        if full and not _is_hr(request.user):
            return Response({"detail": "The full master export is HR only."},
                            status=403)
        emp_label = {"PERMANENT": "Permanent", "CONTRACT": "Contract"}
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        headers = ["Emp No", "Name", "Nationality"]
        if full:
            headers += ["Gender", "Marital Status", "Date of Birth",
                        "Passport No", "Passport Expiry"]
        headers += ["Category", "Site", "Employment",
                    "Work Permit No", "WP Expiry"]
        if full:
            headers += ["Work Visa No", "Medical Expiry", "Insurance Expiry"]
        headers += ["Permit Status"]
        if seespay:
            headers += ["Basic Pay", "Currency", "USD Basic", "Overtime"]
        headers += ["Join Date"]
        if full:
            headers += ["Emergency Contact"]
        headers += ["Status"]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)
        for e in rows:
            row = [e["emp_no"], e["full_name"], e.get("nationality", "")]
            if full:
                row += [e.get("gender", ""), e.get("marital_status", ""),
                        e.get("date_of_birth") or "", e.get("passport_no", ""),
                        e.get("passport_expiry") or ""]
            row += [e.get("job_category_name") or "", e.get("site_code") or "",
                    emp_label.get(e["employment_type"], e["employment_type"]),
                    e.get("work_permit_no", ""),
                    e.get("work_permit_expiry") or ""]
            if full:
                row += [e.get("work_visa_number", ""),
                        e.get("medical_expiry") or "",
                        e.get("insurance_expiry") or ""]
            row += [e.get("permit_state") or ""]
            if seespay:
                row += [e.get("basic_pay") or "", e.get("currency") or "",
                        e.get("usd_basic_pay") or "",
                        f'{e["ot_rate"]}/hr' if e.get("ot_effective") else ""]
            row += [e.get("join_date") or ""]
            if full:
                row += [e.get("emergency_contact", "")]
            row += ["Active" if e["is_active"] else "Inactive"]
            ws.append(row)
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 3)
        resp = HttpResponse(content_type="application/vnd.openxmlformats-"
                            "officedocument.spreadsheetml.sheet")
        fname = "employees-master.xlsx" if full else "employees.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(resp)
        return resp

    def perform_create(self, serializer):
        from .numbering import next_ref

        with transaction.atomic():
            n = int(next_ref("EMP", None).split("-")[1])
            employee = serializer.save(emp_no=f"EMP-{n:04d}")
            # Optional initial posting — a project site or Head Office (MLE).
            sid = self.request.data.get("site_id")
            site = Site.objects.filter(pk=sid).first() if sid else None
            if site:
                EmployeeSiteAllocation.objects.create(
                    employee=employee, site=site, from_date=date.today())
        audit("employee", employee.id, "EMPLOYEE_CREATED",
              actor=self.request.user,
              detail={"emp_no": employee.emp_no,
                      "site": site.code if site else None})

    def perform_update(self, serializer):
        employee = serializer.save()
        audit("employee", employee.id, "EMPLOYEE_UPDATED",
              actor=self.request.user,
              detail={"fields": sorted(
                  k for k in self.request.data
                  if k not in SENSITIVE_FIELDS + PAY_FIELDS)})

    @action(detail=True, methods=["post"])
    def allocate(self, request, pk=None):
        """Transfer to a site; history kept for payroll (spec §6A.1)."""
        employee = self.get_object()
        try:
            site = Site.objects.get(pk=request.data.get("site_id"))
        except Site.DoesNotExist:
            return Response({"detail": "Unknown site_id."}, status=400)
        today = date.today()
        employee.site_allocations.filter(to_date__isnull=True) \
            .update(to_date=today)
        EmployeeSiteAllocation.objects.create(employee=employee, site=site,
                                              from_date=today)
        audit("employee", employee.id, "EMPLOYEE_ALLOCATED",
              actor=request.user, to_state=site.code)
        return Response(self.get_serializer(employee).data)

    @action(detail=True, methods=["post"])
    def deactivate(self, request, pk=None):
        employee = self.get_object()
        employee.is_active = False
        employee.save(update_fields=["is_active"])
        employee.site_allocations.filter(to_date__isnull=True) \
            .update(to_date=date.today())
        audit("employee", employee.id, "EMPLOYEE_DEACTIVATED",
              actor=request.user)
        return Response(self.get_serializer(employee).data)

    @action(detail=True, methods=["get"], url_path="permit-renewals")
    def permit_renewals(self, request, pk=None):
        """Renewal history for a worker (HR/Admin). Pending rows (PYR not yet
        paid) have no new_expiry and applied=False."""
        if not _is_hr(request.user):
            return Response({"detail": "HR/Admin only."}, status=403)
        employee = self.get_object()
        rows = [{
            "months": r.months, "previous_expiry": r.previous_expiry,
            "new_expiry": r.new_expiry, "applied": r.applied,
            "note": r.note, "fee": r.fee, "by": r.created_by.full_name,
            "pyr": r.document.ref if r.document_id else None,
            "pyr_status": r.document.status if r.document_id else None,
            "at": r.created_at,
        } for r in employee.permit_renewals.select_related(
            "created_by", "document")]
        return Response(rows)


@api_view(["GET"])
def permit_alerts(request):
    """Work permits expiring within 30 days (or already expired), for the HR
    view. Site users see their own roster; HR/Admin see everyone."""
    from . import permits
    site_ids = scoped_site_ids(request.user)
    rows = permits.alerts(site_ids=site_ids)
    return Response({
        "within_days": permits.ALERT_DAYS,
        "expired": [r for r in rows if r["state"] == "EXPIRED"],
        "expiring": [r for r in rows if r["state"] == "EXPIRING"],
    })


@api_view(["POST"])
def permit_batch_renew(request):
    """HR raises ONE PYR to renew several permits at once. Each selected
    worker's permit is extended by its months and the PYR carries the total
    renewal fee (Permits & Fees cost head) through the payment workflow, at
    Head Office. Body: {lines:[{employee_id, months, fee, permit_no?}],
    payee?, cost_head_id?, currency?, purpose?}."""
    from .models import CostHead, Document, DocumentRevision
    from .numbering import next_ref
    from .payments import create_payment_request
    if not _is_hr(request.user):
        return Response({"detail": "HR/Admin only."}, status=403)
    lines = request.data.get("lines") or []
    if not lines:
        return Response({"detail": "Select at least one worker to renew."},
                        status=400)
    site = Site.objects.filter(is_head_office=True).first()
    if site is None:
        return Response({"detail": "No Head Office site is configured."},
                        status=400)
    cost_head_id = request.data.get("cost_head_id")
    if not cost_head_id:
        ch = CostHead.objects.filter(name="Permits & Fees",
                                     is_active=True).first()
        cost_head_id = ch.id if ch else None
    payee = request.data.get("payee") or "Work-permit renewals"
    purpose = request.data.get("purpose") or "Work-permit renewals (batch)"
    data = {
        "permit_lines": lines, "cost_head_id": cost_head_id, "payee": payee,
        "currency": request.data.get("currency", "MVR"), "purpose": purpose,
        "payment_method": request.data.get("payment_method", "BANK"),
        "has_supporting_doc": bool(request.data.get("has_supporting_doc")),
        "no_doc_reason": request.data.get("no_doc_reason", ""),
    }
    with transaction.atomic():
        ref = next_ref("PYR", site)
        doc = Document.objects.create(
            doc_type="PYR", ref=ref, site=site, doc_date=date.today(),
            status="DRAFT", created_by=request.user)
        rev = DocumentRevision.objects.create(
            document=doc, rev_label="R0", created_by=request.user,
            payload={"purpose": purpose, "payee": payee,
                     "kind": "permit_renewal"})
        doc.current_revision = rev
        doc.save(update_fields=["current_revision"])
        pr, err = create_payment_request(doc, data, request.user)
        if err:
            transaction.set_rollback(True)
            return Response({"detail": err}, status=400)
    audit("document", doc.id, "DOC_CREATED", actor=request.user,
          to_state="DRAFT", detail={"ref": ref, "kind": "permit_renewal"})
    # Raising the batch IS the request — leaving it in DRAFT for HR to submit
    # separately just stranded it (owner 2026-08-13). Renewals need no
    # approval, so this puts it straight in front of Finance. Goes through the
    # normal action so the transition, audit trail and notifications are the
    # same as any other PYR.
    from .payments import pyr_action

    err = pyr_action(request, doc, "submit")
    doc.refresh_from_db()
    return Response({"ref": doc.ref, "amount": str(pr.amount_requested),
                     "currency": pr.currency, "count": len(lines),
                     "status": doc.status,
                     "submit_error": (err.data.get("detail") if err else None)},
                    status=201)


# ===== Attendance =====


def _month_locked(site_id, day):
    return TimesheetMonth.objects.filter(
        site_id=site_id, year=day.year, month=day.month, status="LOCKED"
    ).exists()


def _allocated_elsewhere(employee, site, day):
    """The other site whose allocation covers `day`, when this site's doesn't.

    A worker has one attendance row per day wherever he is, so when two sites
    claim the same day one of them loses it — silently, because the second
    save overwrites the first. OUT's clerk marked MD SADAM HUSSAN absent for
    1–6 August on the strength of an open-ended allocation that only began on
    the 20th; SSL had him all six days, could not mark them because the rows
    already existed, and he lost six days' pay (owner 2026-09-03).

    Late-filed allocations still allow the back-entry sites depend on — BVR's
    whole crew carries the day the site was loaded into the app — because
    this only speaks up when ANOTHER site's allocation actually covers the
    day in question.
    """
    covers = EmployeeSiteAllocation.objects.filter(
        employee=employee, from_date__lte=day).filter(
        Q(to_date__isnull=True) | Q(to_date__gte=day))
    if covers.filter(site=site).exists():
        return None
    # Head office is not a competing claim: leave parks a man on the MLE
    # allocation while he is away, and his own site still has to be able to
    # mark the day he comes back.
    other = (covers.exclude(site=site).exclude(site__is_head_office=True)
             .select_related("site").first())
    return other.site.code if other else None


def _normal_hours(site, check_in, check_out, remark, shift=None):
    """Hours inside the worker's window — his shift's if he has one, else
    the site's. A night shift's window (and its punches) run past midnight."""
    from .shifts import schedule_for, window_datetimes
    win_from, win_to, _, _ = schedule_for(site, shift)
    win_s, win_e = window_datetimes(date.today(), win_from, win_to)
    full = Decimal(str((win_e - win_s).total_seconds())) / 3600
    if remark in ("ABSENT", "SICK", "LEAVE"):
        return Decimal("0")
    if remark == "HALF_DAY":
        return (full / 2).quantize(Decimal("0.01"))
    if check_in and check_out:
        ci = datetime.combine(date.today(), check_in)
        co = datetime.combine(date.today(), check_out)
        if co <= ci:
            co += timedelta(days=1)      # out after midnight
        overlap = (min(co, win_e) - max(ci, win_s)).total_seconds()
        return max(Decimal(str(overlap)) / 3600, Decimal("0")) \
            .quantize(Decimal("0.01"))
    return full.quantize(Decimal("0.01"))  # present, default window


def _site_scope_ok(request, site):
    site_ids = scoped_site_ids(request.user)
    return site_ids is None or site.id in site_ids


@api_view(["GET"])
def attendance_grid(request):
    """Whole crew on one screen (spec §6A.2): roster with existing rows."""
    try:
        site = Site.objects.get(pk=request.GET.get("site"))
        day = date.fromisoformat(request.GET.get("date"))
    except (Site.DoesNotExist, TypeError, ValueError):
        return Response({"detail": "site and date required."}, status=400)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not found."}, status=404)

    # A rest day is any weekday outside the site's working week (usually
    # Friday). Working it is the 7th-day work paid as an extra day, so the grid
    # defaults everyone to OFF and only marks those who actually worked.
    is_rest_day = day.isoweekday() not in site.working_days
    # The crew as at THIS DAY. Two faults came out of rostering by "today":
    # a leaver disappeared from days he had worked, and a man who had not
    # joined yet was offered for marking — which is how Hossain sharif and
    # Robiul picked up 1 and 2 July while a clerk was fixing somebody else's
    # row, months before they joined on 5 August (owner 2026-08-14).
    # Anyone whose allocation covers the day, plus anyone on the site now —
    # allocation start dates are entered late and in bulk (BVR's whole crew
    # carries the day the site was loaded), so being strict about them would
    # block the back-entry the sites actually do. The join date is the one
    # hard edge, because HR owns it and the owner is explicit: nothing before
    # a man joined.
    on_day = EmployeeSiteAllocation.objects.filter(site=site).filter(
        Q(to_date__isnull=True) | Q(to_date__gte=day, from_date__lte=day))
    existing = {a.employee_id: a for a in Attendance.objects.filter(
        site=site, day=day)}
    # A day another site's allocation covers is that site's day to mark, not
    # this one's — see _allocated_elsewhere. Anyone already marked here stays
    # on the grid, so an existing row can always be corrected or taken back.
    covers_day = EmployeeSiteAllocation.objects.filter(
        from_date__lte=day).filter(
        Q(to_date__isnull=True) | Q(to_date__gte=day))
    blocked = (set(covers_day.exclude(site=site)
                   .exclude(site__is_head_office=True)
                   .values_list("employee_id", flat=True))
               - set(covers_day.filter(site=site)
                     .values_list("employee_id", flat=True))
               - set(existing))
    # Anyone who ALREADY has a mark for this day, whatever the roster says.
    # The month register reads the rows directly, so it showed a man whose
    # allocation had since been closed — while this grid, built only from
    # allocations, could not show him at all. His three wrong days were
    # visible and uncorrectable at the same time, and the clerk's attempts to
    # mark him absent vanished (owner 2026-09-02, EMP-0524 on SSL).
    #
    # Whatever the register displays has to be reachable here, or there is no
    # way to take it back.
    roster = Employee.objects.filter(
        Q(id__in=on_day.values_list("employee_id", flat=True))
        | Q(id__in=list(existing))).exclude(
        id__in=blocked).exclude(
        join_date__gt=day).select_related("job_category", "subcontractor") \
        .order_by("emp_no").distinct()
    rostered = set(on_day.values_list("employee_id", flat=True))
    # What the gate terminals saw (phase 2, owner 2026-08-24). Proposals are
    # computed from the raw punch log at read time and never stored — the only
    # write path into attendance is still the clerk's save below, with every
    # guard it already carries.
    device = None
    if site.attendance_devices.filter(is_active=True).exists():
        from . import biometric
        device = biometric.day_proposals(site, day)
    # Shift sites: each worker's In/Out defaults to HIS shift's window;
    # workers without one follow the site's normal hours (mixed staff).
    from .shifts import site_shifts, shifts_map
    shifts = site_shifts(site)
    smap = (shifts_map(site, day, [e.id for e in roster])
            if shifts else {})
    rows = []
    for employee in roster:
        att = existing.get(employee.id)
        shift = smap.get(employee.id)
        default_remark = "OFF" if is_rest_day else "PRESENT"
        is_sub = employee.engagement_type == Employee.Engagement.SUBCONTRACT
        rows.append({
            "attendance_id": att.id if att else None,
            "employee_id": employee.id,
            "emp_no": employee.emp_no,
            "full_name": employee.full_name,
            "photo_url": employee.photo.url if employee.photo else None,
            "category": employee.job_category.name
            if employee.job_category else "",
            # Internal split (this form is not client-facing): a subcontract
            # worker attends like everyone else but takes extra hours, not OT.
            "is_subcontract": is_sub,
            "subcontractor": employee.subcontractor.name
            if is_sub and employee.subcontractor_id else "",
            "check_in": att.check_in if att
            else (shift.start if shift else site.working_hours_from),
            "check_out": att.check_out if att
            else (shift.end if shift else site.working_hours_to),
            "shift_id": shift.id if shift else None,
            "shift_name": shift.name if shift else None,
            "ot_requested": att.ot_requested if att else 0,
            "ot_approved": att.ot_approved if att else None,
            "sub_extra_hours": att.sub_extra_hours if att else 0,
            "remark": att.remark if att else default_remark,
            "saved": att is not None,
            # Here only because he carries a mark — no longer on the site's
            # roster for this day. Mark him OFF to take the record back.
            "off_roster": employee.id not in rostered,
            "device": (device["rows"].get(employee.id)
                       if device is not None else None),
        })
    return Response({
        "site": site.code, "date": day.isoformat(),
        "is_rest_day": is_rest_day,
        "locked": _month_locked(site.id, day),
        "has_devices": device is not None,
        "device_unmatched": device["unmatched"] if device else [],
        "shifts": [{"id": s.id, "name": s.name,
                    "start": s.start.strftime("%H:%M"),
                    "end": s.end.strftime("%H:%M"),
                    "overnight": s.overnight} for s in shifts],
        "rows": rows,
    })


@api_view(["GET"])
def attendance_register(request):
    """Whole-month attendance for a site: a per-worker day grid plus totals
    (present / absent / leave / OT hours / Fridays worked) and a site summary.
    Site team + HR/Finance/Admin. Also serves 'as of today' for the current
    month, since days beyond today simply carry no record yet."""
    import calendar

    try:
        site = Site.objects.get(pk=request.GET.get("site"))
        year = int(request.GET.get("year"))
        month = int(request.GET.get("month"))
    except (Site.DoesNotExist, TypeError, ValueError):
        return Response({"detail": "site, year, month required."}, status=400)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not found."}, status=404)

    ndays = calendar.monthrange(year, month)[1]
    work_week = set(site.working_days)
    days = []
    for d in range(1, ndays + 1):
        wd = date(year, month, d).isoweekday()
        days.append({"day": d, "dow": ["Mon", "Tue", "Wed", "Thu", "Fri",
                     "Sat", "Sun"][wd - 1], "rest": wd not in work_week})
    today = date.today()
    today_day = today.day if (today.year == year and today.month == month) \
        else None

    # Who was here THIS MONTH, not who is here today. Rostering by the
    # current allocation hid every leaver from the month they actually
    # worked: Asish Rai has 29 days of July at BVR and vanished off the July
    # register the day his allocation closed, which is how a man with a full
    # month's attendance looked like someone who was never there
    # (owner 2026-08-14). Same rule as the payroll run — see
    # payroll.eligible_workers.
    m_start, m_end = date(year, month, 1), date(year, month, ndays)
    here = EmployeeSiteAllocation.objects.filter(
        site=site, from_date__lte=m_end).filter(
        Q(to_date__isnull=True) | Q(to_date__gte=m_start))
    marked_ids = Attendance.objects.filter(
        site=site, day__year=year, day__month=month).values_list(
        "employee_id", flat=True)
    roster = Employee.objects.filter(
        Q(id__in=here.values_list("employee_id", flat=True))
        | Q(id__in=marked_ids)).select_related(
        "job_category").order_by("emp_no").distinct()
    att = {}
    for a in Attendance.objects.filter(site=site, day__year=year,
                                       day__month=month):
        att[(a.employee_id, a.day.day)] = a

    def code(a, is_rest):
        if a is None:
            return ""
        if a.remark == "PRESENT":
            return "F" if is_rest else "P"
        # PL = away on PAID leave. Only ever seen on the Head Office register:
        # granting leave moves the man there, and the days are pre-marked.
        return {"ABSENT": "A", "SICK": "S", "LEAVE": "L", "PAID_LEAVE": "PL",
                "HALF_DAY": "½"}.get(a.remark, "")

    rest_days = {d["day"] for d in days if d["rest"]}
    rows, sums = [], {"present": 0, "absent": 0, "leave": 0, "sick": 0,
                      "ot_hours": Decimal("0"), "fridays": 0}
    for emp in roster:
        cells, t = {}, {"present": 0, "absent": 0, "leave": 0, "sick": 0,
                        "half": 0, "ot_hours": Decimal("0"), "fridays": 0}
        # A worker is only on the roster from their join date: days before it
        # are outside their engagement and shouldn't be marked (owner
        # 2026-07-31). start_day = 1 for anyone who joined in a prior month.
        start_day = 1
        jd = emp.join_date
        if jd and jd.year == year and jd.month == month:
            start_day = jd.day
        elif jd and (jd.year, jd.month) > (year, month):
            start_day = ndays + 1          # joined after this month — all N/A
        for d in range(1, ndays + 1):
            a = att.get((emp.id, d))
            c = code(a, d in rest_days)
            if c:
                cells[str(d)] = c
            if a is None:
                continue
            t["ot_hours"] += a.ot_approved or 0
            if a.remark in ("PRESENT", "PAID_LEAVE"):
                # A paid leave day is a paid day — it belongs with the days
                # payroll pays, not with the absences it deducts.
                if d in rest_days and a.remark == "PRESENT":
                    t["fridays"] += 1
                else:
                    t["present"] += 1
            elif a.remark == "HALF_DAY":
                t["half"] += 1
            elif a.remark == "ABSENT":
                t["absent"] += 1
            elif a.remark == "SICK":
                t["sick"] += 1
            elif a.remark == "LEAVE":
                t["leave"] += 1
        rows.append({
            "emp_no": emp.emp_no, "full_name": emp.full_name,
            "category": emp.job_category.name if emp.job_category_id else "",
            "start_day": start_day, "days": cells, **t})
        for k in ("present", "absent", "leave", "sick", "ot_hours", "fridays"):
            sums[k] += t[k]
    return Response({
        "site": site.code, "year": year, "month": month,
        "days": days, "today": today_day,
        "locked": TimesheetMonth.objects.filter(
            site=site, year=year, month=month, status="LOCKED").exists(),
        "rows": rows, "totals": sums,
    })


@api_view(["PUT"])
def attendance_bulk(request):
    """Day-grid upsert by Site Admin / SE; late edits audited (spec §6A.2)."""
    if request.user.role not in ("SITE_ADMIN", "SITE_ENGINEER", "PM",
                                 "HO_HR", "DIRECTOR", "ADMIN"):
        return Response({"detail": "Site team or HR records attendance."},
                        status=403)
    try:
        site = Site.objects.get(pk=request.data.get("site"))
        day = date.fromisoformat(request.data.get("date"))
    except (Site.DoesNotExist, TypeError, ValueError):
        return Response({"detail": "site and date required."}, status=400)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not allocated to this site."}, status=403)
    if day > date.today():
        return Response({"detail": "Attendance cannot be entered for future "
                                   "days."}, status=400)
    if _month_locked(site.id, day):
        return Response({"detail": "This month is locked. Ask HO HR to "
                                   "reopen it."}, status=400)

    def parse_time(value):
        if not value:
            return None
        if isinstance(value, str):
            return datetime.strptime(value[:5], "%H:%M").time()
        return value

    late_edit = day < date.today()
    from .shifts import shifts_map
    smap = shifts_map(site, day,
                      [r.get("employee_id") for r in
                       request.data.get("rows", [])])
    saved = 0
    refused = []
    withdrawn = []
    # Per-employee change record. The audit used to say only "this site,
    # this date, N rows" — you could prove someone edited the day but not
    # whose, from what, to what; and an OFF mark DELETED a record silently,
    # uncounted (audit 2026-08-28). Payroll evidence needs names.
    changes = []
    for row in request.data.get("rows", []):
        try:
            employee = Employee.objects.get(pk=row.get("employee_id"))
        except Employee.DoesNotExist:
            continue
        # A deactivated worker used to be skipped in silence, so a clerk
        # trying to correct a leaver's day saw the save succeed and nothing
        # change — "I tried several times to put absence for him. But it
        # failed." (owner 2026-09-02). His history can still be TAKEN BACK;
        # what he cannot get is a new mark.
        if not employee.is_active:
            has_row = Attendance.objects.filter(employee=employee,
                                                day=day).exists()
            if not (has_row and (row.get("remark") or "") == "OFF"):
                refused.append(f"{employee.emp_no} is no longer active — his "
                               f"existing marks can be cleared (mark him OFF), "
                               f"but no new attendance can be recorded")
                continue
        # A worker's engagement starts on their join date — no attendance
        # before it (owner 2026-07-31). Say so rather than dropping the row in
        # silence: the clerk needs to know the mark did not take.
        if employee.join_date and day < employee.join_date:
            refused.append(f"{employee.emp_no} joined {employee.join_date}")
            continue
        # Leave WITHOUT pay is a decision to withhold these days, so a stray
        # PRESENT would quietly undo it. Refused out loud, like a pre-join
        # date, so the clerk knows the mark did not take (owner 2026-08-20).
        unpaid = _leave.open_leave_for(employee, day)
        if unpaid and unpaid.kind == "UNPAID":
            refused.append(f"{employee.emp_no} is on leave without pay "
                           f"({unpaid.from_date} to {unpaid.to_date})")
            continue
        # Another site's day is not this site's to mark — or to clear. Skipped
        # when the row is already this site's, so its own marks stay
        # correctable (owner 2026-09-03; see _allocated_elsewhere).
        if not Attendance.objects.filter(employee=employee, day=day,
                                         site=site).exists():
            held = _allocated_elsewhere(employee, site, day)
            if held:
                refused.append(
                    f"{employee.emp_no} was allocated to {held} on {day}, so "
                    f"{held} marks that day — move his allocation if he was "
                    f"here")
                continue
        remark = row.get("remark") or "PRESENT"
        if remark == "OFF":
            # Rest day, not worked — clear any existing record, create none.
            # A deletion is a change to the pay record and is recorded as one.
            gone = Attendance.objects.filter(employee=employee, day=day).first()
            if gone is not None:
                changes.append({
                    "emp": employee.emp_no, "action": "DELETED",
                    "was": {"remark": gone.remark, "site": gone.site.code,
                            "in": str(gone.check_in or ""),
                            "out": str(gone.check_out or ""),
                            "ot": str(gone.ot_requested or 0)}})
                gone.delete()
            continue
        check_in = parse_time(row.get("check_in"))
        check_out = parse_time(row.get("check_out"))
        is_sub = employee.engagement_type == Employee.Engagement.SUBCONTRACT
        # Subcontract workers never enter the OT request/approval pipeline —
        # their beyond-window time is plain extra hours for the subcontractor.
        defaults = {
            "site": site, "check_in": check_in, "check_out": check_out,
            "remark": remark, "entered_by": request.user,
        }
        if is_sub:
            defaults["sub_extra_hours"] = Decimal(
                str(row.get("sub_extra_hours") or 0))
            defaults["ot_requested"] = Decimal("0")
            defaults["ot_approved"] = None
            defaults["ot_approved_by"] = None
            defaults["ot_approved_at"] = None
        else:
            defaults["ot_requested"] = Decimal(str(row.get("ot_requested") or 0))
            defaults["sub_extra_hours"] = Decimal("0")
        before = Attendance.objects.filter(employee=employee,
                                           day=day).first()
        # The previous site is part of the change when a day moves between
        # sites: that is a transfer of somebody's pay from one register to
        # another and the trail has to name both ends.
        was = ({"remark": before.remark, "in": str(before.check_in or ""),
                "out": str(before.check_out or ""), "site": before.site.code,
                "ot": str(before.ot_requested or 0)} if before else None)
        record, _created = Attendance.objects.update_or_create(
            employee=employee, day=day, defaults=defaults)
        record.normal_hours = _normal_hours(
            site, record.check_in, record.check_out, remark,
            shift=smap.get(employee.id))
        fields = ["normal_hours"]
        # An approval must never outlive the number it was given. The PM
        # approved 6 hours; the clerk later corrected the request to 4 (or,
        # on one BVR row, 47 became 6) — and payroll went on paying the
        # approved figure, because the save wrote ot_requested and left
        # ot_approved exactly where it was. 32 rows in a month across eight
        # sites (owner 2026-09-03, "revising previous day's OT is not
        # getting updated"). A changed request goes back to the PM.
        if (before is not None and not is_sub
                and before.ot_approved is not None
                and (record.ot_requested or Decimal("0"))
                != (before.ot_requested or Decimal("0"))):
            record.ot_approved = None
            record.ot_approved_by = None
            record.ot_approved_at = None
            fields += ["ot_approved", "ot_approved_by", "ot_approved_at"]
            withdrawn.append(employee.emp_no)
        record.save(update_fields=fields)
        now = {"remark": record.remark, "in": str(record.check_in or ""),
               "out": str(record.check_out or ""),
               "ot": str(record.ot_requested or 0)}
        if was != now:
            changes.append({"emp": employee.emp_no,
                            "action": "CREATED" if before is None else "EDITED",
                            **({"was": was} if was else {}), "now": now,
                            **({"ot_approval": "withdrawn"}
                               if employee.emp_no in withdrawn else {})})
        saved += 1
    audit("attendance", site.id, "ATTENDANCE_SAVED", actor=request.user,
          detail={"site": site.code, "date": day.isoformat(), "rows": saved,
                  "late_edit": late_edit,
                  "changed": len(changes), "changes": changes,
                  "ot_approval_withdrawn": withdrawn})
    return Response({"saved": saved, "late_edit": late_edit,
                     "refused": refused, "ot_approval_withdrawn": withdrawn})


@api_view(["GET", "POST"])
def site_shifts_view(request, pk):
    """List / define a site's shifts. Reading is open to whoever can see the
    site; defining takes Admin/Director or the site's own PM."""
    from .shifts import can_manage_shifts

    try:
        site = Site.objects.get(pk=pk)
    except Site.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not found."}, status=404)
    if request.method == "POST":
        if not can_manage_shifts(request.user, site):
            return Response({"detail": "Admin, Director or the site PM "
                                       "defines shifts."}, status=403)
        from .models import SiteShift
        name = (request.data.get("name") or "").strip()
        if not name:
            return Response({"detail": "Give the shift a name."}, status=400)
        try:
            start = datetime.strptime(
                str(request.data.get("start"))[:5], "%H:%M").time()
            end = datetime.strptime(
                str(request.data.get("end"))[:5], "%H:%M").time()
        except (TypeError, ValueError):
            return Response({"detail": "Start and end times are required "
                                       "(HH:MM)."}, status=400)
        ot_raw = request.data.get("ot_counts_from")
        ot_from = (datetime.strptime(str(ot_raw)[:5], "%H:%M").time()
                   if ot_raw else None)
        SiteShift.objects.create(
            site=site, name=name, start=start, end=end,
            ot_counts_from=ot_from)
        audit("site", site.id, "SHIFT_DEFINED", actor=request.user,
              detail={"shift": name, "start": str(start), "end": str(end)})
    return Response([{
        "id": s.id, "name": s.name,
        "start": s.start.strftime("%H:%M"), "end": s.end.strftime("%H:%M"),
        "ot_counts_from": (s.ot_counts_from.strftime("%H:%M")
                           if s.ot_counts_from else None),
        "overnight": s.overnight, "is_active": s.is_active,
        "workers": s.assignments.filter(to_date__isnull=True).count(),
    } for s in site.shifts.all()])


@api_view(["PATCH"])
def shift_update(request, pk):
    """Rename / retime / retire one shift (same roles as defining)."""
    from .models import SiteShift
    from .shifts import can_manage_shifts

    try:
        shift = SiteShift.objects.select_related("site").get(pk=pk)
    except SiteShift.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if not can_manage_shifts(request.user, shift.site):
        return Response({"detail": "Admin, Director or the site PM edits "
                                   "shifts."}, status=403)
    fields = []
    if "name" in request.data:
        shift.name = (request.data["name"] or "").strip() or shift.name
        fields.append("name")
    for key in ("start", "end", "ot_counts_from"):
        if key in request.data:
            raw = request.data[key]
            try:
                value = (datetime.strptime(str(raw)[:5], "%H:%M").time()
                         if raw else None)
            except ValueError:
                return Response({"detail": f"Bad time for {key}."},
                                status=400)
            if key != "ot_counts_from" and value is None:
                return Response({"detail": f"{key} cannot be blank."},
                                status=400)
            setattr(shift, key, value)
            fields.append(key)
    if "is_active" in request.data:
        shift.is_active = bool(request.data["is_active"])
        fields.append("is_active")
    if fields:
        shift.save(update_fields=fields)
        audit("site", shift.site_id, "SHIFT_UPDATED", actor=request.user,
              detail={"shift": shift.name, "fields": fields})
    return Response({"ok": True})


@api_view(["POST"])
def shift_assign(request):
    """Put workers on a shift (or back on normal hours with shift_id null),
    effective from the given day. Same roles as the day grid."""
    from .models import SiteShift
    from . import shifts as shift_svc

    if request.user.role not in ("SITE_ADMIN", "SITE_ENGINEER", "PM",
                                 "HO_HR", "DIRECTOR", "ADMIN"):
        return Response({"detail": "Site team or HR assigns shifts."},
                        status=403)
    try:
        site = Site.objects.get(pk=request.data.get("site"))
        day = date.fromisoformat(request.data.get("date"))
    except (Site.DoesNotExist, TypeError, ValueError):
        return Response({"detail": "site and date required."}, status=400)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not allocated to this site."}, status=403)
    shift = None
    if request.data.get("shift_id"):
        shift = SiteShift.objects.filter(pk=request.data["shift_id"],
                                         site=site, is_active=True).first()
        if shift is None:
            return Response({"detail": "No such shift on this site."},
                            status=400)
    ids = request.data.get("employee_ids") or []
    moved = 0
    for employee in Employee.objects.filter(id__in=ids, is_active=True):
        shift_svc.assign(employee, shift, day)
        moved += 1
    audit("site", site.id, "SHIFT_ASSIGNED", actor=request.user,
          detail={"shift": shift.name if shift else None,
                  "employees": moved, "from": day.isoformat()})
    return Response({"assigned": moved})


@api_view(["POST"])
def ot_approve(request):
    """PM approves OT per day or in batch; unapproved OT can never flow
    into payroll (spec §6A.2)."""
    # Either the legacy shape — ids + one optional hours override — or a
    # per-row decision: rows=[{id, hours}]. The PM reviews each man's hours
    # against their cost; a single button over 200 rows is how OT got
    # abused (owner 2026-09-03).
    per_row = {}
    for r in request.data.get("rows") or []:
        try:
            per_row[int(r.get("id"))] = Decimal(str(r.get("hours")))
        except (TypeError, ValueError, ArithmeticError):
            return Response({"detail": "Each row needs an id and hours."},
                            status=400)
    ids = list(per_row) or (request.data.get("ids") or [])
    rows = Attendance.objects.filter(pk__in=ids).select_related(
        "site", "employee__job_category")
    if not rows:
        return Response({"detail": "ids required."}, status=400)
    for row in rows:
        pm = row.site.current_pm()
        if not (request.user.role in ("ADMIN", "HO_HR", "PA") or
                (request.user.role == "PM" and pm and pm.id == request.user.id)):
            return Response({"detail": f"Only the site PM or HR approves OT "
                                       f"({row.site.code})."}, status=403)
        if _month_locked(row.site_id, row.day):
            return Response({"detail": "Month is locked."}, status=400)
    hours_override = request.data.get("hours")
    decided, total_cost = [], Decimal("0")
    for row in rows:
        if row.id in per_row:
            hours = per_row[row.id]
        elif hours_override is not None:
            hours = Decimal(str(hours_override))
        else:
            hours = row.ot_requested or Decimal("0")
        if hours < 0:
            return Response({"detail": "Hours cannot be negative."},
                            status=400)
        row.ot_approved = hours
        row.ot_approved_by = request.user
        row.ot_approved_at = timezone.now()
        row.save(update_fields=["ot_approved", "ot_approved_by",
                                "ot_approved_at"])
        rate = row.employee.ot_rate()
        cost = (hours * rate).quantize(Decimal("0.01"))
        total_cost += cost
        decided.append({"emp": row.employee.emp_no, "day": row.day.isoformat(),
                        "requested": str(row.ot_requested or 0),
                        "approved": str(hours), "rate": str(rate),
                        "cost": str(cost)})
    # The audit used to say only "count: 197". Payroll evidence needs the
    # men, the hours and what they cost.
    audit("attendance", rows[0].site_id, "OT_APPROVED", actor=request.user,
          detail={"count": len(rows), "total_cost": str(total_cost),
                  "rows": decided})
    return Response({"approved": len(rows), "total_cost": total_cost})


@api_view(["GET"])
def attendance_range_pdf(request, site_id):
    """The client's attendance record for a date range — headcount and
    marks, no overtime (owner 2026-09-03: housekeeping and food)."""
    from django.template.loader import render_to_string

    from . import attendance_report, pdf as pdf_mod
    from .views_payroll import _pdf_response
    if request.user.role not in ("SITE_ADMIN", "SITE_ENGINEER", "PM", "QS",
                                 "HO_HR", "FINANCE", "DIRECTOR", "PA",
                                 "ADMIN"):
        return Response({"detail": "Not allowed."}, status=403)
    try:
        site = Site.objects.get(pk=site_id)
    except Site.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not found."}, status=404)
    try:
        start = date.fromisoformat(request.GET.get("from") or "")
        end = date.fromisoformat(request.GET.get("to") or "")
    except ValueError:
        return Response({"detail": "from and to are required, YYYY-MM-DD."},
                        status=400)
    try:
        ctx = attendance_report.build(site, start, end)
    except ValueError as exc:
        return Response({"detail": str(exc)}, status=400)
    ctx["logo_src"] = pdf_mod.logo_src()
    ctx["co"] = pdf_mod.company_info()
    ctx["subline"] = (f"{site.code}  |  {start:%d %b %Y} – {end:%d %b %Y}"
                      "  |  for the client")
    html = render_to_string("pdf/attendance_range.html", ctx)
    return _pdf_response(html, f"{site.code}-attendance-{start}-to-{end}.pdf")


def ot_pending_summary(site, days_back=31):
    """What is waiting for the PM's OT decision at a site: rows, days, hours
    and cost, over the last month of unlocked days. The PM used to learn of
    pending OT only by opening the right day — and a revised request that
    had gone back to "awaiting" could sit unnoticed (owner 2026-09-03,
    "how does pm see if there is any pending ot approval?")."""
    from .models import TimesheetMonth
    since = date.today() - timedelta(days=days_back)
    locked = set(TimesheetMonth.objects.filter(
        site=site, status="LOCKED").values_list("year", "month"))
    qs = (Attendance.objects.filter(site=site, day__gte=since,
                                    day__lte=date.today(),
                                    ot_requested__gt=0, ot_approved__isnull=True)
          .exclude(employee__engagement_type=Employee.Engagement.SUBCONTRACT)
          .select_related("employee__job_category"))
    rows, days, hours, cost = 0, set(), Decimal("0"), {}
    for a in qs:
        if (a.day.year, a.day.month) in locked:
            continue
        rows += 1
        days.add(a.day)
        hours += a.ot_requested
        ccy = a.employee.currency or "MVR"
        cost[ccy] = cost.get(ccy, Decimal("0")) + a.ot_requested * a.employee.ot_rate()
    return {
        "rows": rows, "days": len(days),
        "oldest_day": min(days).isoformat() if days else None,
        "newest_day": max(days).isoformat() if days else None,
        "hours": hours,
        "cost": [{"currency": c, "amount": v.quantize(Decimal("0.01"))}
                 for c, v in sorted(cost.items())],
    }


def _ot_flag_hours():
    """Requested hours above which a row is highlighted for the PM. A
    company parameter, not a rule — it draws the eye, it decides nothing."""
    from .models import CompanyParameter
    try:
        v = (CompanyParameter.objects.get(key="ot_flag_hours").value
             or "").strip()
        return Decimal(v) if v else Decimal("4")
    except (CompanyParameter.DoesNotExist, ArithmeticError, ValueError):
        return Decimal("4")


@api_view(["GET"])
def ot_review(request):
    """The PM's OT approval table for a day: every request with its rate and
    what it costs, the day's totals, and the month so far.

    Approving used to be one button over the whole day — 170 to 200 rows at
    a click, 455 times in a month — and the PM saw hours, never money.
    A man's 6 hours is a number; 6 hours at MVR 45 beside a month already at
    MVR 60,000 is a decision (owner 2026-09-03).
    """
    if request.user.role not in ("PM", "HO_HR", "ADMIN", "DIRECTOR", "PA",
                                 "FINANCE", "SITE_ENGINEER", "SITE_ADMIN"):
        return Response({"detail": "Not allowed."}, status=403)
    try:
        site = Site.objects.get(pk=request.GET.get("site"))
        day = date.fromisoformat(request.GET.get("date"))
    except (Site.DoesNotExist, TypeError, ValueError):
        return Response({"detail": "site and date required."}, status=400)
    if not _site_scope_ok(request, site):
        return Response({"detail": "Not found."}, status=404)
    from django.db.models import Q
    flag = _ot_flag_hours()
    rows, totals = [], {}

    def bucket(ccy):
        return totals.setdefault(ccy, {
            "currency": ccy, "requested_hours": Decimal("0"),
            "requested_cost": Decimal("0"), "approved_hours": Decimal("0"),
            "approved_cost": Decimal("0"), "pending_rows": 0})

    qs = (Attendance.objects.filter(site=site, day=day)
          .exclude(employee__engagement_type=Employee.Engagement.SUBCONTRACT)
          .filter(Q(ot_requested__gt=0) | Q(ot_approved__isnull=False))
          .select_related("employee__job_category")
          .order_by("employee__job_category__name", "employee__emp_no"))
    for a in qs:
        e = a.employee
        rate = e.ot_rate()
        req = a.ot_requested or Decimal("0")
        appr = a.ot_approved
        b = bucket(e.currency or "MVR")
        b["requested_hours"] += req
        b["requested_cost"] += req * rate
        if appr is None:
            b["pending_rows"] += 1
        else:
            b["approved_hours"] += appr
            b["approved_cost"] += appr * rate
        rows.append({
            "attendance_id": a.id, "employee_id": e.id, "emp_no": e.emp_no,
            "full_name": e.full_name, "photo_url": (e.photo.url if e.photo
                                                    else None),
            "category": e.job_category.name if e.job_category_id else "",
            "check_in": a.check_in, "check_out": a.check_out,
            "normal_hours": a.normal_hours,
            "ot_requested": req, "ot_approved": appr,
            "ot_rate": rate, "currency": e.currency or "MVR",
            "cost_requested": (req * rate).quantize(Decimal("0.01")),
            "cost_approved": ((appr * rate).quantize(Decimal("0.01"))
                              if appr is not None else None),
            "pending": appr is None,
            "no_rate": rate == 0,
            "flag": req > flag,
        })
    # The month so far: what this site has already committed to in OT.
    mtd = {}
    for a in (Attendance.objects.filter(site=site, day__year=day.year,
                                        day__month=day.month,
                                        ot_approved__isnull=False)
              .exclude(employee__engagement_type=Employee.Engagement.SUBCONTRACT)
              .select_related("employee__job_category")):
        ccy = a.employee.currency or "MVR"
        m = mtd.setdefault(ccy, {"currency": ccy, "hours": Decimal("0"),
                                 "cost": Decimal("0")})
        m["hours"] += a.ot_approved
        m["cost"] += a.ot_approved * a.employee.ot_rate()
    for b in totals.values():
        for k in ("requested_cost", "approved_cost"):
            b[k] = b[k].quantize(Decimal("0.01"))
    for m in mtd.values():
        m["cost"] = m["cost"].quantize(Decimal("0.01"))
    return Response({
        "site": site.code, "date": day.isoformat(),
        "locked": _month_locked(site.id, day),
        "flag_hours": flag, "rows": rows,
        "totals": list(totals.values()), "month_to_date": list(mtd.values()),
    })


# ===== Overtime rate master (owner: managed, not hardcoded) =====


@api_view(["GET", "POST"])
def overtime_rates(request):
    """GET: every DPR job category with its MVR/USD OT rate (if set) so the
    management page can show and fill them. POST: upsert one category+currency
    rate. HR, Finance and Admin — the OT rate is a payroll input, and Finance
    runs payroll alongside HR (owner 2026-08-19)."""
    if request.method == "POST":
        if request.user.role not in (*HR_ROLES, "FINANCE"):
            return Response({"detail": "HR, Finance or Admin manage OT rates."},
                            status=403)
        try:
            cat = ManpowerCategory.objects.get(pk=request.data.get("category_id"))
        except ManpowerCategory.DoesNotExist:
            return Response({"detail": "Unknown category."}, status=400)
        currency = request.data.get("currency", "MVR")
        try:
            rate = Decimal(str(request.data.get("rate_per_hour") or 0))
        except (TypeError, ValueError):
            return Response({"detail": "Rate is invalid."}, status=400)
        row, _ = OvertimeRate.objects.update_or_create(
            category=cat, currency=currency,
            defaults={"rate_per_hour": rate,
                      "applies_by_default": bool(
                          request.data.get("applies_by_default", True))})
        audit("overtime_rate", row.id, "OT_RATE_SET", actor=request.user,
              detail={"category": cat.name, "currency": currency,
                      "rate": str(rate)})
        return Response(_ot_rate_info(row), status=200)

    cats = ManpowerCategory.objects.filter(
        list_type="DPR", is_active=True).order_by("grp", "sort_order")
    rates = {(r.category_id, r.currency): r
             for r in OvertimeRate.objects.all()}
    out = []
    for cat in cats:
        row = {"category_id": cat.id, "category_name": cat.name,
               "grp": cat.grp, "rates": {}}
        for cur in ("MVR", "USD"):
            r = rates.get((cat.id, cur))
            row["rates"][cur] = {
                "rate_per_hour": r.rate_per_hour if r else None,
                "applies_by_default": r.applies_by_default if r else True,
            } if r else None
        out.append(row)
    return Response(out)


def _ot_rate_info(r):
    return {"id": r.id, "category_id": r.category_id, "currency": r.currency,
            "rate_per_hour": r.rate_per_hour,
            "applies_by_default": r.applies_by_default}


# ===== Month close & payroll (spec §6A.3) =====


@api_view(["POST"])
def timesheet_lock(request, site_id, year, month):
    try:
        site = Site.objects.get(pk=site_id)
    except Site.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    pm = site.current_pm()
    # HR can sign off any month (needed for Head Office, which has no PM, and
    # for corrections); otherwise the site PM signs off (spec §6A.3).
    if not (request.user.role in ("ADMIN", "HO_HR", "PA") or
            (request.user.role == "PM" and pm and pm.id == request.user.id)):
        return Response({"detail": "The site PM or HR signs off the month."},
                        status=403)
    row, _ = TimesheetMonth.objects.get_or_create(site=site, year=year,
                                                  month=month)
    if row.status == "LOCKED":
        return Response({"detail": "Already locked."}, status=400)
    row.status = "LOCKED"
    row.signed_off_by = request.user
    row.signed_off_at = timezone.now()
    row.save()
    # Staff cost is Incurred at month lock (§6C.3.5) — one Labour & Staff
    # posting per site for the period
    from . import staff_cost

    staff_cost.post_staff_cost(site, year, month, request.user)
    audit("timesheet", row.id, "TIMESHEET_LOCKED", actor=request.user,
          detail={"site": site.code, "period": f"{year}-{month:02d}"})
    return Response({"status": "LOCKED",
                     "signed_off_by": request.user.full_name})


@api_view(["POST"])
def timesheet_reopen(request, site_id, year, month):
    """Reopen a locked month with a reason — audited (spec §6A.3). Whoever may
    lock may unlock: the site PM (accidental locks) as well as HO HR / Admin
    (owner 2026-07-14)."""
    try:
        site = Site.objects.get(pk=site_id)
    except Site.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    pm = site.current_pm()
    if not (request.user.role in ("ADMIN", "HO_HR", "PA") or
            (request.user.role == "PM" and pm and pm.id == request.user.id)):
        return Response({"detail": "The site PM or HR reopens a month."},
                        status=403)
    reason = (request.data.get("reason") or "").strip()
    if not reason:
        return Response({"detail": "A reason is required."}, status=400)
    try:
        row = TimesheetMonth.objects.get(site_id=site_id, year=year,
                                         month=month, status="LOCKED")
    except TimesheetMonth.DoesNotExist:
        return Response({"detail": "Month is not locked."}, status=400)
    row.status = "OPEN"
    row.reopened_by = request.user
    row.reopen_reason = reason
    row.save()
    # Reverse the month's staff cost so it can be recomputed at the next lock
    from . import staff_cost

    staff_cost.reverse_staff_cost(row.site, year, month, request.user)
    audit("timesheet", row.id, "TIMESHEET_REOPENED", actor=request.user,
          detail={"reason": reason, "period": f"{year}-{month:02d}"})
    return Response({"status": "OPEN"})


def _param_decimal(key, default):
    try:
        return Decimal(str(CompanyParameter.objects.get(key=key).value))
    except CompanyParameter.DoesNotExist:
        return Decimal(str(default))


@api_view(["GET"])
def payroll_export(request, year, month):
    """Per employee: days worked, absences, hours, approved OT, computed
    gross = basic + OT x hourly x multiplier. HR/Finance/Admin (R3)."""
    if not _sees_pay(request.user):
        return Response({"detail": "HO HR/Payroll or Finance only."},
                        status=403)
    multiplier = _param_decimal("ot_multiplier", 1.25)
    divisor = _param_decimal("hourly_rate_divisor", 240)

    qs = Attendance.objects.filter(day__year=year, day__month=month) \
        .select_related("employee", "site")
    if request.GET.get("site"):
        qs = qs.filter(site_id=request.GET["site"])

    by_employee = {}
    for att in qs:
        entry = by_employee.setdefault(att.employee_id, {
            "employee": att.employee, "site_codes": set(),
            "days_worked": 0, "absences": 0,
            "normal_hours": Decimal("0"), "ot_hours": Decimal("0"),
        })
        entry["site_codes"].add(att.site.code)
        if att.remark in ("ABSENT", "SICK", "LEAVE"):
            entry["absences"] += 1
        elif att.remark == "HALF_DAY":
            entry["days_worked"] += 0.5
        else:
            entry["days_worked"] += 1
        entry["normal_hours"] += att.normal_hours or 0
        entry["ot_hours"] += att.ot_approved or 0  # approved OT ONLY

    rows = []
    for entry in sorted(by_employee.values(),
                        key=lambda e: e["employee"].emp_no):
        employee = entry["employee"]
        basic = employee.basic_pay or Decimal("0")
        hourly = (basic / divisor).quantize(Decimal("0.01")) if divisor else 0
        ot_amount = (entry["ot_hours"] * hourly * multiplier) \
            .quantize(Decimal("0.01"))
        rows.append({
            "emp_no": employee.emp_no, "full_name": employee.full_name,
            "sites": "/".join(sorted(entry["site_codes"])),
            "days_worked": entry["days_worked"],
            "absences": entry["absences"],
            "normal_hours": entry["normal_hours"],
            "ot_hours_approved": entry["ot_hours"],
            "basic_pay": basic, "hourly_rate": hourly,
            "ot_amount": ot_amount, "gross": basic + ot_amount,
        })

    # NB: "format" is reserved by DRF content negotiation — use "export"
    if request.GET.get("export") == "xlsx":
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = f"Payroll {year}-{month:02d}"
        headers = ["Emp No", "Name", "Site(s)", "Days Worked", "Absences",
                   "Normal Hours", "Approved OT (h)", "Basic Pay (MVR)",
                   "Hourly Rate", "OT Amount", "Gross (MVR)"]
        ws.append(headers)
        for row in rows:
            ws.append([row["emp_no"], row["full_name"], row["sites"],
                       row["days_worked"], row["absences"],
                       float(row["normal_hours"]),
                       float(row["ot_hours_approved"]),
                       float(row["basic_pay"]), float(row["hourly_rate"]),
                       float(row["ot_amount"]), float(row["gross"])])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet")
        response["Content-Disposition"] = \
            f'attachment; filename="payroll-{year}-{month:02d}.xlsx"'
        wb.save(response)
        return response

    return Response({
        "period": f"{year}-{month:02d}",
        "ot_multiplier": multiplier, "hourly_rate_divisor": divisor,
        "rows": rows,
    })


@api_view(["GET"])
def dashboard_hr(request):
    """HR/Payroll dashboard (spec §7.4): month-lock board, permit-expiry
    and reallocation alerts, workforce today and OT summaries."""
    if request.user.role not in PAYROLL_ROLES:
        return Response({"detail": "HR/Finance/Admin only."}, status=403)
    today = date.today()
    sites = Site.objects.exclude(is_head_office=True).order_by("code")
    locks = {
        (t.site_id): t for t in TimesheetMonth.objects.filter(
            year=today.year, month=today.month)
    }
    active_sites = [s for s in sites if s.status == "ACTIVE"]
    board = [{
        "site_id": s.id, "code": s.code, "name": s.name,
        "status": locks[s.id].status if s.id in locks else "OPEN",
        "signed_off_at": locks[s.id].signed_off_at if s.id in locks else None,
    } for s in active_sites]
    all_locked = bool(board) and all(b["status"] == "LOCKED" for b in board)

    from . import permits
    expiring = permits.alerts()[:30]  # permanent workers, permit ≤30 days

    closed_ids = sites.filter(status="CLOSED").values_list("id", flat=True)
    stranded = list(EmployeeSiteAllocation.objects.filter(
        to_date__isnull=True, site_id__in=closed_ids,
        employee__is_active=True,
    ).select_related("employee", "site").values(
        "employee__emp_no", "employee__full_name", "site__code")[:30])

    todays = Attendance.objects.filter(day=today)
    present = todays.exclude(remark__in=("ABSENT", "LEAVE", "SICK")).count()
    ot_pending = Attendance.objects.filter(
        ot_requested__gt=0, ot_approved__isnull=True,
        employee__engagement_type="DIRECT").count()

    return Response({
        "month": f"{today.year}-{today.month:02d}",
        "lock_board": board,
        "all_locked": all_locked,
        "permit_expiries": expiring,
        "reallocation_alerts": stranded,
        "workforce_today": present,
        "ot_pending_approval": ot_pending,
        "active_employees": Employee.objects.hr_managed().filter(
            is_active=True).count(),
    })


def site_manpower_data(site, day=None):
    """Roster vs attendance per manpower category for one site (R9):
    the employee DB says who is stationed here; today's attendance says
    who actually turned up."""
    day = day or date.today()
    allocations = EmployeeSiteAllocation.objects.filter(
        site=site, to_date__isnull=True, employee__is_active=True
    ).select_related("employee__job_category")
    cats = {}

    def bucket(category):
        key = category.id if category else 0
        if key not in cats:
            cats[key] = {"id": key,
                         "name": category.name if category else "Uncategorised",
                         "grp": category.grp if category else "",
                         "roster": 0, "present": 0, "absent": 0}
        return cats[key]

    emp_ids = []
    for a in allocations:
        bucket(a.employee.job_category)["roster"] += 1
        emp_ids.append(a.employee_id)
    todays = Attendance.objects.filter(
        site=site, day=day).select_related("employee__job_category")
    for att in todays:
        b = bucket(att.employee.job_category)
        if att.remark in ("ABSENT", "SICK", "LEAVE"):
            b["absent"] += 1
        else:
            b["present"] += 1
    rows = sorted(cats.values(), key=lambda c: -c["roster"])
    return {
        "attendance_entered": todays.exists(),
        "roster_total": len(emp_ids),
        "present": sum(c["present"] for c in rows),
        "absent": sum(c["absent"] for c in rows),
        "categories": rows,
    }


@api_view(["GET"])
def site_manpower(request, site_id):
    """Full manpower breakdown for the site page (R9 'more data' view):
    every category plus the roster with today's status. Site users see
    names and categories only — never pay or passports."""
    try:
        site = Site.objects.get(pk=site_id)
    except Site.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None and site.id not in site_ids:
        return Response({"detail": "Not found."}, status=404)
    today = date.today()
    data = site_manpower_data(site, today)
    status_by_emp = {
        a.employee_id: a.remark
        for a in Attendance.objects.filter(site=site, day=today)
    }
    employees = [{
        "emp_no": a.employee.emp_no,
        "full_name": a.employee.full_name,
        "category": a.employee.job_category.name
        if a.employee.job_category else "—",
        "today": status_by_emp.get(a.employee_id, "NOT RECORDED"),
    } for a in EmployeeSiteAllocation.objects.filter(
        site=site, to_date__isnull=True, employee__is_active=True,
    ).select_related("employee__job_category")
        .order_by("employee__emp_no")]
    data["employees"] = employees
    data["date"] = today
    data["site"] = site.code
    return Response(data)


@api_view(["GET", "POST"])
def employee_merge(request):
    """Fold a duplicate employee record into the one that keeps the history.

    GET  ?keeper=EMP-0020&duplicate=EMP-0603      -> the plan, writes nothing
    POST {keeper, duplicate, to_site, from_date}  -> merge, then optionally
                                                     re-site the days

    Admin only, and a plan first: it moves work history and there is no undo
    (owner 2026-08-15).
    """
    from . import merge_employees

    if request.user.role != User.Role.ADMIN:
        return Response({"detail": "Admin only."}, status=403)
    src = request.GET if request.method == "GET" else request.data
    try:
        keeper = Employee.objects.get(emp_no=src.get("keeper"))
        dup = Employee.objects.get(emp_no=src.get("duplicate"))
    except Employee.DoesNotExist:
        return Response({"detail": "Give the emp_no of both records."},
                        status=400)
    if request.method == "GET":
        return Response(merge_employees.plan(keeper, dup))

    res, err = merge_employees.merge(keeper, dup, request.user)
    if err:
        return Response({"detail": err}, status=400)
    site_code, from_date = src.get("to_site"), src.get("from_date")
    if site_code and from_date:
        site = Site.objects.filter(code=site_code).first()
        if not site:
            return Response({"detail": f"No site {site_code}."}, status=400)
        try:
            moved_on = date.fromisoformat(from_date)
        except (TypeError, ValueError):
            return Response({"detail": "from_date must be YYYY-MM-DD."},
                            status=400)
        res["resited"] = merge_employees.transfer_from(
            keeper, site, moved_on, request.user)
    return Response(res)


@api_view(["GET"])
def duplicate_passports(request):
    """Every passport number on more than one employee record.

    38 of them when this was written, about half the same man twice and the
    rest a mistyped number shared with somebody else entirely — the two need
    opposite fixes, so the list says which records hold the work history
    (owner 2026-08-15).
    """
    if request.user.role not in (User.Role.ADMIN, User.Role.HO_HR):
        return Response({"detail": "HR or Admin only."}, status=403)
    groups = {}
    for e in Employee.objects.exclude(passport_no="").exclude(
            passport_no=None).order_by("emp_no"):
        groups.setdefault(e.passport_no.strip().upper(), []).append(e)
    out = []
    for pno, emps in groups.items():
        if len(emps) < 2:
            continue
        out.append({
            "passport_no": pno,
            "records": [{
                "emp_no": e.emp_no, "full_name": e.full_name,
                "is_active": e.is_active,
                "join_date": e.join_date.isoformat() if e.join_date else None,
                "basic_pay": e.basic_pay,
                "attendance_rows": Attendance.objects.filter(
                    employee=e).count(),
                "site": (e.current_site_id() and Site.objects.filter(
                    pk=e.current_site_id()).values_list("code", flat=True)
                    .first()) or None,
            } for e in emps],
        })
    out.sort(key=lambda g: -sum(r["attendance_rows"] for r in g["records"]))
    return Response({"count": len(out), "groups": out})
