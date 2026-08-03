"""Procurement Schedule Phase 6 — client xlsx export.

The export is a client-facing allowlist: qty/unit + a required-on-site date and
an overall Status are added over the SFR sheet; internal money and the supplier
name never appear; source country stays.
"""
import io
from datetime import date

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from .models import (Project, ScheduleLine, Site, SitePmHistory,
                     User)
from .tests import make_user


class ProcurementExportTests(TestCase):
    def setUp(self):
        self.site = Site.objects.create(code="SJR", name="Jani",
                                        status=Site.Status.ACTIVE,
                                        client_name="Soneva Fushi")
        self.pm = make_user("px_pm", User.Role.PM, site=self.site)
        SitePmHistory.objects.create(site=self.site, pm_user=self.pm,
                                     from_date=date(2026, 1, 1))
        self.project = Project.objects.create(
            site=self.site, code="SJR-01", title="Villa Upgrades", pm=self.pm)
        self.purch = make_user("px_buy", User.Role.HO_PURCHASING)
        self.director = make_user("px_dir", User.Role.DIRECTOR)
        self.client = APIClient()

    def _signed_schedule(self, lines):
        """Propose→confirm→sign-off a schedule with the given line payloads."""
        self.client.force_authenticate(self.pm)
        self.pk = self.client.post(
            f"/api/v1/projects/{self.project.id}/procurement-schedule").data["id"]
        for ln in lines:
            self.client.post(f"/api/v1/procurement-schedules/{self.pk}/lines",
                             ln, format="json")
        self.client.post(f"/api/v1/procurement-schedules/{self.pk}/submit")
        self.client.force_authenticate(self.purch)
        # give each line its commercial (internal) fields
        for line in ScheduleLine.objects.filter(schedule__document_id=self.pk):
            self.client.patch(
                f"/api/v1/procurement-schedule-lines/{line.id}",
                {"estimated_value": "5000", "planned_supplier": "Dixon Seals",
                 "source_country": "UK"}, format="json")
        self.client.post(f"/api/v1/procurement-schedules/{self.pk}/action",
                         {"action": "confirm"}, format="json")
        self.client.force_authenticate(self.director)
        self.client.post(f"/api/v1/procurement-schedules/{self.pk}/action",
                         {"action": "sign_off"}, format="json")

    def _export(self, user):
        self.client.force_authenticate(user)
        r = self.client.get(
            f"/api/v1/procurement-schedules/{self.pk}/export")
        self.assertEqual(r.status_code, 200, getattr(r, "data", None))
        self.assertIn("spreadsheetml", r["Content-Type"])
        return load_workbook(io.BytesIO(r.content))

    def test_export_layout_and_allowlist(self):
        self._signed_schedule([
            {"description": "Door seals", "section_code": "A",
             "section_title": "Villa Upgrades", "quantity": "12", "uom": "nos",
             "required_date": "2026-09-01", "tds_required": True,
             "supply_by": "CONTRACTOR"},
            {"description": "Client tiles", "section_code": "A",
             "supply_by": "CLIENT", "quantity": "50", "uom": "m2"},
        ])
        ws = self._export(self.director).active

        # header block (labels + values in merged ranges)
        self.assertEqual(ws["A1"].value, "Project:")
        self.assertIn("SJR-01", str(ws["C1"].value))
        self.assertEqual(ws["A2"].value, "Sub:")
        self.assertEqual(ws["H2"].value, "Client:")
        self.assertEqual(ws["J2"].value, "Soneva Fushi")

        # column headers (rows 3-4)
        headers = [ws.cell(row=3, column=c).value
                   for c in range(1, ws.max_column + 1)]
        self.assertIn("Qty", headers)            # added over SFR
        self.assertIn("Required On Site", headers)
        self.assertIn("Source Country", headers)
        self.assertIn("Status", headers)
        self.assertIn("Remarks", headers)

        # gather all cell text
        allcells = "\n".join(
            str(ws.cell(row=r, column=c).value)
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None)

        # allowlist: no internal money or supplier name
        self.assertNotIn("Dixon Seals", allcells)   # supplier hidden
        self.assertNotIn("5000", allcells)           # estimate hidden
        # source country IS shown
        self.assertIn("UK", allcells)
        # supply-by relabelled
        self.assertIn("Sand Planet", allcells)
        self.assertIn("Client", allcells)
        # section band + both items present
        self.assertIn("Door seals", allcells)
        self.assertIn("Client tiles", allcells)

    def test_share_link_public_page_and_allowlist(self):
        self._signed_schedule([
            {"description": "Door seals", "section_code": "A",
             "section_title": "Villa Upgrades", "quantity": "12", "uom": "nos",
             "required_date": "2026-09-01", "supply_by": "CONTRACTOR"}])
        # no link before it's minted
        self.client.force_authenticate(self.director)
        d = self.client.get(
            f"/api/v1/procurement-schedules/{self.pk}").data
        self.assertEqual(d["share"]["path"], "")
        self.assertTrue(d["share"]["can_share"])

        # mint the link
        d = self.client.post(
            f"/api/v1/procurement-schedules/{self.pk}/share").data
        path = d["share"]["path"]
        self.assertTrue(path.startswith("/share/procurement/"))
        token = path.rsplit("/", 1)[-1]
        self.assertGreaterEqual(len(token), 20)

        # the public page loads with NO login and carries only allowlist data
        anon = APIClient()
        r = anon.get(path)
        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("Door seals", html)
        self.assertIn("Sand Planet", html)
        self.assertNotIn("Dixon Seals", html)   # supplier hidden
        self.assertNotIn("5000", html)          # estimate hidden
        self.assertIn("UK", html)               # source country shown
        self.assertIn(f"{path}/plan.xlsx", html)  # download button

        # the same link serves the full spreadsheet, no login
        x = anon.get(f"{path}/plan.xlsx")
        self.assertEqual(x.status_code, 200)
        self.assertIn("spreadsheetml", x["Content-Type"])
        xhtml = load_workbook(io.BytesIO(x.content)).active
        allx = "\n".join(str(xhtml.cell(row=rr, column=cc).value)
                         for rr in range(1, xhtml.max_row + 1)
                         for cc in range(1, xhtml.max_column + 1)
                         if xhtml.cell(row=rr, column=cc).value is not None)
        self.assertIn("Door seals", allx)
        self.assertNotIn("Dixon Seals", allx)   # allowlist holds on the xlsx too

        # rotating the token revokes the old URL
        self.client.force_authenticate(self.director)
        d2 = self.client.post(
            f"/api/v1/procurement-schedules/{self.pk}/share").data
        self.assertNotEqual(d2["share"]["path"], path)
        self.assertEqual(anon.get(path).status_code, 404)   # old link dead

        # revoke clears it
        d3 = self.client.delete(
            f"/api/v1/procurement-schedules/{self.pk}/share").data
        self.assertEqual(d3["share"]["path"], "")
        self.assertEqual(anon.get(d2["share"]["path"]).status_code, 404)

    def test_share_bad_token_404(self):
        anon = APIClient()
        self.assertEqual(
            anon.get("/share/procurement/nope-not-a-token").status_code, 404)

    def test_site_engineer_cannot_share(self):
        se = make_user("px_se2", User.Role.SITE_ENGINEER, site=self.site)
        self._signed_schedule([
            {"description": "Glass", "section_code": "A"}])
        self.client.force_authenticate(se)
        d = self.client.get(
            f"/api/v1/procurement-schedules/{self.pk}").data
        self.assertFalse(d["share"]["can_share"])
        r = self.client.post(f"/api/v1/procurement-schedules/{self.pk}/share")
        self.assertEqual(r.status_code, 403)

    def test_site_engineer_can_export_but_no_values(self):
        # a client sheet carries no values, so site staff may generate it
        se = make_user("px_se", User.Role.SITE_ENGINEER, site=self.site)
        self._signed_schedule([
            {"description": "Glass", "section_code": "A", "quantity": "3",
             "uom": "nos", "required_date": "2026-09-01"}])
        ws = self._export(se).active
        allcells = "\n".join(
            str(ws.cell(row=r, column=c).value)
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None)
        self.assertIn("Glass", allcells)
        self.assertNotIn("5000", allcells)
