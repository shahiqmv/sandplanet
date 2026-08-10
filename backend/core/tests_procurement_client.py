"""Procurement Schedule Phase 4 — client-supplied lines + staleness chase."""
from datetime import date, timedelta

from django.test import TestCase
from django.utils import timezone
from rest_framework.test import APIClient

from . import procurement_pipeline as pp
from .models import (Notification, Project, ScheduleLine, Site, SitePmHistory,
                     User)
from .tests import make_user


def _stage(payload_line, key):
    return next(s for s in payload_line["pipeline"] if s["key"] == key)


class ProcurementClientTests(TestCase):
    def setUp(self):
        self.site = Site.objects.create(code="SJR", name="Jani",
                                        status=Site.Status.ACTIVE)
        self.pm = make_user("pc_pm", User.Role.PM, site=self.site)
        SitePmHistory.objects.create(site=self.site, pm_user=self.pm,
                                     from_date=date(2026, 1, 1))
        self.project = Project.objects.create(
            site=self.site, code="SJR-01", title="Villa Upgrades", pm=self.pm)
        self.purch = make_user("pc_buy", User.Role.HO_PURCHASING)
        self.director = make_user("pc_dir", User.Role.DIRECTOR)
        self.client = APIClient()
        self.today = timezone.localdate()

    def _signed_client_line(self, **fields):
        self.client.force_authenticate(self.pm)
        self.pk = self.client.post(
            f"/api/v1/projects/{self.project.id}/procurement-schedule").data["id"]
        body = {"description": "Client sofas", "section_code": "A",
                "supply_by": "CLIENT", **fields}
        line_id = self.client.post(
            f"/api/v1/procurement-schedules/{self.pk}/lines", body,
            format="json").data["lines"][0]["id"]
        self.client.post(f"/api/v1/procurement-schedules/{self.pk}/submit")
        self.client.force_authenticate(self.purch)
        self.client.post(f"/api/v1/procurement-schedules/{self.pk}/action",
                         {"action": "confirm"}, format="json")
        self.client.force_authenticate(self.director)
        self.client.post(f"/api/v1/procurement-schedules/{self.pk}/action",
                         {"action": "sign_off"}, format="json")
        self.client.force_authenticate(self.pm)
        return ScheduleLine.objects.get(pk=line_id)

    def _line_payload(self, line_id):
        d = self.client.get(f"/api/v1/procurement-schedules/{self.pk}").data
        return next(l for l in d["lines"] if l["id"] == line_id)

    def test_client_pipeline_has_no_planet_docs(self):
        line = self._signed_client_line(
            required_date=(self.today + timedelta(days=30)).isoformat())
        p = self._line_payload(line.id)
        self.assertEqual(_stage(p, "tds")["state"], "na")
        self.assertEqual(_stage(p, "order")["state"], "na")
        self.assertEqual(_stage(p, "delivery")["state"], "pending")

    def test_record_update_clears_staleness(self):
        # old update → stale
        line = self._signed_client_line()
        ScheduleLine.objects.filter(pk=line.id).update(
            client_last_update=self.today - timedelta(days=40))
        self.assertTrue(pp.client_is_stale(
            ScheduleLine.objects.get(pk=line.id)))
        # recording an update refreshes it
        r = self.client.post(
            f"/api/v1/procurement-schedule-lines/{line.id}/client-update",
            {"note": "On the boat next week"}, format="json")
        self.assertEqual(r.status_code, 200, r.data)
        self.assertFalse(pp.client_is_stale(
            ScheduleLine.objects.get(pk=line.id)))

    def test_mark_delivered_sets_pipeline_done(self):
        line = self._signed_client_line(
            required_date=(self.today + timedelta(days=5)).isoformat())
        self.client.post(
            f"/api/v1/procurement-schedule-lines/{line.id}/client-update",
            {"note": "Arrived", "delivered": True}, format="json")
        p = self._line_payload(line.id)
        self.assertEqual(_stage(p, "delivery")["state"], "done")
        self.assertEqual(p["risk"]["level"], "DELIVERED")

    def test_contractor_line_rejects_client_update(self):
        self.client.force_authenticate(self.pm)
        pk = self.client.post(
            f"/api/v1/projects/{self.project.id}/procurement-schedule").data["id"]
        line_id = self.client.post(
            f"/api/v1/procurement-schedules/{pk}/lines",
            {"description": "Steel", "section_code": "A",
             "supply_by": "CONTRACTOR"}, format="json").data["lines"][0]["id"]
        r = self.client.post(
            f"/api/v1/procurement-schedule-lines/{line_id}/client-update",
            {"note": "x"}, format="json")
        self.assertEqual(r.status_code, 400)

    def test_staleness_sweep_chases_pm_then_dedupes(self):
        line = self._signed_client_line(
            required_date=(self.today + timedelta(days=30)).isoformat())
        ScheduleLine.objects.filter(pk=line.id).update(
            client_last_update=self.today - timedelta(days=40))
        sent = pp.sweep_client_staleness()
        self.assertEqual(sent, 1)
        chase = Notification.objects.filter(
            recipient=self.pm, title__startswith="Chase client update")
        self.assertEqual(chase.count(), 1)
        # watermark prevents a same-window re-chase
        self.assertEqual(pp.sweep_client_staleness(), 0)

    def test_delivered_client_line_not_chased(self):
        line = self._signed_client_line()
        ScheduleLine.objects.filter(pk=line.id).update(
            client_last_update=self.today - timedelta(days=40),
            client_delivered_on=self.today)
        self.assertEqual(pp.sweep_client_staleness(), 0)


class BundleVariantsTests(ProcurementClientTests):
    def test_client_plan_bundle_carries_variants(self):
        from . import procurement_client as pc
        from .models import ProcurementSchedule
        self.client.force_authenticate(self.pm)
        pk = self.client.post(
            f"/api/v1/projects/{self.project.id}/procurement-schedule"
        ).data["id"]
        for desc in ("uPVC 20mm", "uPVC 25mm"):
            self.client.post(
                f"/api/v1/procurement-schedules/{pk}/lines",
                {"description": desc, "section_code": "D",
                 "bundle": "uPVC fittings", "quantity": "10", "uom": "nos"},
                format="json")
        ScheduleLine.objects.filter(schedule__document_id=pk).update(
            planned_supplier="S-LON")
        sched = ProcurementSchedule.objects.get(document_id=pk)
        rows = [r for sec in pc.client_plan(sched)["sections"]
                for r in sec["rows"]]
        bundle = next(r for r in rows if r.get("is_bundle"))
        self.assertEqual(bundle["remarks"], "2 variants")
        self.assertEqual({v["description"] for v in bundle["variants"]},
                         {"uPVC 20mm", "uPVC 25mm"})

    def test_xlsx_export_includes_expanded_variants(self):
        import io

        from openpyxl import load_workbook

        from . import procurement_client as pc, procurement_export as pe
        from .models import ProcurementSchedule
        self.client.force_authenticate(self.pm)
        pk = self.client.post(
            f"/api/v1/projects/{self.project.id}/procurement-schedule"
        ).data["id"]
        for desc in ("uPVC 20mm", "uPVC 25mm"):
            self.client.post(
                f"/api/v1/procurement-schedules/{pk}/lines",
                {"description": desc, "section_code": "D",
                 "bundle": "uPVC fittings", "quantity": "10", "uom": "nos"},
                format="json")
        ScheduleLine.objects.filter(schedule__document_id=pk).update(
            planned_supplier="S-LON")
        sched = ProcurementSchedule.objects.get(document_id=pk)
        wb = pe.build_client_xlsx_from_plan(pc.client_plan(sched), expand=True)
        buf = io.BytesIO()
        wb.save(buf)
        text = " ".join(
            str(c.value) for row in load_workbook(io.BytesIO(buf.getvalue()))
            .active.iter_rows() for c in row if c.value)
        self.assertIn("uPVC 20mm", text)          # variant expanded into xlsx
        self.assertIn("uPVC 25mm", text)
