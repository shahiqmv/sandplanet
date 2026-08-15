from datetime import date

from django.utils import timezone
from django.contrib.auth import authenticate, login, logout
from django.db import connection
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response

from .audit import audit
from .models import (
    CompanyParameter,
    Holiday,
    ManpowerCategory,
    Site,
    SitePmHistory,
    User,
    UserSiteAllocation,
)
from .permissions import (
    IsAdmin,
    IsAdminOrReadOnly,
    IsHrAdminOrReadOnly,
    IsSiteManagerOrReadOnly,
    scoped_site_ids,
)
from .serializers import (
    AllocationSerializer,
    HolidaySerializer,
    ManpowerCategorySerializer,
    ParameterSerializer,
    SiteSerializer,
    UserSerializer,
)


@api_view(["GET"])
@permission_classes([AllowAny])
def health(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT 1")
        db_ok = cursor.fetchone()[0] == 1
    return Response(
        {"status": "ok", "db": "ok" if db_ok else "error", "engine": connection.vendor}
    )


# ===== Auth =====


def _client_ip(request):
    xff = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if xff:
        return xff.split(",")[0].strip()[:45]
    return (request.META.get("REMOTE_ADDR") or "")[:45]


def record_login_event(request, kind, user=None, username="", source="WEB"):
    """Append a sign-in / sign-out / failed-attempt row for the admin security
    view. Never raises — auth must not break on a logging hiccup."""
    from .models import LoginEvent
    try:
        LoginEvent.objects.create(
            user=user if (user and user.pk) else None,
            username=(username or (user.username if user else ""))[:150],
            kind=kind, source=source, ip_address=_client_ip(request),
            user_agent=(request.META.get("HTTP_USER_AGENT") or "")[:300])
    except Exception:                                   # pragma: no cover
        pass


@api_view(["POST"])
@permission_classes([AllowAny])
def auth_login(request):
    username = request.data.get("username", "")
    user = authenticate(request, username=username,
                        password=request.data.get("password", ""))
    if user is None or not user.is_active:
        record_login_event(request, "FAILED", username=username)
        return Response({"detail": "Invalid credentials."}, status=400)
    login(request, user)
    record_login_event(request, "LOGIN", user=user)
    return Response(_me_payload(user))


@api_view(["POST"])
def auth_logout(request):
    if request.user.is_authenticated:
        record_login_event(request, "LOGOUT", user=request.user)
    logout(request)
    return Response({"detail": "Logged out."})


@ensure_csrf_cookie
@api_view(["GET"])
@permission_classes([AllowAny])
def auth_me(request):
    """Also sets the CSRF cookie — the SPA calls this first."""
    if not request.user.is_authenticated:
        return Response({"authenticated": False})
    return Response(_me_payload(request.user))


def _me_payload(user):
    allocations = AllocationSerializer(
        user.site_allocations.filter(to_date__isnull=True).select_related("site"),
        many=True,
    ).data
    # Single-site roles land directly on their site (brief: no site picker)
    landing_site = allocations[0]["site"] if len(allocations) == 1 else None
    return {
        "authenticated": True,
        "id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "role": user.role,
        "is_ho": user.is_ho,
        "allocations": allocations,
        "landing_site_id": landing_site,
        "must_change_password": user.must_change_password,
        "phone": user.phone,
        "notify_external": user.notify_external,
    }


@api_view(["POST"])
def auth_change_password(request):
    """Any signed-in user sets a new password (also clears the
    must-change flag from an invite)."""
    current = request.data.get("current_password", "")
    new = request.data.get("new_password", "")
    user = request.user
    if not user.check_password(current):
        return Response({"detail": "Current password is incorrect."},
                        status=400)
    if len(new) < 8:
        return Response({"detail": "New password must be at least 8 "
                                   "characters."}, status=400)
    user.set_password(new)
    user.must_change_password = False
    user.save(update_fields=["password", "must_change_password"])
    from django.contrib.auth import update_session_auth_hash
    update_session_auth_hash(request, user)  # keep the session alive
    audit("user", user.id, "PASSWORD_CHANGED", actor=user)
    return Response({"detail": "Password updated."})


# ===== Sites =====


class SiteViewSet(viewsets.ModelViewSet):
    serializer_class = SiteSerializer
    permission_classes = [IsSiteManagerOrReadOnly]  # Admin + Director (R4)
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_queryset(self):
        qs = Site.objects.all().order_by("code")
        site_ids = scoped_site_ids(self.request.user)
        if site_ids is not None:
            qs = qs.filter(id__in=site_ids)
        return qs

    def perform_create(self, serializer):
        site = serializer.save()
        audit("site", site.id, "SITE_CREATED", actor=self.request.user,
              to_state=site.status, detail={"code": site.code})

    def perform_update(self, serializer):
        site = serializer.save()
        audit("site", site.id, "SITE_UPDATED", actor=self.request.user,
              detail={"fields": sorted(self.request.data.keys())})

    @action(detail=True, methods=["post"])
    def status(self, request, pk=None):
        """Lifecycle transition; reason required; every change audited (§2.2)."""
        site = self.get_object()
        new_status = request.data.get("status")
        reason = (request.data.get("reason") or "").strip()
        if new_status not in Site.Status.values:
            return Response({"detail": f"Unknown status '{new_status}'."}, status=400)
        if not reason:
            return Response({"detail": "A reason is required."}, status=400)
        allowed = Site.TRANSITIONS.get(site.status, set())
        if new_status not in allowed:
            return Response(
                {"detail": f"Cannot move {site.status} → {new_status}."}, status=400
            )
        if site.status == Site.Status.CLOSED and request.user.role != User.Role.ADMIN:
            return Response(
                {"detail": "Only Admin can reopen a closed site."}, status=403
            )
        old = site.status
        site.status = new_status
        if new_status == Site.Status.CLOSED and not site.actual_completion:
            site.actual_completion = request.data.get("actual_completion") or date.today()
        site.save()
        audit("site", site.id, "SITE_STATUS_CHANGED", actor=request.user,
              from_state=old, to_state=new_status, detail={"reason": reason})
        return Response(self.get_serializer(site).data)

    @action(detail=True, methods=["post"], url_path="assign-pm")
    def assign_pm(self, request, pk=None):
        """Manage the site's PM(s); history kept (spec §2.1). `mode`:
          - "replace" (default): this PM becomes the sole current PM (any
            existing PMs are closed today) — the classic reassignment.
          - "add": add this PM as a co-PM alongside the existing one(s). A busy
            site can carry several — co-PMs share full PM authority.
          - "remove": close this PM's current assignment (leaves any others)."""
        site = self.get_object()
        mode = (request.data.get("mode") or "replace").lower()
        try:
            pm = User.objects.get(pk=request.data.get("pm_user_id"),
                                  role=User.Role.PM, is_active=True)
        except User.DoesNotExist:
            return Response({"detail": "pm_user_id must be an active PM."},
                            status=400)
        today = date.today()
        previous = site.current_pm()
        if mode == "remove":
            n = site.pm_history.filter(
                pm_user=pm, to_date__isnull=True).update(to_date=today)
            if not n:
                return Response({"detail": "That PM is not a current PM here."},
                                status=400)
            audit("site", site.id, "SITE_PM_REMOVED", actor=request.user,
                  from_state=pm.username, to_state="")
            return Response(self.get_serializer(site).data)
        if mode != "add":                        # replace — close the others
            site.pm_history.filter(to_date__isnull=True).update(to_date=today)
        # add / replace — open a row for this PM if they aren't already current
        if not site.pm_history.filter(
                pm_user=pm, to_date__isnull=True).exists():
            SitePmHistory.objects.create(site=site, pm_user=pm, from_date=today)
        # PM approval routing needs a read allocation on the site
        if not UserSiteAllocation.objects.filter(
            user=pm, site=site, to_date__isnull=True
        ).exists():
            UserSiteAllocation.objects.create(user=pm, site=site, from_date=today)
        audit("site", site.id,
              "SITE_PM_ADDED" if mode == "add" else "SITE_PM_ASSIGNED",
              actor=request.user,
              from_state=previous.username if previous else "",
              to_state=pm.username)
        return Response(self.get_serializer(site).data)


# ===== Users (admin-managed accounts, no self-registration) =====


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_directory(request):
    """Lightweight internal directory (id / name / role) for people-pickers —
    meeting attendees, action-item owners, etc. Read-only, no management or
    sensitive fields, so it's open to any authenticated user (managing users
    stays admin-only on the UserViewSet)."""
    users = User.objects.filter(is_active=True).order_by("full_name",
                                                         "username")
    return Response([{"id": u.id, "full_name": u.full_name or u.username,
                      "role": u.role, "username": u.username} for u in users])


class UserViewSet(viewsets.ModelViewSet):
    serializer_class = UserSerializer
    permission_classes = [IsAdmin]
    queryset = User.objects.all().order_by("username")
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def destroy(self, request, *args, **kwargs):
        from django.db.models import ProtectedError
        user = self.get_object()
        if user.id == request.user.id:
            return Response({"detail": "You can't delete your own account."},
                            status=400)
        username = user.username
        try:
            user.delete()
        except ProtectedError:
            return Response(
                {"detail": "This user has records (documents, approvals, "
                           "payments) and can't be deleted — deactivate them "
                           "instead."}, status=400)
        audit("user", 0, "USER_DELETED", actor=request.user,
              detail={"username": username})
        return Response(status=204)

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        # Report whether the welcome email went out (set on the instance in
        # perform_create) so the admin sees it immediately.
        response.data["invite_sent"] = getattr(self, "_invite_sent", False)
        if getattr(self, "_invite_error", None):
            response.data["invite_error"] = self._invite_error
        return response

    def perform_create(self, serializer):
        user = serializer.save()
        audit("user", user.id, "USER_CREATED", actor=self.request.user,
              detail={"username": user.username, "role": user.role})
        self._invite_sent, self._invite_error = self._maybe_invite(
            user, getattr(user, "_temp_password", None))

    def _maybe_invite(self, user, temp_password):
        """Email login details if a temp password was issued and an address is
        on file. Returns (sent, error)."""
        if not temp_password or not user.email:
            return False, None
        from .invites import send_user_invite
        try:
            send_user_invite(user, temp_password)
            audit("user", user.id, "USER_INVITE_SENT", actor=self.request.user,
                  detail={"email": user.email})
            return True, None
        except Exception as exc:  # noqa: BLE001 — surface send failures to admin
            return False, str(exc)

    def perform_update(self, serializer):
        user = serializer.save()
        audit("user", user.id, "USER_UPDATED", actor=self.request.user,
              detail={"fields": sorted(self.request.data.keys())})

    @action(detail=True, methods=["post"], url_path="change-role")
    def change_role(self, request, pk=None):
        """Promote/demote a user to another role. When promoting to PM, an
        optional `assign_site_id` also makes them a (co-)PM of that site in the
        same step. Changing AWAY from PM closes any open site-PM assignments."""
        user = self.get_object()
        new_role = request.data.get("role")
        if new_role not in User.Role.values:
            return Response({"detail": "Unknown role."}, status=400)
        if user.id == request.user.id:
            return Response(
                {"detail": "You can't change your own role."}, status=400)
        old_role = user.role
        if new_role == old_role and not request.data.get("assign_site_id"):
            return Response({"detail": "That's already their role."}, status=400)
        from django.db import transaction
        today = date.today()
        with transaction.atomic():
            if old_role != new_role:
                user.role = new_role
                user.save(update_fields=["role"])
                audit("user", user.id, "USER_ROLE_CHANGED", actor=request.user,
                      from_state=old_role, to_state=new_role)
                # No longer a PM anywhere → close open site-PM assignments.
                if old_role == User.Role.PM and new_role != User.Role.PM:
                    SitePmHistory.objects.filter(
                        pm_user=user, to_date__isnull=True).update(to_date=today)
            assigned = None
            site_id = request.data.get("assign_site_id")
            if new_role == User.Role.PM and site_id:
                site = Site.objects.filter(pk=site_id).first()
                if site is None:
                    return Response({"detail": "Unknown site."}, status=400)
                if not site.pm_history.filter(
                        pm_user=user, to_date__isnull=True).exists():
                    SitePmHistory.objects.create(
                        site=site, pm_user=user, from_date=today)
                if not UserSiteAllocation.objects.filter(
                        user=user, site=site, to_date__isnull=True).exists():
                    UserSiteAllocation.objects.create(
                        user=user, site=site, from_date=today)
                assigned = site.code
                audit("site", site.id, "SITE_PM_ADDED", actor=request.user,
                      to_state=user.username)
        data = self.get_serializer(user).data
        data["assigned_pm_site"] = assigned
        return Response(data)

    @action(detail=True, methods=["post"])
    def resend_invite(self, request, pk=None):
        """Re-issue a temporary password and email it (e.g. lost invite)."""
        from .invites import make_temp_password, send_user_invite
        user = self.get_object()
        if not user.email:
            return Response({"detail": "This user has no email address."},
                            status=400)
        temp = make_temp_password()
        user.set_password(temp)
        user.must_change_password = True
        user.save(update_fields=["password", "must_change_password"])
        try:
            send_user_invite(user, temp)
        except Exception as exc:  # noqa: BLE001
            return Response({"detail": f"Email failed: {exc}"}, status=502)
        audit("user", user.id, "USER_INVITE_RESENT", actor=request.user,
              detail={"email": user.email})
        return Response({"invite_sent": True})

    @action(detail=True, methods=["post"])
    def deactivate(self, request, pk=None):
        """Deactivation, never deletion (NFR §9)."""
        user = self.get_object()
        user.is_active = False
        user.save()
        UserSiteAllocation.objects.filter(user=user, to_date__isnull=True).update(
            to_date=date.today()
        )
        audit("user", user.id, "USER_DEACTIVATED", actor=request.user)
        return Response(self.get_serializer(user).data)

    @action(detail=True, methods=["post"])
    def allocate(self, request, pk=None):
        """Allocate a user to a site (closes previous open allocation for
        single-site roles)."""
        user = self.get_object()
        try:
            site = Site.objects.get(pk=request.data.get("site_id"))
        except Site.DoesNotExist:
            return Response({"detail": "Unknown site_id."}, status=400)
        today = date.today()
        if user.role in User.SINGLE_SITE_ROLES:
            user.site_allocations.filter(to_date__isnull=True).update(to_date=today)
        UserSiteAllocation.objects.create(user=user, site=site, from_date=today)
        audit("user", user.id, "USER_ALLOCATED", actor=request.user,
              to_state=site.code)
        return Response(self.get_serializer(user).data)


# ===== Master data =====


class ManpowerCategoryViewSet(viewsets.ModelViewSet):
    serializer_class = ManpowerCategorySerializer
    permission_classes = [IsHrAdminOrReadOnly]
    queryset = ManpowerCategory.objects.all()
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def perform_destroy(self, instance):
        # Deactivate rather than delete if the category is referenced by an
        # employee — history must survive (spec §6A.1)
        if instance.employees.exists():
            instance.is_active = False
            instance.save(update_fields=["is_active"])
        else:
            instance.delete()


class HolidayViewSet(viewsets.ModelViewSet):
    serializer_class = HolidaySerializer
    permission_classes = [IsAdminOrReadOnly]
    queryset = Holiday.objects.all().order_by("day")
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def cost_heads(request):
    """Cost heads for PYR / petty cash pickers (§6C.1). Project heads only
    by default; ?pools=1 includes the three HO pools."""
    from .models import CostHead

    qs = CostHead.objects.filter(is_active=True)
    if request.GET.get("pools") != "1":
        qs = qs.filter(is_pool=False)
    return Response([{"id": c.id, "name": c.name, "is_pool": c.is_pool}
                     for c in qs])


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def pm_list(request):
    """Active PM users, for site/project PM assignment pick-lists."""
    pms = User.objects.filter(role=User.Role.PM, is_active=True) \
        .order_by("full_name")
    return Response([{"id": u.id, "full_name": u.full_name} for u in pms])


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def pm_overview(request):
    """PM assignments board (R5): every active PM with the sites they run
    (current site PM), the projects they run (project PM), and the site-PM
    history. Site PM is a special duty — managed on its own page."""
    if request.user.role not in ("ADMIN", "DIRECTOR"):
        return Response({"detail": "Admin/Director only."}, status=403)
    from .models import Project

    pms = list(User.objects.filter(role=User.Role.PM).order_by("full_name"))
    current = SitePmHistory.objects.filter(to_date__isnull=True) \
        .select_related("site")
    sites_by_pm = {}
    for h in current:
        sites_by_pm.setdefault(h.pm_user_id, []).append(
            {"site_id": h.site_id, "code": h.site.code, "name": h.site.name,
             "since": h.from_date}
        )
    projects_by_pm = {}
    for p in Project.objects.filter(pm__isnull=False).select_related("site"):
        projects_by_pm.setdefault(p.pm_id, []).append(
            {"project_id": p.id, "code": p.code, "title": p.title,
             "site_code": p.site.code, "status": p.status}
        )
    history = [
        {"pm_id": h.pm_user_id, "pm_name": h.pm_user.full_name,
         "site_code": h.site.code, "site_name": h.site.name,
         "from_date": h.from_date, "to_date": h.to_date}
        for h in SitePmHistory.objects.select_related("site", "pm_user")
        .order_by("-from_date")[:100]
    ]
    return Response({
        "pms": [{
            "id": u.id, "username": u.username, "full_name": u.full_name,
            "email": u.email, "is_active": u.is_active,
            "sites": sites_by_pm.get(u.id, []),
            "projects": projects_by_pm.get(u.id, []),
        } for u in pms],
        "history": history,
    })


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def company_logo(request):
    """Company logo image used on every PDF letterhead. Stored via the
    configured storage — Spaces in production, local disk in dev — at
    company/logo.png|jpg; PDFs fall back to the bundled stationery logo when
    nothing is uploaded."""
    from django.core.files.base import ContentFile
    from django.core.files.storage import default_storage

    names = ("company/logo.png", "company/logo.jpg")
    if request.method == "POST":
        if request.user.role != User.Role.ADMIN:
            return Response({"detail": "Admin only."}, status=403)
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Attach the logo as 'file'."}, status=400)
        ext = {"image/png": "png", "image/jpeg": "jpg"}.get(file.content_type)
        if not ext:
            return Response({"detail": "PNG or JPEG only."}, status=400)
        for old in names:  # one logo at a time
            if default_storage.exists(old):
                default_storage.delete(old)
        default_storage.save(f"company/logo.{ext}", ContentFile(file.read()))
        audit("parameter", 0, "COMPANY_LOGO_UPDATED", actor=request.user,
              detail={"file_name": file.name, "size": file.size})
    for name in names:
        if default_storage.exists(name):
            return Response({"url": default_storage.url(name), "uploaded": True})
    return Response({"url": None, "uploaded": False})


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def company_stamp(request):
    """The round company seal overlaid on the IM30 visa form and official
    letters, so HR doesn't have to print, stamp and scan. Stored at
    company/stamp.png|jpg via the configured storage (owner 2026-08-05)."""
    from django.core.files.base import ContentFile
    from django.core.files.storage import default_storage

    names = ("company/stamp.png", "company/stamp.jpg")
    if request.method == "POST":
        if request.user.role != User.Role.ADMIN:
            return Response({"detail": "Admin only."}, status=403)
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Attach the stamp as 'file'."}, status=400)
        ext = {"image/png": "png", "image/jpeg": "jpg"}.get(file.content_type)
        if not ext:
            return Response({"detail": "PNG or JPEG only."}, status=400)
        for old in names:  # one stamp at a time
            if default_storage.exists(old):
                default_storage.delete(old)
        default_storage.save(f"company/stamp.{ext}", ContentFile(file.read()))
        audit("parameter", 0, "COMPANY_STAMP_UPDATED", actor=request.user,
              detail={"file_name": file.name, "size": file.size})
    for name in names:
        if default_storage.exists(name):
            return Response({"url": default_storage.url(name), "uploaded": True})
    return Response({"url": None, "uploaded": False})


def _paginate(request, default=50, cap=200):
    try:
        limit = min(int(request.GET.get("limit", default)), cap)
    except (TypeError, ValueError):
        limit = default
    try:
        offset = max(int(request.GET.get("offset", 0)), 0)
    except (TypeError, ValueError):
        offset = 0
    return limit, offset


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def login_activity(request):
    """Sign-in activity for the admin security view — successful and failed
    sign-ins and sign-outs, newest first (owner 2026-08-06)."""
    from .models import LoginEvent
    if request.user.role != User.Role.ADMIN:
        return Response({"detail": "Administrators only."}, status=403)
    qs = LoginEvent.objects.select_related("user")
    kind = request.GET.get("kind")
    if kind:
        qs = qs.filter(kind=kind)
    if request.GET.get("user_id"):
        qs = qs.filter(user_id=request.GET["user_id"])
    if request.GET.get("q"):
        qs = qs.filter(username__icontains=request.GET["q"])
    if request.GET.get("since"):
        qs = qs.filter(at__date__gte=request.GET["since"])
    if request.GET.get("until"):
        qs = qs.filter(at__date__lte=request.GET["until"])
    total = qs.count()
    limit, offset = _paginate(request)
    rows = [{
        "id": e.id, "at": e.at, "kind": e.kind, "source": e.source,
        "username": e.username or (e.user.username if e.user_id else ""),
        "full_name": e.user.full_name if e.user_id else "",
        "role": e.user.role if e.user_id else "",
        "ip_address": e.ip_address, "user_agent": e.user_agent,
    } for e in qs[offset:offset + limit]]
    return Response({"items": rows, "total": total,
                     "limit": limit, "offset": offset})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def audit_trail(request):
    """The append-only audit log for the admin security view — every state
    change and administrative action, filterable, newest first."""
    from .models import AuditLog
    if request.user.role != User.Role.ADMIN:
        return Response({"detail": "Administrators only."}, status=403)
    qs = AuditLog.objects.select_related("actor")
    for field in ("entity", "event"):
        if request.GET.get(field):
            qs = qs.filter(**{field: request.GET[field]})
    if request.GET.get("actor_id"):
        qs = qs.filter(actor_id=request.GET["actor_id"])
    if request.GET.get("entity_id"):
        qs = qs.filter(entity_id=request.GET["entity_id"])
    if request.GET.get("since"):
        qs = qs.filter(at__date__gte=request.GET["since"])
    if request.GET.get("until"):
        qs = qs.filter(at__date__lte=request.GET["until"])
    total = qs.count()
    limit, offset = _paginate(request)
    rows = [{
        "id": a.id, "at": a.at, "entity": a.entity, "entity_id": a.entity_id,
        "event": a.event, "from_state": a.from_state, "to_state": a.to_state,
        "actor": a.actor.full_name if a.actor_id else "",
        "actor_role": a.actor.role if a.actor_id else "",
        "detail": a.detail,
    } for a in qs[offset:offset + limit]]
    # distinct entity/event lists power the filter dropdowns
    return Response({"items": rows, "total": total, "limit": limit,
                     "offset": offset,
                     "entities": sorted(AuditLog.objects.values_list(
                         "entity", flat=True).distinct())})


@api_view(["GET", "PUT"])
@permission_classes([IsAuthenticated])
def parameter_detail(request, key):
    if request.method == "PUT":
        if request.user.role != User.Role.ADMIN:
            return Response({"detail": "Admin only."}, status=403)
        param, _ = CompanyParameter.objects.get_or_create(
            key=key, defaults={"value": ""})
        serializer = ParameterSerializer(param, data={**request.data, "key": key})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        audit("parameter", 0, "PARAMETER_UPDATED", actor=request.user,
              detail={"key": key})
        return Response(serializer.data)
    try:
        param = CompanyParameter.objects.get(key=key)
    except CompanyParameter.DoesNotExist:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
    return Response(ParameterSerializer(param).data)


# ===== Item Master (spec §5.0) — owned by HO Purchasing =====


from rest_framework.permissions import BasePermission  # noqa: E402

from .models import Item  # noqa: E402
from rest_framework.parsers import FormParser, MultiPartParser  # noqa: E402

from .procurement import next_item_code  # noqa: E402
from .serializers_documents import ItemSerializer  # noqa: E402


class IsPurchasingOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.method in ("GET", "HEAD", "OPTIONS"):
            return True  # sites need the catalog for MR autocomplete
        return request.user.role in ("HO_PURCHASING", "ADMIN")


class CanEditCatalogItem(BasePermission):
    """HO Purchasing/Admin manage the catalogue; site teams may CREATE a
    missing item (flagged provisional) while receiving goods / adding tools."""
    OWNER = ("HO_PURCHASING", "ADMIN")
    CREATOR = OWNER + ("SITE_ADMIN", "SITE_ENGINEER", "PM", "DIRECTOR")

    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        if request.method in ("GET", "HEAD", "OPTIONS"):
            return True
        if request.method == "POST":
            return request.user.role in self.CREATOR
        return request.user.role in self.OWNER   # PATCH: owners only


class ItemCategoryViewSet(viewsets.ModelViewSet):
    """Controlled item categories, managed by HO Purchasing on their own
    page (owner, 2026-07-08)."""
    from .models import ItemCategory
    from .serializers import ItemCategorySerializer

    serializer_class = ItemCategorySerializer
    permission_classes = [IsPurchasingOrReadOnly]
    queryset = ItemCategory.objects.all()
    http_method_names = ["get", "post", "patch", "delete", "head", "options"]

    def perform_destroy(self, instance):
        # Keep categories still in use by items — deactivate instead
        if Item.objects.filter(category=instance.name).exists():
            instance.is_active = False
            instance.save(update_fields=["is_active"])
        else:
            instance.delete()


class ItemViewSet(viewsets.ModelViewSet):
    serializer_class = ItemSerializer
    permission_classes = [CanEditCatalogItem]
    http_method_names = ["get", "post", "patch", "head", "options"]

    def get_queryset(self):
        qs = Item.objects.filter(merged_into__isnull=True).order_by("code")
        search = self.request.GET.get("search")
        if search:
            from django.db.models import Q

            qs = qs.filter(Q(description__icontains=search) |
                           Q(code__icontains=search) |
                           Q(category__icontains=search))
        if self.request.GET.get("active") != "all":
            qs = qs.filter(is_active=True)
        return qs

    def perform_create(self, serializer):
        from django.db import transaction

        # Provisional/approval gate (owner 2026-07-14): items created by site
        # staff are provisional until HO Purchasing/Admin review the spelling &
        # category and approve them. HO/Admin create permanent items directly.
        provisional = self.request.user.role not in ("HO_PURCHASING", "ADMIN")
        with transaction.atomic():  # row-locked counter needs a transaction
            item = serializer.save(code=next_item_code(),
                                   is_provisional=provisional)
        audit("item", item.id, "ITEM_CREATED", actor=self.request.user,
              detail={"code": item.code, "provisional": provisional})

    def perform_update(self, serializer):
        item = serializer.save()
        audit("item", item.id, "ITEM_UPDATED", actor=self.request.user,
              detail={"fields": sorted(self.request.data.keys())})

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """A ready-to-fill Excel template for the bulk item import (owner
        2026-07-14)."""
        import io

        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font

        from .models import ItemCategory

        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        headers = ["Description", "Unit", "Category", "Brand", "Spec Ref",
                   "Key material (yes/no)"]
        ws.append(headers)
        widths = [42, 10, 20, 18, 18, 20]
        for i, w in enumerate(widths, 1):
            ws.cell(row=1, column=i).font = Font(bold=True)
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"

        info = wb.create_sheet("Instructions")
        cats = ", ".join(ItemCategory.objects.filter(is_active=True)
                         .order_by("name").values_list("name", flat=True))
        for line in [
            "Fill the 'Items' sheet — one row per catalogue item.",
            "",
            "Description  — required (item name / spec).",
            "Unit         — required (bag, kg, nos, m, m2, m3, ltr, roll…).",
            "Category     — optional; must match an existing Item Category.",
            "Brand        — optional.",
            "Spec Ref     — optional.",
            "Key material — 'yes' to flag it as a DPR key material.",
            "",
            f"Item Categories: {cats or '(none defined yet)'}",
            "",
            "Codes (ITM-…) are assigned automatically. A row whose description "
            "already exists is skipped, so re-uploading a file is safe.",
        ]:
            info.append([line])
        info.column_dimensions["A"].width = 95

        buf = io.BytesIO()
        wb.save(buf)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet")
        resp["Content-Disposition"] = \
            'attachment; filename="item-import-template.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="import",
            parser_classes=[MultiPartParser, FormParser])
    def bulk_import(self, request):
        """Create catalogue items from a filled Excel sheet (owner
        2026-07-14). HO Purchasing / Admin only."""
        if request.user.role not in ("HO_PURCHASING", "ADMIN"):
            return Response({"detail": "Head Office manages the catalogue."},
                            status=403)
        upload = request.FILES.get("file")
        if upload is None:
            return Response({"detail": "Attach the filled Excel (.xlsx) file."},
                            status=400)
        from openpyxl import load_workbook

        from .item_import import import_item_rows, normalise_header
        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
        except Exception:
            return Response({"detail": "Could not read the file — save it as "
                                       ".xlsx and try again."}, status=400)
        ws = wb["Items"] if "Items" in wb.sheetnames else wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if header is None:
            return Response({"detail": "The sheet is empty."}, status=400)
        keys = [normalise_header(h) for h in header]
        if "description" not in keys:
            return Response({"detail": "The sheet needs a 'Description' "
                                       "column (use the template)."},
                            status=400)
        rows = []
        for raw in it:
            if not any(str(v).strip() for v in raw if v is not None):
                continue                       # blank row
            rows.append({k: v for k, v in zip(keys, raw) if k})
        if not rows:
            return Response({"detail": "No item rows found below the header."},
                            status=400)
        result = import_item_rows(rows)
        audit("item", 0, "ITEMS_BULK_IMPORTED", actor=request.user,
              detail={"created": result["created"],
                      "skipped": result["skipped"]})
        return Response(result)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        """HO Purchasing / Admin confirm a site-created provisional item."""
        if request.user.role not in ("HO_PURCHASING", "ADMIN"):
            return Response({"detail": "HO Purchasing/Admin approve items."},
                            status=403)
        item = self.get_object()
        item.is_provisional = False
        item.save(update_fields=["is_provisional", "updated_at"])
        audit("item", item.id, "ITEM_APPROVED", actor=request.user)
        return Response(self.get_serializer(item).data)

    @action(detail=True, methods=["post"])
    def merge(self, request, pk=None):
        """Duplicate resolution (spec §5.0): this item merges into target;
        existing document lines keep their history."""
        item = self.get_object()
        try:
            target = Item.objects.get(pk=request.data.get("target_id"),
                                      merged_into__isnull=True)
        except Item.DoesNotExist:
            return Response({"detail": "target_id must be an unmerged item."},
                            status=400)
        if target.pk == item.pk:
            return Response({"detail": "Cannot merge an item into itself."},
                            status=400)
        item.merged_into = target
        item.is_active = False
        item.save(update_fields=["merged_into", "is_active"])
        audit("item", item.id, "ITEM_MERGED", actor=request.user,
              to_state=target.code)
        return Response(self.get_serializer(target).data)


@api_view(["GET"])
def sites_summary(request):
    """Live signal per site for the landing page.

    The list said code, name and status and nothing else, so it answered "what
    exists" when the question people actually open it with is "where does my
    attention go" (owner 2026-08-15). Everything here is one aggregate query
    across all sites — no per-site loop, because this is the first screen
    after signing in.
    """
    from datetime import timedelta

    from django.db.models import Count, Max, Q

    from .models import Attendance, Document, Site

    sites = Site.objects.exclude(status="CLOSED")
    ids = scoped_site_ids(request.user)
    if ids is not None:
        sites = sites.filter(id__in=ids)
    site_ids = list(sites.values_list("id", flat=True))
    today = timezone.localdate()

    # Manpower: the last day anyone was marked, per site, and how many were
    # present on it — "28 on site" means more than a headcount on paper.
    last_day = dict(Attendance.objects.filter(site_id__in=site_ids)
                    .values("site_id").annotate(d=Max("day"))
                    .values_list("site_id", "d"))
    present = {}
    for sid, day in last_day.items():
        present[sid] = Attendance.objects.filter(
            site_id=sid, day=day, remark__in=("PRESENT", "HALF_DAY")).count()

    last_dpr = dict(Document.objects.filter(
        site_id__in=site_ids, doc_type="DPR", is_void=False)
        .values("site_id").annotate(d=Max("doc_date"))
        .values_list("site_id", "d"))

    # Anything sitting in a decision state — what the site is waiting on.
    WAITING = ("SUBMITTED", "PM_APPROVED", "DIRECTOR_APPROVED", "ISSUED",
               "PM_VERIFIED", "AMENDMENT_PENDING")
    open_docs = dict(Document.objects.filter(
        site_id__in=site_ids, is_void=False, status__in=WAITING)
        .values("site_id").annotate(n=Count("id"))
        .values_list("site_id", "n"))

    out = []
    for s in sites.order_by("code"):
        day = last_day.get(s.id)
        dpr = last_dpr.get(s.id)
        out.append({
            "id": s.id, "code": s.code, "name": s.name,
            "status": s.status, "is_head_office": s.is_head_office,
            "pms": [p.full_name for p in s.current_pms()],
            "manpower": present.get(s.id, 0),
            "manpower_day": day.isoformat() if day else None,
            "manpower_stale": bool(day and (today - day).days > 1),
            "last_dpr": dpr.isoformat() if dpr else None,
            "dpr_days_ago": (today - dpr).days if dpr else None,
            "open_docs": open_docs.get(s.id, 0),
        })
    return Response({"sites": out, "as_of": today.isoformat()})
