"""Capture a BOQ from a client PDF or Excel into a reviewable draft.

The document is read to plain text (pdfplumber for PDF, openpyxl for Excel —
all sheets), then Claude structures it into BOQ rows in the committer's shape
(section / item_code / description / unit / qty / rate_supply / rate_install /
rate_combined / is_heading). The rows land in a BoqImport staging record the QS
corrects before committing into the live BOQ. Nothing is auto-committed.

The Claude call reads ANTHROPIC_API_KEY from the environment and the model from
the `boq_extract_model` company parameter (default Claude Sonnet). The rest of
the pipeline (text extraction, normalisation, reconciliation) is pure and
testable without the API.
"""
import os
from decimal import Decimal, InvalidOperation

DEFAULT_MODEL = "claude-sonnet-5"
# Smaller, focused batches — a model asked to extract ~250 lines from one huge
# prompt gets "lazy" and returns a fraction. One-ish sheet per call (batches run
# concurrently, so more of them is cheap) keeps extraction complete.
_MAX_BATCH_CHARS = 11000
_MAX_OUTPUT_TOKENS = 16000      # ~150 BOQ rows per call without truncation


class ExtractionError(Exception):
    """A user-facing extraction failure (bad file, missing key, model error)."""


# ---- text extraction -----------------------------------------------------

def pdf_pages(fileobj):
    """Per-page text of a (digital) PDF, using layout-preserving extraction so a
    BOQ's columns stay aligned on each row.

    We deliberately do NOT use pdfplumber's table detector here: on borderless /
    tender-style BOQ grids (no internal row rules) it collapses each column into
    a single cell — every item code in one blob, every description in another —
    which destroys the row-to-row alignment the model needs to pair a
    description with its qty/rate/amount (and it even splits numbers). Laid-out
    text keeps each row on one line with columns positioned by whitespace, which
    the model reads reliably across both ruled and borderless formats."""
    import pdfplumber
    pages = []
    try:
        with pdfplumber.open(fileobj) as pdf:
            for page in pdf.pages:
                text = page.extract_text(layout=True) or ""
                if not text.strip():             # some pages lay out to nothing
                    text = page.extract_text() or ""
                # Drop blank lines and trailing padding, but KEEP leading spaces
                # — that indentation is the column alignment.
                text = "\n".join(ln.rstrip() for ln in text.splitlines()
                                 if ln.strip())
                pages.append(text)
    except Exception as e:                       # pragma: no cover - lib/env
        raise ExtractionError(f"Could not read that PDF: {e}")
    if not any(p.strip() for p in pages):
        raise ExtractionError(
            "No text found in that PDF — it looks scanned. Scanned-BOQ support "
            "is coming; for now upload a digital (text) PDF or the Excel.")
    return pages


def excel_pages(fileobj):
    """One 'page' of tab-separated text per worksheet (all sheets).

    Hidden columns and rows are skipped: in client tender/comparison workbooks
    those hold the other vendors' prices and working rows, not the BOQ we're
    capturing — feeding them to the model mixes up which rate/amount is ours.
    Fully-empty columns are dropped too so the real BOQ columns stay adjacent
    (spreadsheets often have spacer columns between fields)."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    try:                                    # not read_only: need hidden flags
        wb = load_workbook(fileobj, data_only=True)
    except Exception as e:
        raise ExtractionError(f"Could not read that Excel file: {e}")
    pages = []
    for ws in wb.worksheets:
        hidden_cols = {c for c, dim in ws.column_dimensions.items() if dim.hidden}
        vis = [i for i in range(1, (ws.max_column or 0) + 1)
               if get_column_letter(i) not in hidden_cols]
        rows = []
        for r, raw in enumerate(ws.iter_rows(values_only=True), start=1):
            if ws.row_dimensions[r].hidden:
                continue
            cells = ["" if raw[i - 1] is None else str(raw[i - 1]) for i in vis]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if not rows:
            continue
        keep = [j for j in range(len(vis))
                if any(row[j].strip() for row in rows)]   # drop empty columns
        lines = [f"# Sheet: {ws.title}"]
        lines += ["\t".join(row[j] for j in keep) for row in rows]
        if len(lines) > 1:
            pages.append("\n".join(lines))
    if not pages:
        raise ExtractionError("That workbook has no data rows.")
    return pages


def _batches(pages, max_chars=_MAX_BATCH_CHARS):
    """Group pages into batches under the size cap so big BOQs are extracted in
    several model calls (page boundaries never split a row)."""
    batches, cur, size = [], [], 0
    for i, text in enumerate(pages):
        if cur and size + len(text) > max_chars:
            batches.append(cur)
            cur, size = [], 0
        cur.append((i + 1, text))
        size += len(text)
    if cur:
        batches.append(cur)
    return batches


# ---- model structuring ---------------------------------------------------

_SYSTEM = (
    "You extract a construction Bill of Quantities (BOQ) from the given "
    "document text into structured rows. Rules:\n"
    "- COMPLETENESS IS CRITICAL: return EVERY line as its own row, in document "
    "order — never skip, merge, abbreviate, or summarise. A tender BOQ repeats "
    "many near-identical lines (the same work for different areas / villas / "
    "blocks); return each one. Returning only some of the lines is a failure.\n"
    "- Return one row per BOQ line, in document order.\n"
    "- If a line's rate is a note rather than a figure ('Included', 'Rate "
    "only', 'PS', 'Provisional'), keep that text as its rate.\n"
    "- A bill/section/trade title or a preamble note with no quantity or rate "
    "is a heading (is_heading=true); set its description and leave money "
    "fields empty.\n"
    "- For a priced item, fill item_code, description, unit, qty and the "
    "rate. If the schedule prices material (supply) and labour (install) "
    "separately, use rate_supply and rate_install; if it's a single rate, use "
    "rate_combined. Never invent a split that isn't in the document.\n"
    "- Carry the current bill/section title onto each item's `section`.\n"
    "- Skip page headers/footers, column headers, page numbers, and running "
    "'brought forward'/'carried forward' subtotal lines.\n"
    "- Keep numbers as plain digits (no thousands separators, no currency).\n"
    "- If you see a printed bill or grand total figure, report it in "
    "printed_totals so the import can be reconciled.\n"
    "- rate_mode is SPLIT if the document prices supply and labour separately, "
    "else SINGLE."
)

_TOOL = {
    "name": "emit_boq",
    "description": "Return the structured BOQ rows extracted from the document.",
    "input_schema": {
        "type": "object",
        "properties": {
            "rate_mode": {"type": "string", "enum": ["SINGLE", "SPLIT"]},
            "rows": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "section": {"type": "string"},
                        "item_code": {"type": "string"},
                        "description": {"type": "string"},
                        "unit": {"type": "string"},
                        "qty": {"type": "string"},
                        "rate_supply": {"type": "string"},
                        "rate_install": {"type": "string"},
                        "rate_combined": {"type": "string"},
                        "is_heading": {"type": "boolean"},
                        "page": {"type": "integer"},
                    },
                    "required": ["description"],
                },
            },
            "printed_totals": {
                "type": "array",
                "items": {"type": "number"},
                "description": "Any printed bill/page/grand total figures seen.",
            },
        },
        "required": ["rate_mode", "rows"],
    },
}


def _model_name():
    from .models import CompanyParameter
    try:
        v = CompanyParameter.objects.get(key="boq_extract_model").value
        return (v or "").strip() or DEFAULT_MODEL
    except CompanyParameter.DoesNotExist:
        return DEFAULT_MODEL


def _call_claude(content, model):
    """One structured extraction call. Isolated so tests can monkeypatch it."""
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        raise ExtractionError(
            "BOQ extraction needs an ANTHROPIC_API_KEY — ask the administrator "
            "to set it in the server environment.")
    try:
        import anthropic
    except ImportError:                          # pragma: no cover - env dep
        raise ExtractionError("The anthropic SDK isn't installed on the server.")
    try:
        client = anthropic.Anthropic(api_key=key)
        msg = client.messages.create(
            model=model, max_tokens=_MAX_OUTPUT_TOKENS, system=_SYSTEM,
            tools=[_TOOL],
            tool_choice={"type": "tool", "name": "emit_boq"},
            messages=[{"role": "user", "content": content}])
    except Exception as e:                       # pragma: no cover - network
        raise ExtractionError(f"The extraction model failed: {e}")
    for block in msg.content:
        if getattr(block, "type", None) == "tool_use":
            return block.input
    raise ExtractionError("The model returned no structured rows.")


def structure(pages, model=None):
    """Run the (possibly batched) model extraction over the document pages.
    Returns {rate_mode, rows, printed_totals}.

    Each batch is one independent model call, so a big BOQ's batches run
    CONCURRENTLY — wall time is the slowest single call, not the sum. A large
    bill that took 3 sequential calls (and blew the request timeout) now
    finishes in roughly one call's time. Order is preserved so rows stay in
    document order."""
    from concurrent.futures import ThreadPoolExecutor
    model = model or _model_name()
    batches = _batches(pages)

    def run(batch):
        header = ("Extract the BOQ from this document text. Page markers are "
                  "shown as [PAGE n].\n\n")
        body = "\n\n".join(f"[PAGE {n}]\n{text}" for n, text in batch)
        return _call_claude(header + body, model) or {}

    if len(batches) <= 1:
        outs = [run(b) for b in batches]
    else:
        with ThreadPoolExecutor(max_workers=min(len(batches), 5)) as ex:
            outs = list(ex.map(run, batches))       # ex.map preserves order

    all_rows, totals, modes = [], [], []
    for out in outs:
        all_rows.extend(out.get("rows", []))
        totals.extend(out.get("printed_totals") or [])
        if out.get("rate_mode"):
            modes.append(out["rate_mode"])
    rate_mode = "SPLIT" if "SPLIT" in modes else (modes[0] if modes else "")
    return {"rate_mode": rate_mode, "rows": all_rows,
            "printed_totals": totals, "model": model}


# ---- normalisation, warnings, reconciliation -----------------------------

FIELDS = ("section", "item_code", "description", "unit", "qty",
          "rate_supply", "rate_install", "rate_combined", "is_heading")


def _num(v):
    """A cleaned numeric string (strip commas/currency/spaces) or ""."""
    if v is None:
        return ""
    s = str(v).strip().replace(",", "")
    for junk in ("$", "USD", "MVR", "Rf", "rf"):
        s = s.replace(junk, "")
    s = s.strip()
    if not s:
        return ""
    try:
        Decimal(s)
        return s
    except (InvalidOperation, ValueError):
        return s          # keep the raw text; a warning will flag it


def _dec_or_none(s):
    try:
        return Decimal(s) if s not in ("", None) else None
    except (InvalidOperation, ValueError):
        return None


def normalise_rows(rows):
    """Coerce model rows to the committer's shape with clean numbers."""
    out = []
    for r in rows:
        row = {
            "section": str(r.get("section") or "").strip(),
            "item_code": str(r.get("item_code") or "").strip(),
            "description": str(r.get("description") or "").strip(),
            "unit": str(r.get("unit") or "").strip(),
            "qty": _num(r.get("qty")),
            "rate_supply": _num(r.get("rate_supply")),
            "rate_install": _num(r.get("rate_install")),
            "rate_combined": _num(r.get("rate_combined")),
            "is_heading": bool(r.get("is_heading")),
        }
        if not (row["description"] or row["section"] or row["item_code"]):
            continue
        out.append(row)
    return out


def row_warnings(row):
    """Per-row review flags. Headings are exempt."""
    if row.get("is_heading"):
        return []
    w = []
    qty = _dec_or_none(row.get("qty"))
    supply = _dec_or_none(row.get("rate_supply"))
    install = _dec_or_none(row.get("rate_install"))
    combined = _dec_or_none(row.get("rate_combined"))
    has_rate = any(x is not None for x in (supply, install, combined))
    # a value present but unparseable
    for f in ("qty", "rate_supply", "rate_install", "rate_combined"):
        val = row.get(f)
        if val and _dec_or_none(val) is None:
            w.append(f"{f.replace('_', ' ')} isn't a number")
    if (qty is not None or has_rate) and not row.get("unit"):
        w.append("missing unit")
    if qty is not None and not has_rate:
        w.append("no rate")
    if has_rate and qty is None:
        w.append("no quantity")
    return w


def _row_amount(row):
    if row.get("is_heading"):
        return Decimal("0")
    qty = _dec_or_none(row.get("qty")) or Decimal("0")
    supply = _dec_or_none(row.get("rate_supply")) or Decimal("0")
    install = _dec_or_none(row.get("rate_install")) or Decimal("0")
    combined = _dec_or_none(row.get("rate_combined")) or Decimal("0")
    rate = combined if combined else (supply + install)
    return qty * rate


def extracted_total(rows):
    return sum((_row_amount(r) for r in rows), Decimal("0"))


def reconcile(rows, printed_totals):
    """Compare the sum of extracted line amounts to the largest printed total
    the model saw. Returns a dict for the review banner."""
    got = extracted_total(rows)
    printed = None
    nums = []
    for t in printed_totals or []:
        d = _dec_or_none(str(t))
        if d is not None:
            nums.append(d)
    if nums:
        printed = max(nums)
    ok = None
    if printed and printed > 0:
        ok = abs(got - printed) <= printed * Decimal("0.005")   # 0.5%
    return {
        "extracted_total": str(got.quantize(Decimal("0.01"))),
        "printed_total": (str(printed.quantize(Decimal("0.01")))
                          if printed is not None else None),
        "reconciled": ok,
    }


# ---- orchestration -------------------------------------------------------

def run_import(project, upload, actor):
    """Read + extract an uploaded BOQ file into a draft BoqImport. Returns
    (boq_import, error)."""
    from .models import BoqImport
    name = (getattr(upload, "name", "") or "").lower()
    if name.endswith(".pdf"):
        source, pages = "PDF", pdf_pages(upload)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        source, pages = "XLSX", excel_pages(upload)
    else:
        return None, "Upload a BOQ as a PDF or an Excel (.xlsx) file."
    result = structure(pages)
    rows = normalise_rows(result["rows"])
    if not rows:
        return None, ("No BOQ lines were found in that document — check it's "
                      "the priced schedule and try again.")
    rec = reconcile(rows, result["printed_totals"])
    imp = BoqImport.objects.create(
        project=project, source=source,
        filename=getattr(upload, "name", "") or "",
        rate_mode=result.get("rate_mode") or "", rows=rows,
        meta={"page_count": len(pages), "model": result.get("model"),
              **rec},
        created_by=actor)
    return imp, None


def commit(boq_import, actor):
    """Load the (reviewed) draft rows into the live BOQ via the shared
    committer. Returns (boq, error)."""
    from . import commercial
    from .models import BoqImport
    if boq_import.status == BoqImport.Status.COMMITTED:
        return None, "This import has already been loaded into the BOQ."
    boq, msg = commercial.import_boq_rows(
        boq_import.project, boq_import.rows, actor)
    if msg:
        return None, msg
    boq_import.status = BoqImport.Status.COMMITTED
    boq_import.save(update_fields=["status", "updated_at"])
    return boq, None


def import_payload(boq_import):
    """Serialise a draft for review, with per-row warnings + totals."""
    rows = []
    warn_count = 0
    for i, r in enumerate(boq_import.rows):
        w = row_warnings(r)
        warn_count += len(w)
        rows.append({**r, "i": i, "warnings": w,
                     "amount": str(_row_amount(r).quantize(Decimal("0.01")))})
    return {
        "id": boq_import.id, "status": boq_import.status,
        "source": boq_import.source, "filename": boq_import.filename,
        "rate_mode": boq_import.rate_mode, "rows": rows,
        "warning_count": warn_count, "meta": boq_import.meta,
    }
