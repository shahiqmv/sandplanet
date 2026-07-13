from datetime import date, timedelta

from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from .audit import audit
from .models import (
    Approval,
    Attachment,
    Document,
    DocumentRevision,
    Holiday,
    Item,
    PendingItem,
    ProgrammeActivity,
    Project,
    Site,
    User,
)
from .numbering import next_ref
from .pdf import generate_pdf
from .permissions import scoped_site_ids
from .procurement import (
    sync_pr_vendor_rows,
    grn_lines_from_lm,
    link_documents,
    on_grn_verified,
    on_lm_departed,
    on_pr_approved,
    po_lm_prefill_lines,
    resolve_refs,
    save_lines,
    validate_mr_lines,
)
from .views_quotes import pr_coverage_data
from .serializers_documents import (
    AttachmentSerializer,
    DocumentSerializer,
    PendingItemSerializer,
)

CREATE_ROLES = {  # spec §3 "can create"
    "DPR": {"SITE_ENGINEER", "SITE_ADMIN", "PM"},
    "TWS": {"SITE_ENGINEER", "PM"},
    "IR": {"SITE_ENGINEER", "PM"},       # SE submits with QA/QC
    "MAR": {"SITE_ENGINEER", "PM"},      # SE submits with QS
    # Site Engineer has full site-task parity with Site Admin (owner,
    # 2026-07-13): both raise MRs, receive goods, etc.
    "MR": {"SITE_ADMIN", "SITE_ENGINEER", "PM"},
    "GRN": {"SITE_ADMIN", "SITE_ENGINEER", "PM"},
    "PR": {"HO_PURCHASING"},             # Head Office documents
    "LM": {"HO_PURCHASING"},
    "PO": {"HO_PURCHASING"},             # normally auto-generated (R2)
    "DMA": {"SITE_ENGINEER", "PM"},      # SE may prepare; PM issues (R5)
    "PYR": {"SITE_ADMIN", "SITE_ENGINEER", "PM"},  # payment request (M6)
    "PMR": {"SITE_ENGINEER", "SITE_ADMIN", "PM"},  # import requirement (§5.10)
}
SITE_TEAM = {"SITE_ENGINEER", "SITE_ADMIN", "PM"}  # record client results (dec. 1)
RESULTS = {  # client results per type (spec §5.3/§5.4)
    "IR": {"APPROVED", "APPROVED_WITH_COMMENTS", "REJECTED"},
    "MAR": {"APPROVED", "APPROVED_WITH_COMMENTS", "REVISE_RESUBMIT", "REJECTED"},
}
LINE_TYPES = {"MR", "PR", "LM", "GRN", "PMR"}
MIN_DPR_PHOTOS = 0  # owner (Phase C): no photo floor — attach any number


def _serialize(doc, request):
    return DocumentSerializer(doc, context={"request": request}).data


def _get_scoped_document(request, ref):
    try:
        doc = Document.objects.select_related("site", "current_revision").get(ref=ref)
    except Document.DoesNotExist:
        return None, Response({"detail": "Not found."}, status=404)
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None and doc.site_id not in site_ids:
        return None, Response({"detail": "Not found."}, status=404)
    return doc, None


def _is_site_pm(user, site):
    pm = site.current_pm()
    return user.role == User.Role.PM and pm is not None and pm.id == user.id


def _is_pm_for(user, doc):
    """PM gate for a document: the PROJECT PM when the document belongs to
    a project that has one (R4 management round); the site PM always
    remains authorised as fallback."""
    if user.role != User.Role.PM:
        return False
    if doc.project_id and doc.project.pm_id and doc.project.pm_id == user.id:
        return True
    return _is_site_pm(user, doc.site)


def _can(request, doc_type, roles):
    return request.user.role in roles or request.user.role == "ADMIN"


def _previous_revision(doc):
    return (
        doc.revisions.exclude(pk=doc.current_revision_id).order_by("-id").first()
    )


@api_view(["POST"])
def document_create(request):
    doc_type = request.data.get("doc_type")
    if doc_type not in Document.TRANSITIONS:
        return Response({"detail": f"Unsupported doc_type '{doc_type}'."}, status=400)
    try:
        site = Site.objects.get(pk=request.data.get("site_id"))
    except Site.DoesNotExist:
        return Response({"detail": "Unknown site_id."}, status=400)

    site_ids = scoped_site_ids(request.user)
    if site_ids is not None and site.id not in site_ids:
        return Response({"detail": "Not allocated to this site."}, status=403)
    if not _can(request, doc_type, CREATE_ROLES[doc_type]):
        return Response({"detail": "Role cannot create this document."}, status=403)
    # Lifecycle rules (spec §2.2); MAR/MR may begin at AWARDED (mobilization)
    if site.status == Site.Status.CLOSED:
        return Response({"detail": "Site is closed — no new documents."}, status=400)
    if site.status == Site.Status.ON_HOLD and request.user.role not in ("PM", "ADMIN") \
            and not request.user.is_ho:
        return Response(
            {"detail": "Site on hold — only PM/HO can create documents."}, status=403
        )

    # IR/MAR belong to a project (R4). DPR/TWS are SITE-WIDE — one daily
    # report to the one client per site, each work/planned row tagged with
    # its project (owner, R8 2026-07-08; supersedes R4's per-project DPR).
    PROJECT_TYPES = ("IR", "MAR", "PMR")  # PMR imports are raised per project
    project = None
    if doc_type in PROJECT_TYPES:
        project_id = request.data.get("project_id")
        if project_id:
            try:
                project = Project.objects.get(pk=project_id, site=site)
            except Project.DoesNotExist:
                return Response({"detail": "Unknown project for this site."},
                                status=400)
            if project.status == "CLOSED":
                return Response({"detail": "Project is closed — no new "
                                           "documents."}, status=400)
        elif site.projects.filter(status="ACTIVE").exists():
            return Response({"detail": "Select the project this document "
                                       "belongs to."}, status=400)

    doc_date = request.data.get("doc_date") or date.today().isoformat()
    if doc_type in ("DPR", "TWS", "DMA"):  # one per SITE per day (R8/R5)
        clash = Document.objects.filter(
            doc_type=doc_type, site=site, doc_date=doc_date, is_void=False
        ).first()
        if clash:
            return Response(
                {"detail": f"{clash.ref} already exists for {doc_date}."}, status=400
            )

    # IR resubmission: new number quoting the previous, rejected IR (§4.2)
    previous_ir = None
    if doc_type == "IR" and request.data.get("previous_ir_ref"):
        try:
            previous_ir = Document.objects.get(
                ref=request.data["previous_ir_ref"], doc_type="IR", is_void=False
            )
        except Document.DoesNotExist:
            return Response({"detail": "Unknown previous IR."}, status=400)
        if previous_ir.status != "REJECTED":
            return Response(
                {"detail": f"{previous_ir.ref} is {previous_ir.status} — only a "
                           f"rejected IR is resubmitted under a new number."},
                status=400,
            )
        if previous_ir.site_id != site.id:
            return Response({"detail": "Previous IR is for a different site."},
                            status=400)

    payload = request.data.get("payload") or {}
    lines_data = request.data.get("lines") or []

    # GRN is normally raised FROM a manifest: prefill lines + link (spec §5.6).
    # Without a manifest it captures the site's existing / opening stock, keyed
    # in by the site team (owner, temporary) — no LM link, lines added manually.
    lm = None
    if doc_type == "GRN":
        lm_ref = request.data.get("lm_ref") or payload.get("manifest_ref")
        if lm_ref:
            try:
                lm = Document.objects.get(ref=lm_ref, doc_type="LM",
                                          is_void=False)
            except Document.DoesNotExist:
                return Response({"detail": f"Unknown manifest '{lm_ref}'."},
                                status=400)
            if lm.site_id != site.id:
                return Response({"detail": "Manifest is for a different site."},
                                status=400)
            if not lines_data:
                lines_data = grn_lines_from_lm(lm)
            payload.setdefault("manifest_ref", lm.ref)
            payload.setdefault("vessel",
                               (lm.current_revision.payload or {}).get(
                                   "vessel", ""))
            payload.setdefault(
                "mr_refs",
                list(lm.links_from.filter(link_type="MR_LM")
                     .values_list("to_document__ref", flat=True)),
            )
        else:
            payload.setdefault("opening_stock", True)

    if doc_type == "MR":
        error = validate_mr_lines(lines_data)
        if error:
            return Response({"detail": error}, status=400)

    # Cross-references arrive as refs, stored as FK links (spec §4.3)
    mr_docs, missing = resolve_refs(request.data.get("mr_refs"), "MR")
    if missing:
        return Response({"detail": f"Unknown MR refs: {sorted(missing)}"}, status=400)
    pr_docs, missing = resolve_refs(request.data.get("pr_refs"), "PR")
    if missing:
        return Response({"detail": f"Unknown PR refs: {sorted(missing)}"}, status=400)
    po_docs, missing = resolve_refs(request.data.get("po_refs"), "PO")
    if missing:
        return Response({"detail": f"Unknown PO refs: {sorted(missing)}"}, status=400)
    # A PR/LM belongs to the same site/project as the MRs it answers
    wrong = [d.ref for d in mr_docs + pr_docs + po_docs if d.site_id != site.id]
    if wrong:
        return Response(
            {"detail": f"References belong to a different site: {wrong}. "
                       f"Pick the site the MR was raised for."}, status=400
        )

    with transaction.atomic():
        ref = next_ref(doc_type, site)  # locks counter until commit — gap-free
        doc = Document.objects.create(
            doc_type=doc_type, ref=ref, site=site, doc_date=doc_date,
            status="DRAFT", created_by=request.user, previous_ir=previous_ir,
            project=project,
        )
        revision = DocumentRevision.objects.create(
            document=doc, rev_label="R0", payload=payload, created_by=request.user
        )
        doc.current_revision = revision
        doc.save(update_fields=["current_revision"])
        if doc_type in LINE_TYPES:
            save_lines(revision, lines_data)
        for mr in mr_docs:
            link_documents(doc, mr, "MR_PR" if doc_type == "PR" else "MR_LM")
        for pr in pr_docs:
            link_documents(doc, pr, "PR_LM")
        for po in po_docs:
            link_documents(doc, po, "PO_LM")
        if lm is not None:
            link_documents(doc, lm, "LM_GRN")
        if doc_type == "PYR":  # typed payment-request row alongside (§5.9)
            from .payments import create_payment_request

            pr, pyr_err = create_payment_request(doc, request.data,
                                                 request.user)
            if pyr_err:
                transaction.set_rollback(True)
                return Response({"detail": pyr_err}, status=400)
    audit("document", doc.id, "DOC_CREATED", actor=request.user,
          to_state="DRAFT", detail={"ref": ref})
    return Response(_serialize(doc, request), status=201)


@api_view(["GET", "PATCH"])
def document_detail(request, ref):
    doc, err = _get_scoped_document(request, ref)
    if err:
        return err
    if request.method == "GET":
        return Response(_serialize(doc, request))

    # PATCH — draft revisions only (issued revisions immutable, spec §7.2)
    revision = doc.current_revision
    if doc.is_void or doc.status != "DRAFT" or revision.issued_at is not None:
        return Response({"detail": "Only draft documents can be edited."}, status=400)
    if not _can(request, doc.doc_type, CREATE_ROLES.get(doc.doc_type, set())):
        return Response({"detail": "Role cannot edit this document."}, status=403)
    if "payload" in request.data:
        revision.payload = request.data["payload"]
        revision.save(update_fields=["payload"])
    if "lines" in request.data and doc.doc_type in LINE_TYPES:
        if doc.doc_type == "MR":
            error = validate_mr_lines(request.data["lines"])
            if error:
                return Response({"detail": error}, status=400)
        save_lines(revision, request.data["lines"],
                   previous_revision=_previous_revision(doc))
    if "doc_date" in request.data and doc.doc_type == "DPR":
        clash = Document.objects.filter(
            doc_type="DPR", site=doc.site, doc_date=request.data["doc_date"],
            is_void=False,
        ).exclude(pk=doc.pk).first()
        if clash:
            return Response(
                {"detail": f"{clash.ref} already exists for that date."}, status=400
            )
        doc.doc_date = request.data["doc_date"]
        doc.save(update_fields=["doc_date"])
    elif "doc_date" in request.data:
        doc.doc_date = request.data["doc_date"]
        doc.save(update_fields=["doc_date"])
    doc.refresh_from_db()
    return Response(_serialize(doc, request))


@api_view(["POST"])
def document_revise(request, ref):
    """MR amendment / MAR resubmission: same number, next revision label,
    restarts at Draft; previous revision stays visible (spec §4.2)."""
    doc, err = _get_scoped_document(request, ref)
    if err:
        return err
    if doc.doc_type not in ("MR", "MAR"):
        return Response({"detail": "Only MR/MAR use revisions; IR gets a new "
                                   "number quoting the previous IR."}, status=400)
    if doc.is_void:
        return Response({"detail": "Document is void."}, status=400)
    if doc.status == "DRAFT":
        return Response({"detail": "Document is already an editable draft."},
                        status=400)
    if doc.status == "CLOSED":
        return Response({"detail": "Closed documents cannot be amended."}, status=400)
    if not _can(request, doc.doc_type, CREATE_ROLES.get(doc.doc_type, set())):
        return Response({"detail": "Role cannot amend this document."}, status=403)

    old_revision = doc.current_revision
    next_label = f"R{int(old_revision.rev_label[1:]) + 1}"
    with transaction.atomic():
        old_revision.is_current = False
        old_revision.save(update_fields=["is_current"])
        # Carry contractor content; the client's outcome belongs to the old
        # revision only — the new round gets a fresh result (spec §4.2)
        carried = dict(old_revision.payload or {})
        for workflow_key in ("client_result", "closure", "acknowledgement"):
            carried.pop(workflow_key, None)
        new_revision = DocumentRevision.objects.create(
            document=doc, rev_label=next_label,
            payload=carried, created_by=request.user,
        )
        # carry lines over unflagged; edits via PATCH mark is_changed (§5.5 r3)
        save_lines(new_revision, [
            {"item_id": line.item_id, "free_text_desc": line.free_text_desc,
             "unit": line.unit,
             "qty_required": line.qty_required, "qty_stock": line.qty_stock,
             "qty_to_order": line.qty_to_order, "priority": line.priority,
             "urgent_reason": line.urgent_reason, "remarks": line.remarks}
            for line in old_revision.lines.all()
        ])
        old_status = doc.status
        doc.current_revision = new_revision
        doc.status = "DRAFT"
        doc.save(update_fields=["current_revision", "status", "updated_at"])
    audit("document", doc.id, "DOC_REVISED", actor=request.user,
          from_state=old_status, to_state="DRAFT",
          detail={"ref": doc.ref, "rev": next_label})
    return Response(_serialize(doc, request), status=201)


# ===== Workflow actions =====


def _transition(doc, new_status):
    allowed = Document.TRANSITIONS[doc.doc_type].get(doc.status, set())
    return new_status in allowed


def _record(doc, action, request, comment="", result=""):
    Approval.objects.create(
        document=doc, revision=doc.current_revision, action=action,
        result=result, actor=request.user, actor_role=request.user.role,
        comment=comment,
    )


def _apply(request, doc, new_status, action, roles=None, pm_gate=False,
           lock_revision=False, pdf_milestone=None, comment=""):
    if pm_gate:
        if not (_is_pm_for(request.user, doc) or
                request.user.role == "ADMIN"):
            return Response({"detail": "Only the project's PM (or the "
                                       "site's PM) can do this."},
                            status=403)
    elif roles is not None and not _can(request, doc.doc_type, roles):
        return Response({"detail": f"Role cannot {action.lower()} this document."},
                        status=403)
    if not _transition(doc, new_status):
        return Response({"detail": f"Cannot {action.lower()} from {doc.status}."},
                        status=400)
    old = doc.status
    with transaction.atomic():
        if lock_revision and doc.current_revision.issued_at is None:
            doc.current_revision.issued_at = timezone.now()
            doc.current_revision.save(update_fields=["issued_at"])
        doc.status = new_status
        doc.save(update_fields=["status", "updated_at"])
        _record(doc, action, request, comment=comment)
    audit("document", doc.id, f"DOC_{action}", actor=request.user,
          from_state=old, to_state=new_status, detail={"ref": doc.ref})
    if pdf_milestone:
        generate_pdf(doc, doc.current_revision, pdf_milestone)
    return None


@api_view(["POST"])
def document_action(request, ref, action_name):
    doc, err = _get_scoped_document(request, ref)
    if err:
        return err
    if doc.is_void:
        return Response({"detail": "Document is void."}, status=400)
    if doc.doc_type == "PYR":  # dedicated payment-request workflow (§5.9)
        from .payments import pyr_action

        result = pyr_action(request, doc, action_name)
        if isinstance(result, Response):
            return result
        doc.refresh_from_db()
        return Response(_serialize(doc, request))
    handler = {
        "issue": _do_issue, "verify": _do_verify, "void": _do_void,
        "submit": _do_submit, "approve": _do_approve, "return": _do_return,
        "authorise": _do_authorise,
        "withdraw-authorisation": _do_withdraw,
        "send": _do_send, "depart": _do_depart, "count": _do_count,
        "close": _do_close,
        "record-result": _do_record_result, "client-verify": _do_client_verify,
        "acknowledge": _do_acknowledge,
        "ho-review": _do_ho_review, "size-release": _do_size_release,
        "cancel": _do_cancel,
    }.get(action_name)
    if handler is None:
        return Response({"detail": f"Unknown action '{action_name}'."}, status=400)
    comment = request.data.get("comment", "") if hasattr(request, "data") else ""
    result = handler(request, doc, comment)
    if isinstance(result, Response):
        return result
    doc.refresh_from_db()
    return Response(_serialize(doc, request))


def _do_issue(request, doc, comment):
    if doc.doc_type == "DPR" and MIN_DPR_PHOTOS:
        captioned = doc.attachments.filter(kind="PHOTO").exclude(caption="").count()
        if captioned < MIN_DPR_PHOTOS:
            return Response(
                {"detail": f"DPR needs at least {MIN_DPR_PHOTOS} captioned photos "
                           f"to issue ({captioned} attached)."},
                status=400,
            )
    if doc.doc_type == "DMA":
        # The morning allocation is the PM's call — SE may prepare it,
        # but only the project/site PM issues (R5).
        return _apply(request, doc, "ISSUED", "ISSUE", pm_gate=True,
                      lock_revision=True, pdf_milestone="issue",
                      comment=comment)
    if doc.doc_type in ("DPR", "TWS", "IR", "MAR", "PO"):
        # DPR/TWS/PO issue from DRAFT; IR/MAR issue after the PM gate (§7.1)
        err = _apply(request, doc, "ISSUED", "ISSUE",
                     roles=CREATE_ROLES[doc.doc_type], lock_revision=True,
                     pdf_milestone="issue", comment=comment)
        if err is None and doc.doc_type == "DPR":
            _update_programme_progress(doc, request.user)
            _post_dpr_consumption(doc, request.user)
        return err
    return Response({"detail": "Issue applies to DPR/TWS/IR/MAR/PO/DMA."},
                    status=400)


@api_view(["GET"])
def approvals_pending(request):
    """The per-role 'waiting on you' queue (owner, 2026-07-08): each
    approver's landing page lists exactly the documents blocked on them."""
    user = request.user
    site_ids = scoped_site_ids(user)
    groups = []

    def scoped(qs):
        return qs.filter(site_id__in=site_ids) if site_ids is not None else qs

    def rows(qs, hint):
        return [{"ref": d.ref, "doc_type": d.doc_type,
                 "site_code": d.site.code, "project_code":
                 d.project.code if d.project else None,
                 "doc_date": d.doc_date, "status": d.status, "hint": hint}
                for d in qs.select_related("site", "project")[:50]]

    def add(title, items):
        if items:
            groups.append({"title": title, "items": items})

    base = Document.objects.filter(is_void=False).order_by("doc_date", "id")
    if user.role in ("PM", "ADMIN"):
        mine = [d for d in scoped(base.filter(
                    doc_type__in=("MR", "IR", "MAR"), status="SUBMITTED"))
                .select_related("site", "project")
                if user.role == "ADMIN" or _is_pm_for(user, d)]
        add("To approve — submitted MR / IR / MAR",
            [{"ref": d.ref, "doc_type": d.doc_type, "site_code": d.site.code,
              "project_code": d.project.code if d.project else None,
              "doc_date": d.doc_date, "status": d.status,
              "hint": "PM approval"} for d in mine])
        pyrs = [d for d in scoped(base.filter(doc_type="PYR",
                                              status="SUBMITTED"))
                .select_related("site")
                if user.role == "ADMIN" or _is_pm_for(user, d)]
        add("To approve — submitted payment requests",
            [{"ref": d.ref, "doc_type": "PYR", "site_code": d.site.code,
              "project_code": None, "doc_date": d.doc_date,
              "status": d.status, "hint": "PM approval"} for d in pyrs])
        pmrs = [d for d in scoped(base.filter(doc_type="PMR",
                                              status="SUBMITTED"))
                .select_related("site", "project")
                if user.role == "ADMIN" or _is_pm_for(user, d)]
        add("To approve — submitted import requests (PMR)",
            [{"ref": d.ref, "doc_type": "PMR", "site_code": d.site.code,
              "project_code": d.project.code if d.project else None,
              "doc_date": d.doc_date, "status": d.status,
              "hint": "PM approval"} for d in pmrs])
        dprs = [d for d in scoped(base.filter(doc_type="DPR",
                                              status="ISSUED"))
                .select_related("site", "project")
                if user.role == "ADMIN" or _is_pm_for(user, d)]
        add("To verify — issued DPRs",
            [{"ref": d.ref, "doc_type": "DPR", "site_code": d.site.code,
              "project_code": d.project.code if d.project else None,
              "doc_date": d.doc_date, "status": d.status,
              "hint": "PM verification"} for d in dprs])
        dmas = [d for d in scoped(base.filter(doc_type="DMA",
                                              status="DRAFT"))
                .select_related("site")
                if user.role == "ADMIN" or _is_pm_for(user, d)]
        add("To issue — morning manpower allocations",
            [{"ref": d.ref, "doc_type": "DMA", "site_code": d.site.code,
              "project_code": None, "doc_date": d.doc_date,
              "status": d.status, "hint": "PM issues the allocation"}
             for d in dmas])
    if user.role in ("DIRECTOR", "ADMIN"):
        add("To award — submitted PRs",
            rows(scoped(base.filter(doc_type="PR", status="SUBMITTED")),
                 "Director approval awards the suppliers"))
        add("To approve — PM-approved payment requests",
            rows(scoped(base.filter(doc_type="PYR", status="PM_APPROVED")),
                 "Director approval of the requisition"))
        add("To size & release — reviewed import requests (PMR)",
            rows(scoped(base.filter(doc_type="PMR", status="HO_REVIEWED")),
                 "Size the order (MOQ) and release to sourcing"))
    if user.role in ("SIGNATORY", "ADMIN"):
        # Signatory approves whole Payment Vouchers (M6d), not each doc
        add("To approve — payment vouchers",
            [{"ref": pv.ref, "doc_type": "PV", "site_code": "—",
              "project_code": None, "doc_date": pv.doc_date,
              "status": pv.status,
              "hint": "Approve the batch or query lines"}
             for pv in base.filter(doc_type="PV", status="SUBMITTED")[:50]])
    if user.role in ("HO_PURCHASING", "ADMIN"):
        add("To action — MRs sent to Head Office",
            rows(scoped(base.filter(doc_type="MR", status="SENT_TO_HO")),
                 "Raise PR or plan loading"))
        add("To issue — draft POs",
            rows(scoped(base.filter(doc_type="PO", status="DRAFT")),
                 "Issue to supplier"))
        add("To review — PM-approved import requests (PMR)",
            rows(scoped(base.filter(doc_type="PMR", status="PM_APPROVED")),
                 "Review the requirement before the Director sizes it"))
    if user.role in ("FINANCE", "HO_PURCHASING", "ADMIN"):
        add("Payments pending — authorised PRs",
            rows(scoped(base.filter(doc_type="PR",
                                    status__in=("AUTHORISED",
                                                "PAYMENT_PROCESSING"))),
                 "Record vendor payments / PO refs"))
    if user.role in ("FINANCE", "ADMIN"):
        # Finance builds vouchers from Director-approved requisitions and
        # pays authorised ones (M6d)
        from .vouchers import awaiting_voucher

        awaiting = awaiting_voucher()
        if awaiting:
            add("Awaiting a payment voucher",
                [{"ref": d.ref, "doc_type": d.doc_type,
                  "site_code": d.site.code, "project_code": None,
                  "doc_date": d.doc_date, "status": d.status,
                  "hint": "Add to a payment voucher for the signatory"}
                 for d in awaiting[:50]])
        add("To pay — authorised payment requests",
            rows(scoped(base.filter(doc_type="PYR", status="AUTHORISED")),
                 "Execute payment and record the reference"))
    return Response({"groups": groups,
                     "total": sum(len(g["items"]) for g in groups)})


@api_view(["GET"])
def dma_prefill(request):
    """Task rows for the morning allocation, pulled from the TWSs scheduled
    FOR the given day (issued the previous evening, any project) (R5)."""
    site_id = request.GET.get("site")
    day = request.GET.get("date") or date.today().isoformat()
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None and int(site_id or 0) not in site_ids:
        return Response({"detail": "Not allocated to this site."}, status=403)
    rows = []
    tws_qs = Document.objects.filter(
        doc_type="TWS", site_id=site_id, doc_date=day, is_void=False,
        status__in=("ISSUED", "ACKNOWLEDGED"),
    ).select_related("project", "current_revision")
    for tws in tws_qs:
        fallback = tws.project.code if tws.project else ""  # legacy TWS
        for a in (tws.current_revision.payload or {}).get("activities", []):
            rows.append({
                "task": a.get("activity", ""), "location": a.get("location", ""),
                "category": a.get("trade", ""), "workers": "",
                "remarks": "", "project": a.get("project") or fallback,
                "source": tws.ref,
            })
    return Response({
        "tasks": rows,
        "tws_refs": [t.ref for t in tws_qs],
    })


def _update_programme_progress(doc, actor):
    """Issued DPR work-done rows carry cumulative %-to-date per programme
    activity — roll each row into ITS OWN project's programme (R8: the
    DPR is site-wide; rows are tagged per project)."""
    for row in (doc.current_revision.payload or {}).get("work_done", []):
        activity_id = row.get("activity_id")
        todate = row.get("progress_todate")
        if not activity_id or todate in (None, ""):
            continue
        try:
            activity = ProgrammeActivity.objects.select_related("project") \
                .get(pk=activity_id, project__site_id=doc.site_id)
        except ProgrammeActivity.DoesNotExist:
            continue
        new_value = max(min(float(todate), 100), 0)
        if float(activity.progress) == new_value:
            continue
        old_value = float(activity.progress)
        activity.progress = new_value
        activity.progress_updated_from = doc
        activity.save(update_fields=["progress", "progress_updated_from"])
        audit("programme_activity", activity.id, "PROGRESS_UPDATED",
              actor=actor, from_state=str(old_value), to_state=str(new_value),
              detail={"dpr": doc.ref, "activity": activity.name[:80]})


def _post_dpr_consumption(doc, actor):
    """Issued DPR key-material rows loaded from stock carry an item_id and a
    'consumed' figure — post each as an ISSUE so the site stock balance stays
    live from daily reporting. Idempotent: skips if this DPR already posted."""
    from . import stock
    from .models import StockMovement

    if StockMovement.objects.filter(document=doc,
                                    kind=StockMovement.Kind.ISSUE).exists():
        return  # already posted (guards a re-run of the issue transition)
    for row in (doc.current_revision.payload or {}).get("materials", []):
        item_id = row.get("item_id")
        consumed = row.get("consumed")
        if not item_id or consumed in (None, "", 0, "0"):
            continue
        try:
            qty = float(consumed)
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        try:
            item = Item.objects.get(pk=item_id)
        except Item.DoesNotExist:
            continue
        stock.consume(doc.site, item, qty, project=doc.project, document=doc,
                      actor=actor, movement_date=doc.doc_date,
                      reason=f"Consumed — DPR {doc.ref}")


def _do_submit(request, doc, comment):
    roles = {"MR": {"SITE_ADMIN", "SITE_ENGINEER", "PM"},
             "PR": {"HO_PURCHASING"},
             "IR": {"SITE_ENGINEER", "PM"},
             "MAR": {"SITE_ENGINEER", "PM"},
             "PMR": {"SITE_ENGINEER", "SITE_ADMIN", "PM"},
             "IPR": {"HO_PURCHASING"}}.get(doc.doc_type)
    if roles is None:
        return Response({"detail": "Submit applies to MR/PR/IR/MAR/PMR/IPR."},
                        status=400)
    if doc.doc_type == "PR" and doc.quotations.exists():
        # Coverage tally (R2): every MR line quoted AND awarded, or an
        # explicit override with a reason.
        rows = pr_coverage_data(doc)
        uncovered = [r["description"] for r in rows if not r["covered"]]
        unawarded = [r["description"] for r in rows
                     if r["covered"] and not r["awarded"]]
        if (uncovered or unawarded) and not request.data.get("allow_uncovered"):
            return Response({
                "detail": "MR lines are not fully quoted/awarded. Match and "
                          "award them, or resubmit with allow_uncovered and "
                          "a reason in the comment.",
                "uncovered": uncovered, "unawarded": unawarded,
            }, status=400)
        if (uncovered or unawarded) and not comment.strip():
            return Response({"detail": "A reason is required to submit with "
                                       "uncovered MR lines."}, status=400)
        # Vendor summary always reflects the quotes at submit time
        sync_pr_vendor_rows(doc)
    return _apply(request, doc, "SUBMITTED", "SUBMIT", roles=roles,
                  comment=comment)


def _do_approve(request, doc, comment):
    if doc.doc_type in ("MR", "IR", "MAR", "PMR"):  # PM gate (spec §5.3–§5.5)
        return _apply(request, doc, "PM_APPROVED", "APPROVE", pm_gate=True,
                      comment=comment)
    if doc.doc_type == "PR":  # Director approval = award (§5.7); NO
        # commitment or PO here — a signatory authorises next (§6C.2, M6c)
        err = _apply(request, doc, "APPROVED", "APPROVE",
                     roles={"DIRECTOR"}, lock_revision=True,
                     pdf_milestone="approved", comment=comment)
        if err is None:
            on_pr_approved(doc, request.user)
        return err
    if doc.doc_type == "IPR":  # Director awards the order; commitment is next
        # (a signatory authorises it on a Payment Voucher, §6C.2 / D1)
        err = _apply(request, doc, "APPROVED", "APPROVE",
                     roles={"DIRECTOR"}, comment=comment)
        if err is None:
            from .imports import advance_linked_pmrs
            advance_linked_pmrs(doc, "ORDERED", request.user)
        return err
    return Response({"detail": "Approve applies to MR/PR/IR/MAR/IPR."},
                    status=400)


def _do_authorise(request, doc, comment):
    """Retired at M6d — a PR is authorised on a Payment Voucher (a
    signatory approves a batch, not each PR individually)."""
    return Response({"detail": "PRs are authorised on a Payment Voucher "
                               "(Finance builds it, a signatory approves "
                               "it)."}, status=400)


def _do_withdraw(request, doc, comment):
    """Finance withdrawal of a PR's authorisation (§7.5b) — reverses the
    COMMITTED postings, cancels payables, returns to Draft."""
    from .procurement import reverse_pr_authorisation

    if doc.doc_type != "PR":
        return Response({"detail": "Withdraw applies to PR."}, status=400)
    if doc.status != "AUTHORISED":
        return Response({"detail": "Only an authorised, unpaid PR can have "
                                   "its authorisation withdrawn."}, status=400)
    if not comment.strip():
        return Response({"detail": "A reason is required."}, status=400)
    err = _apply(request, doc, "DRAFT", "WITHDRAW_AUTHORISATION",
                 roles={"FINANCE"}, comment=comment)
    if err is None:
        reverse_pr_authorisation(doc, request.user)
    return err


def _do_return(request, doc, comment):
    """Return with comment → back to Draft (spec §7.2 / §7.5a)."""
    if not comment.strip():
        return Response({"detail": "A comment is required to return."}, status=400)
    if doc.doc_type in ("MR", "IR", "MAR"):
        return _apply(request, doc, "DRAFT", "RETURN", pm_gate=True,
                      comment=comment)
    if doc.doc_type == "PMR":
        # Whoever holds it can send it back for rework: PM at Submitted,
        # HO staff at PM-Approved, Director at HO-Reviewed/Sized.
        roles = {"SUBMITTED": {"PM"},
                 "PM_APPROVED": {"HO_PURCHASING"},
                 "HO_REVIEWED": {"DIRECTOR"},
                 "SIZED_RELEASED": {"DIRECTOR"}}.get(doc.status, set())
        return _apply(request, doc, "DRAFT", "RETURN", roles=roles,
                      comment=comment)
    if doc.doc_type == "PR":
        # Director returns before authorisation; signatory/Finance return
        # an approved PR (no commitment yet, so nothing to reverse)
        roles = {"DIRECTOR", "SIGNATORY", "FINANCE"} if doc.status == \
            "APPROVED" else {"DIRECTOR"}
        return _apply(request, doc, "DRAFT", "RETURN", roles=roles,
                      comment=comment)
    if doc.doc_type == "IPR":
        # Director returns a submitted order for rework (no commitment yet)
        return _apply(request, doc, "DRAFT", "RETURN",
                      roles={"DIRECTOR", "HO_PURCHASING"}, comment=comment)
    return Response({"detail": "Return applies to MR/PR/IR/MAR/IPR."},
                    status=400)


def _set_workflow_payload(revision, key, data):
    """Server-controlled, write-once workflow blocks (client result, Part C).
    The contractor's issued content stays immutable; these record the client's
    outcome, which arrives after issue by design (decision 1)."""
    payload = revision.payload or {}
    payload[key] = data
    revision.payload = payload
    revision.save(update_fields=["payload"])


def _do_record_result(request, doc, comment):
    """Site team records the client's result in-app (decision 1)."""
    if doc.doc_type not in RESULTS:
        return Response({"detail": "record-result applies to IR/MAR."}, status=400)
    if request.user.role not in SITE_TEAM and request.user.role != "ADMIN":
        return Response({"detail": "Site team records client results."}, status=403)
    result = (request.data.get("result") or "").upper().replace(" ", "_")
    if result not in RESULTS[doc.doc_type]:
        return Response(
            {"detail": f"result must be one of {sorted(RESULTS[doc.doc_type])}."},
            status=400,
        )
    if (doc.current_revision.payload or {}).get("client_result"):
        return Response({"detail": "A result is already recorded for this "
                                   "revision."}, status=400)
    err = _apply(request, doc, result, "RESULT_RECORDED",
                 roles=SITE_TEAM, comment=comment)
    if err is not None:
        return err
    block = {
        "result": result,
        "comments": comment,
        "recorded_by_user": request.user.full_name,
        "reviewed_by": request.data.get("reviewed_by", ""),
        "position": request.data.get("position", ""),
        "inspection_date": request.data.get("inspection_date", ""),
    }
    if doc.doc_type == "MAR" and result in ("APPROVED",
                                            "APPROVED_WITH_COMMENTS"):
        block["approval_date"] = date.today().isoformat()  # spec §5.4
    _set_workflow_payload(doc.current_revision, "client_result", block)
    Approval.objects.filter(document=doc, revision=doc.current_revision,
                            action="RESULT_RECORDED").update(result=result)
    generate_pdf(doc, doc.current_revision, "result")
    return None


def _do_client_verify(request, doc, comment):
    """IR Part C: client verifies the comment closure (recorded in-app)."""
    if doc.doc_type != "IR":
        return Response({"detail": "client-verify applies to IR."}, status=400)
    if request.user.role not in SITE_TEAM and request.user.role != "ADMIN":
        return Response({"detail": "Site team records client verification."},
                        status=403)
    err = _apply(request, doc, "CLOSED", "CLIENT_VERIFIED",
                 roles=SITE_TEAM, comment=comment)
    if err is not None:
        return err
    payload = doc.current_revision.payload or {}
    closure = payload.get("closure", {})
    closure["verified_by"] = request.data.get("verified_by", "")
    closure["verified_date"] = date.today().isoformat()
    _set_workflow_payload(doc.current_revision, "closure", closure)
    generate_pdf(doc, doc.current_revision, "closed")
    return None


def _do_acknowledge(request, doc, comment):
    """TWS client-rep acknowledgement, recorded in-app (decision 1)."""
    if doc.doc_type != "TWS":
        return Response({"detail": "acknowledge applies to TWS."}, status=400)
    if request.user.role not in SITE_TEAM and request.user.role != "ADMIN":
        return Response({"detail": "Site team records acknowledgement."},
                        status=403)
    err = _apply(request, doc, "ACKNOWLEDGED", "ACKNOWLEDGE",
                 roles=SITE_TEAM, comment=comment)
    if err is not None:
        return err
    _set_workflow_payload(doc.current_revision, "acknowledgement", {
        "acknowledged_by": request.data.get("acknowledged_by", ""),
        "date": date.today().isoformat(),
    })
    return None


def _do_send(request, doc, comment):
    """MR: issue to HO Purchasing after PM approval (spec §5.5 r7)."""
    if doc.doc_type != "MR":
        return Response({"detail": "Send applies to MR."}, status=400)
    return _apply(request, doc, "SENT_TO_HO", "SEND",
                  roles={"SITE_ADMIN", "SITE_ENGINEER", "PM"},
                  lock_revision=True, pdf_milestone="sent", comment=comment)


def _do_depart(request, doc, comment):
    """LM departure = crew countersign moment; locks the manifest, creates
    pending items, updates MR statuses (spec §5.8)."""
    if doc.doc_type != "LM":
        return Response({"detail": "Depart applies to LM."}, status=400)
    err = _apply(request, doc, "DEPARTED", "DEPART",
                 roles={"HO_PURCHASING"}, lock_revision=True,
                 pdf_milestone="departed", comment=comment)
    if err is None:
        on_lm_departed(doc, request.user)
    return err


def _do_count(request, doc, comment):
    if doc.doc_type != "GRN":
        return Response({"detail": "Count applies to GRN."}, status=400)
    return _apply(request, doc, "COUNTED", "COUNT",
                  roles={"SITE_ADMIN", "SITE_ENGINEER"}, comment=comment)


def _do_verify(request, doc, comment):
    if doc.doc_type == "DPR":
        return _apply(request, doc, "VERIFIED", "VERIFY", pm_gate=True,
                      comment=comment)
    if doc.doc_type == "GRN":  # SE/PM verify; immutable afterwards (spec §5.6)
        if not _can(request, "GRN", {"SITE_ENGINEER", "PM"}):
            return Response({"detail": "Only SE/PM can verify a GRN."}, status=403)
        shortage = any(
            (line.qty_received or 0) < (line.qty_manifest or 0)
            for line in doc.current_revision.lines.all()
        )
        new_status = "SHORTAGE_REPORTED" if shortage else "COMPLETE"
        err = _apply(request, doc, new_status, "VERIFY",
                     roles={"SITE_ENGINEER", "PM"}, lock_revision=True,
                     pdf_milestone="verified", comment=comment)
        if err is None:
            on_grn_verified(doc, request.user)
            if shortage:
                audit("document", doc.id, "GRN_SHORTAGE_REPORTED",
                      actor=request.user,
                      detail={"ref": doc.ref,
                              "reported_at": timezone.now().isoformat()})
        return err
    return Response({"detail": "Verify applies to DPR/GRN."}, status=400)


# PR payments are recorded per vendor row via POST /pr/{ref}/vendor-payment
# (views_quotes.pr_vendor_payment, R3 addendum) — no PR-level action.


def _do_close(request, doc, comment):
    if doc.doc_type == "IR":
        # Part C: corrective action taken, closed by the PM (spec §5.3)
        if not (_is_pm_for(request.user, doc)
                or request.user.role == "ADMIN"):
            return Response({"detail": "Only the project's PM closes Part C."},
                            status=403)
        if not comment.strip():
            return Response({"detail": "Describe the corrective action taken."},
                            status=400)
        err = _apply(request, doc, "CLOSED_BY_PM", "CLOSE", pm_gate=True,
                     comment=comment)
        if err is not None:
            return err
        _set_workflow_payload(doc.current_revision, "closure", {
            "corrective_action": comment,
            "closed_by_pm": request.user.full_name,
            "closed_date": date.today().isoformat(),
        })
        return None
    roles = {"MR": {"SITE_ADMIN", "SITE_ENGINEER", "PM", "HO_PURCHASING"},
             "PR": {"HO_PURCHASING"}, "PO": {"HO_PURCHASING"}}.get(doc.doc_type)
    if roles is None:
        return Response({"detail": "Close applies to MR/PR/PO/IR."}, status=400)
    return _apply(request, doc, "CLOSED", "CLOSE", roles=roles, comment=comment)


def _do_ho_review(request, doc, comment):
    """PMR: HO project staff review a PM-approved requirement (§5.10.3)."""
    if doc.doc_type != "PMR":
        return Response({"detail": "ho-review applies to PMR."}, status=400)
    return _apply(request, doc, "HO_REVIEWED", "HO_REVIEW",
                  roles={"HO_PURCHASING"}, comment=comment)


def _do_size_release(request, doc, comment):
    """PMR: the Director sizes the requirement (order qty may exceed what the
    project asked for, MOQ) and releases it to sourcing (§5.10.3). The per-line
    order sizing + surplus-to-general-stock happens on the IPR (later slice);
    here we record the Director's sizing note and advance the thread."""
    if doc.doc_type != "PMR":
        return Response({"detail": "size-release applies to PMR."}, status=400)
    err = _apply(request, doc, "SIZED_RELEASED", "SIZE_RELEASE",
                 roles={"DIRECTOR"}, comment=comment)
    if err is None and comment.strip():
        _set_workflow_payload(doc.current_revision, "sizing", {
            "note": comment, "by": request.user.full_name,
            "date": date.today().isoformat(),
        })
    return err


def _do_cancel(request, doc, comment):
    """Cancellation from an early state (no longer needed). PMR: the raising
    team / project PM before ordering. IPR: HO Purchasing before authorisation."""
    if not comment.strip():
        return Response({"detail": "A reason is required to cancel."}, status=400)
    if doc.doc_type == "PMR":
        return _apply(request, doc, "CANCELLED", "CANCEL",
                      roles={"SITE_ENGINEER", "SITE_ADMIN", "PM"},
                      comment=comment)
    if doc.doc_type == "IPR":
        return _apply(request, doc, "CANCELLED", "CANCEL",
                      roles={"HO_PURCHASING"}, comment=comment)
    return Response({"detail": "cancel applies to PMR/IPR here."}, status=400)


def _do_void(request, doc, comment):
    """Void + reissue path (spec §7.2): reason required, number kept."""
    if not (request.user.role == "ADMIN" or _is_site_pm(request.user, doc.site)
            or (doc.doc_type in ("PR", "LM")
                and request.user.role == "HO_PURCHASING")):
        return Response({"detail": "Only PM/Admin (or HO Purchasing for PR/LM) "
                                   "can void."}, status=403)
    reason = (request.data.get("reason") or "").strip()
    if not reason:
        return Response({"detail": "A reason is required to void."}, status=400)
    doc.is_void = True
    doc.void_reason = reason
    doc.voided_by = request.user
    doc.voided_at = timezone.now()
    doc.save(update_fields=["is_void", "void_reason", "voided_by", "voided_at"])
    _record(doc, "VOID", request, comment=reason)
    audit("document", doc.id, "DOC_VOIDED", actor=request.user,
          from_state=doc.status, to_state="VOID", detail={"reason": reason})
    return None


# ===== Attachments =====


@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
def document_attachments(request, ref):
    doc, err = _get_scoped_document(request, ref)
    if err:
        return err
    if doc.is_void or doc.status != "DRAFT":
        return Response({"detail": "Attachments only on drafts."}, status=400)
    upload = request.FILES.get("file")
    if upload is None:
        return Response({"detail": "file is required."}, status=400)
    line = None
    line_id = request.data.get("line_id")
    if line_id:  # a photo for one specific line (free-text MR items)
        line = doc.current_revision.lines.filter(pk=line_id).first()
        if line is None:
            return Response({"detail": "Unknown line for this document."},
                            status=400)
    attachment = Attachment.objects.create(
        document=doc, revision=doc.current_revision, line=line,
        kind=request.data.get("kind", "PHOTO"),
        file=upload, file_name=upload.name,
        content_type=upload.content_type or "",
        size_bytes=upload.size, caption=request.data.get("caption", ""),
        uploaded_by=request.user,
    )
    return Response(AttachmentSerializer(attachment,
                                         context={"request": request}).data,
                    status=201)


# ===== Lists, registers, dashboards =====


# "Open" = still actionable in the chain (for reference pick-lists)
OPEN_STATUSES = {
    "MR": ["SENT_TO_HO", "PR_RAISED", "LOADING_PLANNED", "PARTIALLY_LOADED"],
    "PR": ["SUBMITTED", "APPROVED", "PAYMENT_PROCESSING", "PAID_PO_ISSUED"],
    "PO": ["DRAFT", "ISSUED"],
    "LM": ["DRAFT", "LOADING", "DEPARTED"],
}


@api_view(["GET"])
def documents_list(request):
    qs = Document.objects.select_related("site").order_by("-doc_date", "-id")
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None:
        qs = qs.filter(site_id__in=site_ids)
    if request.GET.get("site"):
        qs = qs.filter(site_id=request.GET["site"])
    if request.GET.get("doc_type"):
        qs = qs.filter(doc_type=request.GET["doc_type"])
    if request.GET.get("project"):
        qs = qs.filter(project_id=request.GET["project"])
    if request.GET.get("status"):
        qs = qs.filter(status=request.GET["status"])
    if request.GET.get("open"):
        statuses = OPEN_STATUSES.get(request.GET.get("doc_type", ""), None)
        if statuses:
            qs = qs.filter(status__in=statuses, is_void=False)
    if request.GET.get("for_pr"):
        # MRs ready to raise a PR against: reached HO (SENT_TO_HO) and NOT
        # already covered by an active PR — a draft PR links MR_PR the
        # moment it is saved, so an MR with an ongoing PR drops out here.
        from .models import DocumentLink

        covered = DocumentLink.objects.filter(
            link_type="MR_PR", from_document__doc_type="PR",
            from_document__is_void=False,
        ).exclude(
            from_document__status__in=["CANCELLED", "REJECTED"]
        ).values_list("to_document_id", flat=True)
        qs = qs.filter(doc_type="MR", status="SENT_TO_HO",
                       is_void=False).exclude(id__in=covered)
    return Response(
        DocumentSerializer(qs[:200], many=True,
                           context={"request": request}).data
    )


@api_view(["GET"])
def mr_export(request, ref):
    """Export an MR's line items to Excel so Purchasing can source them
    (owner, 2026-07-13). Any role that can see the MR may export it."""
    doc, err = _get_scoped_document(request, ref)
    if err:
        return err
    if doc.doc_type != "MR":
        return Response({"detail": "Export applies to MR."}, status=400)
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rev = doc.current_revision
    payload = rev.payload or {}
    wb = Workbook()
    ws = wb.active
    ws.title = doc.ref[:31]
    bold = Font(bold=True)
    ws.append([f"Material Requisition {doc.ref}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Site", doc.site.name, "", "Status",
               doc.status.replace("_", " ").title()])
    ws.append(["Trades / section", payload.get("trades_covered", ""), "",
               "Required by", payload.get("required_by", "")])
    ws.append(["Planned loading", payload.get("planned_loading", "")])
    ws.append([])
    headers = ["#", "Item Code", "Description", "Unit", "Required", "In Stock",
               "To Order", "Priority", "Urgent reason", "Remarks"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = bold
    for i, ln in enumerate(rev.lines.select_related("item").all(), 1):
        ws.append([
            i, ln.item.code if ln.item_id else "",
            ln.description, ln.unit,
            float(ln.qty_required or 0), float(ln.qty_stock or 0),
            float(ln.qty_to_order or 0), ln.priority or "",
            ln.urgent_reason or "", ln.remarks or "",
        ])
    widths = [4, 12, 40, 8, 10, 10, 10, 10, 24, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col)
                             .column_letter].width = w
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument"
                     ".spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{doc.ref}.xlsx"'
    wb.save(response)
    return response


@api_view(["GET"])
def register_generic(request, doc_type):
    """Registers are views over document data (spec §6) — one row per
    document, showing its current revision (older revisions are history you
    see by opening the document, not separate register rows)."""
    doc_type = doc_type.upper()
    if doc_type not in Document.TRANSITIONS:
        return Response({"detail": "Unknown register."}, status=404)
    qs = Document.objects.filter(doc_type=doc_type).select_related(
        "site", "current_revision", "created_by"
    ).prefetch_related("revisions").order_by("-id")
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None:
        qs = qs.filter(site_id__in=site_ids)
    if request.GET.get("site"):
        qs = qs.filter(site_id=request.GET["site"])
    if request.GET.get("status"):
        qs = qs.filter(status=request.GET["status"])

    rows = []
    for doc in qs[:300]:
        # the current revision (prefer the is_current flag; fall back to the FK)
        current = next((r for r in doc.revisions.all() if r.is_current),
                       doc.current_revision)
        revisions = [current]
        links = {}
        for link in doc.links_from.select_related("to_document"):
            links.setdefault(link.link_type, []).append(link.to_document.ref)
        for link in doc.links_to.select_related("from_document"):
            links.setdefault(link.link_type, []).append(link.from_document.ref)
        for revision in revisions:
            payload = (revision.payload if revision else {}) or {}
            rows.append({
                "ref": doc.ref,
                "rev": revision.rev_label if revision else "R0",
                "is_current_rev": bool(revision and revision.is_current),
                "date": doc.doc_date.isoformat(),
                "site_code": doc.site.code,
                "status": "VOID" if doc.is_void else doc.status,
                "created_by": doc.created_by.full_name,
                "links": links,
                "payload_summary": {
                    k: payload.get(k)
                    for k in ("planned_loading", "trades_covered", "required_by",
                              "vessel", "destination", "expected_arrival",
                              "requested_delivery", "action_taken",
                              "manifest_ref", "discipline", "location",
                              "material_description", "manufacturer")
                    if payload.get(k)
                },
                "prev_ir": doc.previous_ir.ref if doc.previous_ir else None,
                "result": (payload.get("client_result") or {}).get("result"),
            })
    return Response({"doc_type": doc_type, "rows": rows})


@api_view(["GET", "PATCH"])
def pending_items(request, pk=None):
    if request.method == "PATCH":
        if request.user.role not in ("HO_PURCHASING", "ADMIN"):
            return Response({"detail": "HO Purchasing edits the pending log."},
                            status=403)
        try:
            row = PendingItem.objects.get(pk=pk)
        except PendingItem.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)
        for field in ("reason", "action_next"):
            if field in request.data:
                setattr(row, field, request.data[field])
        if request.data.get("clear"):  # manual clear with reason (spec §6)
            reason = (request.data.get("cleared_reason") or "").strip()
            if not reason:
                return Response({"detail": "cleared_reason required for manual "
                                           "clear."}, status=400)
            row.status = "CLEARED"
            row.cleared_date = date.today()
            row.cleared_reason = reason
            audit("pending_item", row.id, "PENDING_CLEARED_MANUAL",
                  actor=request.user, detail={"reason": reason})
        row.save()
        return Response(PendingItemSerializer(row).data)

    qs = PendingItem.objects.select_related(
        "site", "item", "pr_document", "cleared_lm",
        "lm_line__revision__document"
    )
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None:
        qs = qs.filter(site_id__in=site_ids)
    if request.GET.get("site"):
        qs = qs.filter(site_id=request.GET["site"])
    if request.GET.get("status"):
        qs = qs.filter(status=request.GET["status"].upper())
    return Response(PendingItemSerializer(qs[:300], many=True).data)


@api_view(["GET"])
def register_dpr_tws(request):
    try:
        site = Site.objects.get(pk=request.GET.get("site"))
    except (Site.DoesNotExist, ValueError, TypeError):
        return Response({"detail": "site parameter required."}, status=400)
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None and site.id not in site_ids:
        return Response({"detail": "Not found."}, status=404)

    today = date.today()
    default_from = today - timedelta(days=13)
    date_from = date.fromisoformat(request.GET.get("from", default_from.isoformat()))
    date_to = min(date.fromisoformat(request.GET.get("to", today.isoformat())), today)

    holidays = set(
        Holiday.objects.filter(site__isnull=True).values_list("day", flat=True)
    ) | set(Holiday.objects.filter(site=site).values_list("day", flat=True))

    # DPR/TWS are site-wide (R8) — the register is one row per working
    # day per SITE; the old per-project filter is ignored.
    doc_qs = Document.objects.filter(
        site=site, doc_type__in=["DPR", "TWS"],
        doc_date__gte=date_from, doc_date__lte=date_to,
    )
    docs = {}
    for d in doc_qs.order_by("id"):
        key = (d.doc_type, d.doc_date)
        # Voided rows remain visible (spec §4.1) but never satisfy the day
        if key not in docs or docs[key].is_void:
            docs[key] = d
    # Gap-flagging: ACTIVE site (spec §2.2), from the earliest active
    # project's start date (dates live on projects — owner), falling back
    # to the site's legacy start date.
    gaps_active = site.status == Site.Status.ACTIVE
    starts = [p.start_date for p in site.projects.filter(status="ACTIVE")
              if p.start_date]
    gap_start = min(starts) if starts else site.start_date
    rows = []
    day = date_from
    while day <= date_to:
        working = day.isoweekday() in (site.working_days or []) and day not in holidays
        if working:
            dpr = docs.get(("DPR", day))
            tws = docs.get(("TWS", day))
            dpr_missing = dpr is None or dpr.is_void
            in_gap_window = (
                gaps_active and gap_start is not None
                and day >= gap_start
            )
            rows.append({
                "date": day.isoformat(),
                "day": day.strftime("%A"),
                "dpr_ref": dpr.ref if dpr else None,
                "dpr_status": "VOID" if dpr and dpr.is_void else
                              (dpr.status if dpr else None),
                "tws_ref": tws.ref if tws and not tws.is_void else None,
                "gap": in_gap_window and dpr_missing and day < today,
                "due_today": day == today and dpr_missing and in_gap_window,
            })
        day += timedelta(days=1)
    return Response({"site": site.code, "rows": rows})


@api_view(["GET"])
def dashboard_site(request, site_id):
    try:
        site = Site.objects.get(pk=site_id)
    except Site.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    site_ids = scoped_site_ids(request.user)
    if site_ids is not None and site.id not in site_ids:
        return Response({"detail": "Not found."}, status=404)

    today = date.today()
    dpr_today = Document.objects.filter(
        doc_type="DPR", site=site, doc_date=today, is_void=False
    ).first()
    unverified = Document.objects.filter(
        doc_type="DPR", site=site, status="ISSUED", is_void=False
    ).count()
    drafts = Document.objects.filter(
        site=site, status="DRAFT", is_void=False
    ).count()
    incoming_lms = Document.objects.filter(
        doc_type="LM", site=site, status="DEPARTED", is_void=False
    ).count()
    # Morning manpower allocation is a daily site obligation too (owner,
    # 2026-07-08) — surface today's DMA next to the DPR/TWS tiles
    dma_today = Document.objects.filter(
        doc_type="DMA", site=site, doc_date=today, is_void=False
    ).first()
    # Materials snapshot: what is on the water (departed manifests) and
    # what HO still owes the site (open pending items) — headline, not
    # the full detail (owner, 2026-07-08)
    in_transit = []
    for lm in Document.objects.filter(
            doc_type="LM", site=site, status="DEPARTED", is_void=False
    ).select_related("current_revision"):
        for line in lm.current_revision.lines.all():
            in_transit.append({
                "description": line.description, "unit": line.unit,
                "qty": line.qty_loaded, "lm_ref": lm.ref,
            })
    pending_qs = PendingItem.objects.filter(
        site=site, status="PENDING").select_related("item")
    pending_materials = [{
        "description": p.item.description if p.item else p.free_text_desc,
        "unit": p.unit, "qty": p.qty_pending,
    } for p in pending_qs[:8]]
    # Manpower today (R9): roster vs attendance, idle = present − workers
    # allocated in today's DMA, plus the DPR-vs-attendance accuracy check
    from .views_hr import site_manpower_data

    mp = site_manpower_data(site, today)
    dma_total = None
    if dma_today and dma_today.current_revision:
        dma_total = 0
        for t in (dma_today.current_revision.payload or {}).get("tasks", []):
            try:
                dma_total += int(t.get("workers") or 0)
            except (TypeError, ValueError):
                pass
    idle = None
    if mp["attendance_entered"] and dma_total is not None:
        idle = max(mp["present"] - dma_total, 0)
    dpr_total = None
    if dpr_today and dpr_today.current_revision:
        counts = (dpr_today.current_revision.payload or {}) \
            .get("manpower", {}) or {}
        try:
            dpr_total = sum(int(v or 0) for v in counts.values())
        except (TypeError, ValueError):
            dpr_total = None
    manpower = {
        "attendance_entered": mp["attendance_entered"],
        "roster_total": mp["roster_total"],
        "present": mp["present"],
        "absent": mp["absent"],
        "allocated": dma_total,
        "idle": idle,
        "dpr_total": dpr_total,
        "dpr_mismatch": (mp["attendance_entered"] and dpr_total is not None
                         and dpr_total != mp["present"]) or False,
        "top": [{"name": c["name"], "roster": c["roster"],
                 "present": c["present"]}
                for c in mp["categories"][:4]],
        "others_roster": sum(c["roster"] for c in mp["categories"][4:]),
    }
    return Response({
        "manpower": manpower,
        "site": site.code,
        "dpr_today": {"ref": dpr_today.ref, "status": dpr_today.status}
        if dpr_today else None,
        "dma_today": {"ref": dma_today.ref, "status": dma_today.status}
        if dma_today else None,
        "unverified_dprs": unverified,
        "open_drafts": drafts,
        "incoming_lms": incoming_lms,
        "materials_in_transit": in_transit[:8],
        "materials_in_transit_count": len(in_transit),
        "pending_materials": pending_materials,
        "pending_materials_count": pending_qs.count(),
    })


@api_view(["GET"])
def dashboard_ho(request):
    if not request.user.is_ho:
        return Response({"detail": "HO roles only."}, status=403)
    base = Document.objects.filter(is_void=False)
    return Response({
        "mrs_awaiting_action": base.filter(doc_type="MR",
                                           status="SENT_TO_HO").count(),
        "prs_awaiting_approval": base.filter(doc_type="PR",
                                             status="SUBMITTED").count(),
        "prs_awaiting_payment": base.filter(
            doc_type="PR", status__in=["APPROVED", "PAYMENT_PROCESSING"]).count(),
        "lms_in_transit": base.filter(doc_type="LM", status="DEPARTED").count(),
        "pending_items_open": PendingItem.objects.filter(status="PENDING").count(),
        "grn_shortages": base.filter(doc_type="GRN",
                                     status="SHORTAGE_REPORTED").count(),
    })


# ===== Prefill conveniences (design §3) =====


@api_view(["GET"])
def mr_lm_prefill(request, ref):
    doc, err = _get_scoped_document(request, ref)
    if err or doc.doc_type != "MR":
        return err or Response({"detail": "Not an MR."}, status=400)
    rows = []
    for line in doc.current_revision.lines.select_related("item"):
        rows.append({
            "item_id": line.item_id,
            "free_text_desc": line.free_text_desc,
            "unit": line.unit,
            "qty_loaded": float(line.qty_to_order or line.qty_required or 0),
            "qty_pending": 0,
            "remarks": line.remarks,
        })
    return Response({"mr_ref": doc.ref, "site_id": doc.site_id,
                     "site_code": doc.site.code, "lines": rows})


@api_view(["GET"])
def mr_related(request, ref):
    """PRs raised against this MR and their POs — for LM reference
    pick-lists (owner UX request)."""
    doc, err = _get_scoped_document(request, ref)
    if err or doc.doc_type != "MR":
        return err or Response({"detail": "Not an MR."}, status=400)
    prs = [link.from_document
           for link in doc.links_to.filter(link_type="MR_PR")
           .select_related("from_document")]
    pos = []
    for pr in prs:
        for link in pr.links_to.filter(link_type="PR_PO") \
                .select_related("from_document__supplier"):
            pos.append(link.from_document)
    return Response({
        "mr_ref": doc.ref,
        "prs": [{"ref": pr.ref, "status": pr.status}
                for pr in prs if not pr.is_void],
        "pos": [{"ref": po.ref, "status": po.status,
                 "supplier": po.supplier.name if po.supplier else ""}
                for po in pos if not po.is_void],
    })


@api_view(["GET"])
def po_lm_prefill(request, ref):
    doc, err = _get_scoped_document(request, ref)
    if err or doc.doc_type != "PO":
        return err or Response({"detail": "Not a PO."}, status=400)
    return Response({"po_ref": doc.ref, "site_id": doc.site_id,
                     "lines": po_lm_prefill_lines(doc)})


@api_view(["GET"])
def lm_grn_prefill(request, ref):
    doc, err = _get_scoped_document(request, ref)
    if err or doc.doc_type != "LM":
        return err or Response({"detail": "Not an LM."}, status=400)
    return Response({"lm_ref": doc.ref, "lines": grn_lines_from_lm(doc),
                     "payload": {"manifest_ref": doc.ref,
                                 "vessel": (doc.current_revision.payload or {})
                                 .get("vessel", "")}})
