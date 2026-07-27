"""Procurement Schedule — client xlsx export (Phase 6).

The client-facing procurement plan the QS shares with the employer, as a
spreadsheet. It reproduces the Soneva/SFR sheet the owner works from, with the
flaws fixed: qty/unit and a "required on site" date added, an overall risk
Status column, and richer stage wording pulled from the live derived pipeline.

The data (and the strict allowlist — no internal money, no supplier name; source
country kept) comes from `procurement_client.client_plan`, shared with the live
HTML share link so the two never diverge.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from .procurement_client import RISK_COLOR, client_plan, initials

# ---- palette (house blue, matching the PDF letterhead) -------------------
_NAVY = "10344F"
_BAND = "1685CC"
_SECTION = "E7F1F9"
_GREY = "F4F9FD"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="C4D0DA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Column plan: (header, sub-header or None, width). A None sub-header means the
# column spans both header rows; the TDS + Progress groups carry sub-headers.
# The 3rd tuple element pairs each column to a client_row key (see _row_values).
_COLUMNS = [
    ("S.No", None, 6),
    ("Category", None, 16),
    ("Description of Item", None, 34),
    ("Make / Brand", None, 16),
    ("Specification", None, 22),
    ("Qty", None, 7),
    ("Unit", None, 8),
    ("Supply By", None, 12),
    ("Source Country", None, 14),
    ("Required On Site", None, 14),
    ("Technical Data Sheet", "Requirement", 12),
    ("Technical Data Sheet", "Status", 12),
    ("Progress", "Order", 12),
    ("Progress", "Production", 12),
    ("Progress", "Shipment", 12),
    ("Progress", "Delivery", 12),
    ("ETA Resort", None, 13),
    ("Status", None, 11),
    ("Remarks", None, 26),
]


def _row_values(r):
    """The client_row dict → the 19 cells, in column order."""
    return [
        r["s_no"], r["category"], r["description"], r["make_brand"],
        r["specification"], r["quantity"], r["uom"], r["supply_by"],
        r["source_country"], r["required_date"], r["tds_req"], r["tds"],
        r["order"], r["production"], r["shipment"], r["delivery"],
        r["eta"], r["status"], r["remarks"],
    ]


def build_client_xlsx(sched, user=None):
    """An openpyxl Workbook of the client procurement plan for one project.
    `user` is the exporter (for the 'updated by' initials); the public,
    token-gated download passes none."""
    plan = client_plan(sched, updated_by=initials(user) if user else "")
    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Plan"
    ncol = len(_COLUMNS)

    # ---- header block (rows 1-2) ----
    # Three label/value pairs per row, each merged into a wide range so long
    # values (project title, contractor) read on a single line.
    no_wrap = Alignment(horizontal="left", vertical="center", wrap_text=False)
    right = Alignment(horizontal="right", vertical="center")

    def hdr(row, lc, lc_end, key, vc, vc_end, val):
        ws.merge_cells(start_row=row, start_column=lc, end_row=row,
                       end_column=lc_end)
        c = ws.cell(row=row, column=lc, value=key)
        c.font = Font(bold=True, color=_NAVY, size=9)
        c.alignment = right
        ws.merge_cells(start_row=row, start_column=vc, end_row=row,
                       end_column=vc_end)
        v = ws.cell(row=row, column=vc, value=val)
        v.font = Font(size=9, color=_NAVY)
        v.alignment = no_wrap

    hdr(1, 1, 2, "Project:",
        3, 7, f"{plan['project_title']} ({plan['project_code']})")
    hdr(1, 8, 9, "Contractor:", 10, 13, plan["contractor"])
    hdr(1, 14, 16, "Sheet Last Update Date:", 17, 19, plan["last_update"])
    hdr(2, 1, 2, "Sub:", 3, 7, "Procurement Plan")
    hdr(2, 8, 9, "Client:", 10, 13, plan["client"])
    hdr(2, 14, 16, "Sheet Updated By:", 17, 19, plan["updated_by"])
    ws.row_dimensions[1].height = 17
    ws.row_dimensions[2].height = 17

    # ---- column headers (rows 3-4) ----
    head_font = Font(bold=True, color=_WHITE, size=9)
    head_fill = PatternFill("solid", fgColor=_BAND)
    r_top, r_bot = 3, 4
    c = 1
    while c <= ncol:
        header, sub, _ = _COLUMNS[c - 1]
        if sub is None:
            ws.merge_cells(start_row=r_top, start_column=c,
                           end_row=r_bot, end_column=c)
            cell = ws.cell(row=r_top, column=c, value=header)
            cell.font, cell.fill, cell.alignment = head_font, head_fill, _CENTER
            cell.border = _BORDER
            ws.cell(row=r_bot, column=c).border = _BORDER
            ws.cell(row=r_bot, column=c).fill = head_fill
            c += 1
        else:
            span = c
            while (span <= ncol and _COLUMNS[span - 1][0] == header
                   and _COLUMNS[span - 1][1] is not None):
                span += 1
            ws.merge_cells(start_row=r_top, start_column=c,
                           end_row=r_top, end_column=span - 1)
            top = ws.cell(row=r_top, column=c, value=header)
            top.font, top.fill, top.alignment = head_font, head_fill, _CENTER
            for cc in range(c, span):
                subcell = ws.cell(row=r_bot, column=cc,
                                  value=_COLUMNS[cc - 1][1])
                subcell.font, subcell.fill = head_font, head_fill
                subcell.alignment, subcell.border = _CENTER, _BORDER
                ws.cell(row=r_top, column=cc).fill = head_fill
                ws.cell(row=r_top, column=cc).border = _BORDER
            c = span
    ws.row_dimensions[r_top].height = 15
    ws.row_dimensions[r_bot].height = 26
    for i, (_, _, width) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ---- section bands + line rows ----
    section_fill = PatternFill("solid", fgColor=_SECTION)
    alt_fill = PatternFill("solid", fgColor=_GREY)
    row = r_bot + 1
    for sec in plan["sections"]:
        bcell = ws.cell(row=row, column=1, value=sec["code"])
        bcell.font = Font(bold=True, color=_NAVY, size=10)
        bcell.fill, bcell.alignment, bcell.border = section_fill, _CENTER, _BORDER
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=ncol)
        tcell = ws.cell(row=row, column=2, value=sec["title"])
        tcell.font = Font(bold=True, color=_NAVY, size=10)
        tcell.fill, tcell.alignment = section_fill, _LEFT
        for cc in range(2, ncol + 1):
            ws.cell(row=row, column=cc).fill = section_fill
            ws.cell(row=row, column=cc).border = _BORDER
        row += 1

        for i, r in enumerate(sec["rows"]):
            zebra = alt_fill if i % 2 else None
            for cidx, val in enumerate(_row_values(r), start=1):
                cell = ws.cell(row=row, column=cidx, value=val)
                cell.border = _BORDER
                cell.font = Font(size=9)
                cell.alignment = (_LEFT_TOP if cidx in (3, 4, 5, 19)
                                  else _CENTER)
                if zebra:
                    cell.fill = zebra
            colour = RISK_COLOR.get(r["status_level"])
            if colour:
                ws.cell(row=row, column=18).font = Font(size=9, bold=True,
                                                        color=colour)
            row += 1

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return wb
